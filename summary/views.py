from decimal import Decimal

from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import cache_page
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
import json

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Sum, F, DecimalField, Q, Prefetch
from django.db.models.functions import Coalesce

# ========== 核心导入 ==========
from accounts.models import User, Role, Permission, ROLE_SUPER_ADMIN, PERM_ORDER_SUMMARY, PERM_PRODUCT_VIEW
from accounts.views import permission_required, create_operation_log, get_client_ip
from bill.models import Order, OrderItem
from product.models import Product
from customer_manage.models import Customer
from area_manage.models import Area, AreaGroup

from operation_log.models import OperationLog
from django.utils import timezone

# ========== 缓存时长常量配置 ==========
CACHE_HIGH_PRIORITY = 300  # 复杂聚合查询 5分钟
CACHE_MID_PRIORITY = 600  # 静态数据 10分钟


# ========== 通用优化函数 ==========
def parse_datetime(date_str):
    """通用时间解析函数 - 返回 上海时区的 aware datetime
    前端传入的本地时间字符串 → 标记为Asia/Shanghai时区 → ORM自动转UTC查询
    """
    try:
        naive_dt = datetime.strptime(date_str.replace('T', ' '), '%Y-%m-%d %H:%M')
        # 统一标记为当前配置时区（Asia/Shanghai），所有接口复用此逻辑
        return timezone.make_aware(naive_dt, timezone.get_current_timezone())
    except ValueError:
        return None


def get_area_ids_by_group(group_id):
    """【优化】极速获取区域ID列表"""
    if group_id == '0':
        return Area.objects.values_list('id', flat=True)
    try:
        return AreaGroup.objects.filter(id=group_id).values_list('areas__id', flat=True)
    except AreaGroup.DoesNotExist:
        return []


# ========== 通用日志 ==========
def create_summary_operation_log(request, operation_type, object_type, object_id=None, object_name=None,
                                 operation_detail=None):
    create_operation_log(
        request=request, op_type=operation_type, obj_type=object_type,
        obj_id=object_id, obj_name=object_name, detail=operation_detail
    )


# ========== 核心业务视图 ==========
@login_required
@permission_required(PERM_ORDER_SUMMARY)
def summary_page(request):
    """商品汇总页面"""
    return render(request, 'summary/summary.html')



from django.db.models import Case, When, Value, IntegerField, CharField, DecimalField, F, Sum, Q
from django.db.models.functions import Coalesce

@login_required
@permission_required(PERM_ORDER_SUMMARY)
@cache_page(CACHE_HIGH_PRIORITY)
def summary_by_group(request):
    group_id = request.GET.get('group_id')
    start_datetime = request.GET.get('start_date')
    end_datetime = request.GET.get('end_date')
    tag_ids_str = request.GET.get('tag_ids', '')
    creator_id = request.GET.get('creator_id')

    if not all([group_id, start_datetime, end_datetime]):
        return JsonResponse({'code': 0, 'msg': '请选择组和时间范围'})

    area_ids = get_area_ids_by_group(group_id)
    group_name = '全部区域' if group_id == '0' else AreaGroup.objects.get(id=group_id).name

    start = parse_datetime(start_datetime)
    end = parse_datetime(end_datetime)
    if not start or not end:
        return JsonResponse({'code': 0, 'msg': '时间格式错误'})

    filters = {
        'order__area_id__in': area_ids,
        'order__create_time__gte': start,
        'order__create_time__lte': end,
        'order__status__in': ['pending', 'printed', 'reopened']
    }

    if tag_ids_str:
        tag_ids = [int(x) for x in tag_ids_str.split(',') if x]
        if tag_ids:
            filters['product__tags__id__in'] = tag_ids

    if creator_id:
        try:
            creator_id = int(creator_id)
            filters['order__creator_id'] = creator_id
        except (ValueError, TypeError):
            pass

    # 使用带前缀的注释名，避免与模型字段（包括外键的 `_id` 字段）冲突
    items = OrderItem.objects.filter(**filters) \
        .select_related('product') \
        .annotate(
            agg_product_id=Case(
                When(product__isnull=False, then=F('product__id')),
                default=Value(-1),
                output_field=IntegerField()
            ),
            agg_product_name=Case(
                When(product__isnull=False, then=F('product__name')),
                default=F('product_name'),
                output_field=CharField()
            ),
            agg_product_unit=Case(
                When(product__isnull=False, then=F('product__unit')),
                default=F('unit'),
                output_field=CharField()
            ),
            agg_product_price=Case(
                When(product__isnull=False, then=F('product__price')),
                default=Coalesce('actual_unit_price', Value(0, output_field=DecimalField())),
                output_field=DecimalField(max_digits=10, decimal_places=2)
            ),
            agg_product_spec=Case(
                When(product__isnull=False, then=F('product__specification')),
                default=F('specification'),
                output_field=CharField()
            )
        ) \
        .values('agg_product_id', 'agg_product_name', 'agg_product_unit', 'agg_product_price', 'agg_product_spec') \
        .annotate(
            total_qty=Sum('quantity'),
            total_amt=Sum('amount')
        ) \
        .order_by('-total_qty')

    # 批量获取正式商品的标签
    product_ids = [item['agg_product_id'] for item in items if item['agg_product_id'] != -1]
    tags_map = {}
    if product_ids:
        products = Product.objects.filter(id__in=product_ids).prefetch_related('tags')
        tags_map = {p.id: list(p.tags.filter(is_active=True).values_list('id', flat=True)) for p in products}

    data = []
    total_amount = Decimal('0.00')

    for idx, item in enumerate(items, 1):
        item_total_amt = item['total_amt'] or Decimal('0.00')
        total_amount += item_total_amt
        pid = item['agg_product_id']
        tags = tags_map.get(pid, []) if pid != -1 else []
        data.append({
            'serial': idx,
            'pid': pid,                     # 前端据此判断自由开单（pid=-1）
            'name': item['agg_product_name'],
            'unit': item['agg_product_unit'],
            'price': float(item['agg_product_price']),
            'total_qty': item['total_qty'] or 0,
            'total_amt': float(item_total_amt),
            'specification': item['agg_product_spec'] or '',
            'tags': tags,
            'remark': ''
        })

    time_range_str = f"{start.strftime('%Y-%m-%d %H:%M')}至{end.strftime('%Y-%m-%d %H:%M')}"
    create_summary_operation_log(
        request=request, operation_type='query', object_type='product_summary',
        object_name=f'商品汇总-{group_name}',
        operation_detail=f'查询{group_name} {time_range_str}，返回{len(data)}条数据'
    )

    return JsonResponse({'code': 1, 'data': data, 'total_amount': float(total_amount)})

@login_required
@permission_required(PERM_ORDER_SUMMARY)
def user_list(request):
    """返回所有活跃用户（开单人）列表，用于前端下拉"""
    users = User.objects.filter(is_active=True).order_by('username')
    data = [{'id': u.id, 'username': u.username} for u in users]
    return JsonResponse({'code': 1, 'data': data, 'msg': ''})

@login_required
@permission_required(PERM_ORDER_SUMMARY)
@cache_page(CACHE_MID_PRIORITY)
def group_list(request):
    try:
        groups = AreaGroup.objects.all().order_by('name')
        group_list = [{'id': '0', 'name': '全部区域'}]
        group_list.extend([{'id': str(g.id), 'name': g.name} for g in groups])
        return JsonResponse({'code': 1, 'data': group_list, 'msg': ''})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': f'加载失败：{str(e)}'}, status=400)


@login_required
@permission_required(PERM_ORDER_SUMMARY)
def customer_summary_page(request):
    """客户汇总页面"""
    return render(request, 'summary/customer_summary.html')


@login_required
@permission_required(PERM_ORDER_SUMMARY)
@cache_page(CACHE_HIGH_PRIORITY)
def summary_customer_by_group(request):
    """客户汇总接口 - 匹配 Order 统一合并索引"""
    group_id = request.GET.get('group_id')
    start_datetime = request.GET.get('start_date')
    end_datetime = request.GET.get('end_date')

    if not all([group_id, start_datetime, end_datetime]):
        return JsonResponse({'code': 0, 'msg': '请选择组和时间范围'})

    # 时间校验 (统一使用 parse_datetime，已带上海时区)
    start = parse_datetime(start_datetime)
    end = parse_datetime(end_datetime)
    if not start or not end:
        return JsonResponse({'code': 0, 'msg': '时间格式错误'})

    # 区域处理
    area_ids = get_area_ids_by_group(group_id)

    # 🔥 匹配Order合并索引：status → is_settled → area → customer → create_time
    customer_summary = Order.objects.filter(
        status__in=['pending', 'printed', 'reopened'],
        is_settled=False,
        area_id__in=area_ids,
        customer__isnull=False,
        create_time__gte=start,
        create_time__lte=end
    ).select_related('customer', 'area').values(
        'customer__id', 'customer__name', 'customer__remark'
    ).annotate(
        total_amount=Sum('total_amount')
    ).order_by('-total_amount')

    total_amount = customer_summary.aggregate(
        total=Coalesce(Sum('total_amount'), 0, output_field=DecimalField())
    )['total']

    data = [{
        'serial': idx, 'customer_id': item['customer__id'],
        'customer_name': item['customer__name'], 'total_amount': float(item['total_amount'] or 0),
        'remark': item['customer__remark'] or ''
    } for idx, item in enumerate(customer_summary, 1)]

    create_summary_operation_log(request=request, operation_type='query', object_type='customer_summary')
    return JsonResponse(
        {'code': 1, 'data': data, 'total_amount': float(total_amount), 'msg': '查询成功' if data else '无消费数据'})


# ========== Excel导出（优化索引匹配） ==========
from openpyxl.styles import Border, Side  # 新增导入
from openpyxl.utils import get_column_letter

def export_to_excel(data, title, headers, selected_fields, custom_fields, file_name, total_row=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title

    final_fields = selected_fields.copy()
    final_headers = {field: headers[field] for field in selected_fields}

    if custom_fields:
        for cf in custom_fields:
            cf_name = cf.get('name', '')
            cf_position = cf.get('position', 'after')
            cf_target = cf.get('target', '')
            if not cf_name or not cf_target: continue
            custom_field_key = f'custom_{cf_name.replace(" ", "_")}_{len(final_fields)}'
            final_headers[custom_field_key] = cf_name
            try:
                target_index = final_fields.index(cf_target)
                insert_index = target_index + 1 if cf_position == 'after' else target_index
                final_fields.insert(insert_index, custom_field_key)
            except ValueError:
                final_fields.append(custom_field_key)

    # ========== 新增：交换数量和单位的位置 ==========
    if 'unit' in final_fields and 'total_qty' in final_fields:
        unit_idx = final_fields.index('unit')
        qty_idx = final_fields.index('total_qty')
        # 互换
        final_fields[unit_idx], final_fields[qty_idx] = final_fields[qty_idx], final_fields[unit_idx]

    selected_headers = [final_headers[field] for field in final_fields]
    title_font = Font(bold=True, size=12)
    alignment = Alignment(horizontal='center')

    # ---------- 标题行 ----------
    for col, header in enumerate(selected_headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = title_font
        cell.alignment = alignment

    # ---------- 数据行 ----------
    for row, item in enumerate(data, 2):
        for col, field in enumerate(final_fields, 1):
            value = item.get(field, '') if not field.startswith('custom_') else ''
            if isinstance(value, float):
                value = round(value, 2)
            ws.cell(row=row, column=col, value=value)

    # ---------- 总计行 ----------
    if total_row:
        total_row_num = len(data) + 2
        total_font = Font(bold=True, color="FFFFFF")
        total_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.cell(row=total_row_num, column=1, value="总计").font = total_font
        ws.cell(row=total_row_num, column=1).fill = total_fill
        for col, field in enumerate(final_fields, 1):
            if field in total_row:
                cell = ws.cell(row=total_row_num, column=col, value=round(total_row[field], 2))
                cell.font = total_font
                cell.fill = total_fill
                cell.alignment = Alignment(horizontal='center')

    # ---------- 设置列宽 ----------
    for col in range(1, len(selected_headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # ========== 新增：为所有数据区域添加全边框 ==========
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    max_row = len(data) + 1  # 数据从第2行开始
    if total_row:
        max_row += 1
    max_col = len(selected_headers)
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

    # ---------- 保存并返回 ----------
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{file_name}.xlsx"'
    return response


import zipfile
from io import BytesIO
from django.http import HttpResponse

@login_required
@permission_required(PERM_ORDER_SUMMARY)
def export_product_summary(request):
    """商品导出 - 多区域组时打包为ZIP（每个区域组一个独立Excel）"""
    if request.method != 'POST':
        return JsonResponse({'code': 0, 'msg': '请求方式错误'}, status=405)

    try:
        # ---------- 获取通用参数 ----------
        group_ids_str = request.POST.get('group_ids', '').strip()
        if not group_ids_str:
            return JsonResponse({'code': 0, 'msg': '请选择区域组'})

        start_datetime = request.POST.get('start_date')
        end_datetime = request.POST.get('end_date')
        selected_fields = request.POST.getlist('fields[]')
        custom_fields = json.loads(request.POST.get('custom_fields', '[]'))
        tag_ids_str = request.POST.get('tag_ids', '')
        # sorted_data 本次实现不再使用（多区域组时无法共用），忽略即可

        if not all([start_datetime, end_datetime, selected_fields]):
            return JsonResponse({'code': 0, 'msg': '参数不完整'})

        start = parse_datetime(start_datetime)
        end = parse_datetime(end_datetime)
        if not start or not end:
            return JsonResponse({'code': 0, 'msg': '时间格式错误'})

        # ---------- 解析区域组列表 ----------
        group_ids = [gid for gid in group_ids_str.split(',') if gid]
        if not group_ids:
            return JsonResponse({'code': 0, 'msg': '未选择有效区域组'})

        # 若包含 '0'（全部区域），则只处理 '0'
        if '0' in group_ids:
            group_ids = ['0']

        # 定义表头映射
        headers_map = {
            'serial': '序号', 'name': '商品名称', 'unit': '单位', 'price': '单价',
            'total_qty': '数量', 'total_amt': '总金额', 'remark': '备注'
        }

        # ---------- 辅助函数：生成单个区域组的Excel ----------
        def generate_excel_for_group(gid):
            """返回 (file_name, excel_bytes)"""
            group_name = '全部区域' if gid == '0' else AreaGroup.objects.get(id=gid).name

            # 查询数据
            area_ids = get_area_ids_by_group(gid)
            filters = {
                'product__isnull': False,
                'order__area_id__in': area_ids,
                'order__create_time__gte': start,
                'order__create_time__lte': end,
                'order__status__in': ['pending', 'printed', 'reopened']
            }
            if tag_ids_str:
                tag_ids = [int(x) for x in tag_ids_str.split(',') if x]
                if tag_ids:
                    filters['product__tags__id__in'] = tag_ids

            items = OrderItem.objects.filter(**filters) \
                .select_related('product') \
                .values('product__name', 'product__unit', 'product__price') \
                .annotate(
                    total_qty=Sum('quantity'),
                    total_amt=Sum('amount')
                ) \
                .order_by('-total_qty')

            total_amount = items.aggregate(
                total=Coalesce(Sum('total_amt'), 0, output_field=DecimalField())
            )['total']

            export_data = [{
                'serial': idx,
                'name': item['product__name'],
                'unit': item['product__unit'],
                'price': float(item['product__price']),
                'total_qty': item['total_qty'] or 0,
                'total_amt': float(item['total_amt'] or 0),
                'remark': ''
            } for idx, item in enumerate(items, 1)]

            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '商品汇总'

            # 字段处理
            final_fields = selected_fields.copy()
            final_headers = {field: headers_map[field] for field in selected_fields}

            if custom_fields:
                for cf in custom_fields:
                    cf_name = cf.get('name', '').strip()
                    cf_position = cf.get('position', 'after')
                    cf_target = cf.get('target', '')
                    if not cf_name or not cf_target:
                        continue
                    custom_field_key = f'custom_{cf_name.replace(" ", "_")}_{len(final_fields)}'
                    final_headers[custom_field_key] = cf_name
                    try:
                        target_index = final_fields.index(cf_target)
                        insert_index = target_index + 1 if cf_position == 'after' else target_index
                        final_fields.insert(insert_index, custom_field_key)
                    except ValueError:
                        final_fields.append(custom_field_key)

            # 交换单位与数量
            if 'unit' in final_fields and 'total_qty' in final_fields:
                unit_idx = final_fields.index('unit')
                qty_idx = final_fields.index('total_qty')
                final_fields[unit_idx], final_fields[qty_idx] = final_fields[qty_idx], final_fields[unit_idx]

            # 写入表头
            title_font = Font(bold=True, size=12)
            alignment = Alignment(horizontal='center')
            for col, field_name in enumerate(final_fields, 1):
                cell = ws.cell(row=1, column=col, value=final_headers[field_name])
                cell.font = title_font
                cell.alignment = alignment

            # 写入数据
            for row_idx, item in enumerate(export_data, 2):
                for col_idx, field in enumerate(final_fields, 1):
                    value = item.get(field, '') if not field.startswith('custom_') else ''
                    if isinstance(value, float):
                        value = round(value, 2)
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # 总计行
            if export_data:
                total_row_num = len(export_data) + 2
                total_font = Font(bold=True, color="FFFFFF")
                total_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                ws.cell(row=total_row_num, column=1, value="总计").font = total_font
                ws.cell(row=total_row_num, column=1).fill = total_fill
                if 'total_amt' in final_fields:
                    col_idx = final_fields.index('total_amt') + 1
                    cell = ws.cell(row=total_row_num, column=col_idx, value=round(float(total_amount), 2))
                    cell.font = total_font
                    cell.fill = total_fill
                    cell.alignment = Alignment(horizontal='center')

            # 列宽
            for col in range(1, len(final_fields) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15

            # 边框
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            max_row = len(export_data) + 1
            if export_data:
                max_row += 1
            for row in range(1, max_row + 1):
                for col in range(1, len(final_fields) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border

            # 保存到内存
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            # 文件名：日期_区域组名.xlsx
            file_date_str = timezone.localdate().strftime("%Y%m%d")
            safe_name = group_name.replace('/', '_').replace('\\', '_')[:50]
            file_name = f"{file_date_str}商品汇总_{safe_name}.xlsx"
            return file_name, buffer

        # ---------- 判断数量，决定返回单个还是ZIP ----------
        if len(group_ids) == 1:
            # 单个区域组：直接返回Excel
            file_name, buffer = generate_excel_for_group(group_ids[0])
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{file_name}"'
            # 记录日志
            create_summary_operation_log(
                request=request, operation_type='export', object_type='product_summary',
                object_name=f'导出 {file_name}',
                operation_detail=f'时间范围 {start_datetime} ~ {end_datetime}'
            )
            return response

        else:
            # 多个区域组：打包为ZIP
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for gid in group_ids:
                    file_name, excel_buffer = generate_excel_for_group(gid)
                    zip_file.writestr(file_name, excel_buffer.getvalue())

            zip_buffer.seek(0)
            response = HttpResponse(
                zip_buffer.getvalue(),
                content_type='application/zip'
            )
            file_date_str = timezone.localdate().strftime("%Y%m%d")
            response['Content-Disposition'] = f'attachment; filename="{file_date_str}商品汇总_多区域组.zip"'
            # 记录日志
            create_summary_operation_log(
                request=request, operation_type='export', object_type='product_summary',
                object_name=f'批量导出 {len(group_ids)} 个区域组',
                operation_detail=f'时间范围 {start_datetime} ~ {end_datetime}, 组数: {len(group_ids)}'
            )
            return response

    except Exception as e:
        return JsonResponse({'code': 0, 'msg': f'导出失败：{str(e)}'}, status=500)


@login_required
@permission_required(PERM_ORDER_SUMMARY)
def export_customer_summary(request):
    """客户导出 - 命中Order合并索引"""
    if request.method == 'POST':
        try:
            data = request.POST
            group_id = data.get('group_id')
            start_datetime = data.get('start_date')
            end_datetime = data.get('end_date')
            selected_fields = data.getlist('fields[]')
            custom_fields = json.loads(data.get('custom_fields', '[]'))

            if not all([group_id, start_datetime, end_datetime, selected_fields]):
                return JsonResponse({'code': 0, 'msg': '参数不完整'})

            # 统一使用 parse_datetime，已带上海时区
            start = parse_datetime(start_datetime)
            end = parse_datetime(end_datetime)
            if not start or not end:
                return JsonResponse({'code': 0, 'msg': '时间格式错误'})

            area_ids = get_area_ids_by_group(group_id)

            # 🔥 索引匹配查询
            customer_summary = Order.objects.filter(
                status__in=['pending', 'printed', 'reopened'],
                area_id__in=area_ids,
                customer__isnull=False,
                create_time__gte=start,
                create_time__lte=end
            ).select_related('customer').values(
                'customer__name', 'customer__remark'
            ).annotate(total_amount=Sum('total_amount')).order_by('-total_amount')

            total_amount = \
            customer_summary.aggregate(total=Coalesce(Sum('total_amount'), 0, output_field=DecimalField()))['total']
            export_data = [{
                'serial': idx, 'customer_name': item['customer__name'],
                'total_amount': float(item['total_amount'] or 0), 'remark': item['customer__remark'] or ''
            } for idx, item in enumerate(customer_summary, 1)]

            create_summary_operation_log(request=request, operation_type='export', object_type='customer_summary')

            # 已使用 timezone.localdate()，输出上海本地日期
            file_date_str = timezone.localdate().strftime("%Y%m%d")

            return export_to_excel(
                data=export_data, title='客户汇总', headers={
                    'serial': '序号', 'customer_name': '客户名称',
                    'total_amount': '金额', 'remark': '备注'
                }, selected_fields=selected_fields, custom_fields=custom_fields,
                file_name=f'{file_date_str}{"全部区域" if group_id == "0" else AreaGroup.objects.get(id=group_id).name}',
                total_row={'total_amount': total_amount}
            )
        except Exception as e:
            return JsonResponse({'code': 0, 'msg': f'导出失败：{str(e)}'}, status=500)


@login_required
@permission_required(PERM_PRODUCT_VIEW)
@cache_page(CACHE_MID_PRIORITY)
def product_list_for_price(request):
    """商品列表接口"""
    try:
        products = Product.objects.all().order_by('name')
        return JsonResponse([{'id': p.id, 'name': p.name, 'price': float(p.price)} for p in products], safe=False)
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': f'查询失败：{str(e)}'}, status=500)


@login_required
@permission_required(PERM_PRODUCT_VIEW)
def customer_amount_detail_page(request, customer_id):
    """客户金额详情页"""
    customer = get_object_or_404(Customer, id=customer_id)
    return render(request, 'summary/amount_detail.html', {'customer': customer, 'customer_id': customer_id})


@login_required
@permission_required(PERM_PRODUCT_VIEW)
def get_customer_order_source(request, customer_id):
    """客户订单来源 - 无N+1，匹配Order索引"""
    try:
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 10))

        if not all([start_date, end_date]):
            return JsonResponse({'code': 0, 'msg': '缺少时间参数'}, status=400)

        # parse_datetime 已返回上海时区的aware时间，无需二次转换
        start = parse_datetime(start_date)
        end = parse_datetime(end_date)
        if not start or not end:
            return JsonResponse({'code': 0, 'msg': '时间格式错误'}, status=400)

        # 匹配Order索引查询
        orders = Order.objects.filter(
            customer_id=customer_id,
            status__in=['pending', 'printed', 'reopened'],
            create_time__gte=start,
            create_time__lte=end
        ).select_related('customer', 'area').prefetch_related(
            Prefetch('items', queryset=OrderItem.objects.select_related('product'))
        ).order_by('-create_time')

        total_amount = orders.aggregate(total=Coalesce(Sum('total_amount'), 0, output_field=DecimalField()))['total']
        paginator = Paginator(orders, page_size)
        current_page = paginator.page(page) if page in paginator.page_range else paginator.page(1)

        order_list = [{
            'order_no': order.order_no,
            # 转为上海本地时区后格式化
            'create_time': timezone.localtime(order.create_time).strftime('%Y-%m-%d %H:%M:%S'),
            'total_amount': float(order.total_amount),
            'status': dict(order.ORDER_STATUS).get(order.status, '未知状态'),
            'items': [{
                'product_name': item.product.name, 'quantity': item.quantity,
                'unit': item.product.unit, 'price': float(item.product.price),
                'amount': float(item.amount)
            } for item in order.items.all()]
        } for order in current_page]

        return JsonResponse({
            'code': 1, 'data': order_list, 'total_amount': round(float(total_amount), 2),
            'customer_name': Customer.objects.get(id=customer_id).name,
            'page': page, 'page_size': page_size, 'total_count': paginator.count
        })
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': f'查询失败：{str(e)}'}, status=500)


@login_required
@permission_required(PERM_ORDER_SUMMARY)
def product_summary_detail_page(request, product_id):
    """商品汇总详情页"""
    product = get_object_or_404(Product, id=product_id)
    group_id = request.GET.get('group_id', '0')
    group_name = '全部区域' if group_id == '0' else AreaGroup.objects.filter(id=group_id).first().name

    return render(request, 'summary/product_summary_detail.html', {
        'product': product, 'product_id': product_id, 'group_id': group_id,
        'group_name': group_name, 'start_date': request.GET.get('start_date', ''),
        'end_date': request.GET.get('end_date', '')
    })


@login_required
@permission_required(PERM_ORDER_SUMMARY)
def get_product_order_source(request, product_id):
    """商品订单来源 - 100%命中 OrderItem 统一索引"""
    try:
        group_id = request.GET.get('group_id', '0')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 20))

        if not all([start_date, end_date]):
            return JsonResponse({'code': 0, 'msg': '缺少时间参数'}, status=400)
        # parse_datetime 已返回上海时区的aware时间
        start, end = parse_datetime(start_date), parse_datetime(end_date)
        if not start or not end:
            return JsonResponse({'code': 0, 'msg': '时间格式错误'}, status=400)

        area_ids = get_area_ids_by_group(group_id)
        product = get_object_or_404(Product, id=product_id)

        # 🔥 终极优化：完全匹配 OrderItem 索引 [product, order, quantity, amount]
        order_items = OrderItem.objects.filter(
            product_id=product_id,  # 索引第一位，必传
            order__area_id__in=area_ids,
            order__create_time__gte=start,
            order__create_time__lte=end,
            order__status__in=['pending', 'printed', 'reopened']
        ).select_related('order__customer', 'order__area').order_by('-order__create_time')

        # 索引聚合
        aggregate_data = order_items.aggregate(
            total_quantity=Coalesce(Sum('quantity'), 0),
            total_amount=Coalesce(Sum('amount'), 0, output_field=DecimalField())
        )

        paginator = Paginator(order_items, page_size)
        current_page = paginator.page(page) if page in paginator.page_range else paginator.page(1)

        order_list = [{
            'order_no': item.order.order_no,
            # ✅ 修复：转为上海本地时区后再格式化，解决慢8小时问题
            'create_time': timezone.localtime(item.order.create_time).strftime('%Y-%m-%d %H:%M:%S'),
            'customer_name': item.order.customer.name if item.order.customer else '无客户',
            'area_name': item.order.area.name if item.order.area else '无区域',
            'quantity': item.quantity, 'unit': product.unit,
            'price': float(product.price), 'amount': float(item.amount),
            'order_status': dict(item.order.ORDER_STATUS).get(item.order.status, '未知状态')
        } for item in current_page]

        return JsonResponse({
            'code': 1, 'data': order_list,
            'total_quantity': aggregate_data['total_quantity'],
            'total_amount': round(float(aggregate_data['total_amount']), 2),
            'product_name': product.name, 'page': page, 'page_size': page_size, 'total_count': paginator.count
        })
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': f'查询失败：{str(e)}'}, status=500)

from product.models import ProductTag

@login_required
@permission_required(PERM_ORDER_SUMMARY)
@cache_page(CACHE_MID_PRIORITY)
def tag_list(request):
    """获取所有启用的标签，供前端选择"""
    tags = ProductTag.objects.filter(is_active=True).order_by('sort_order', 'id')
    data = [{'id': t.id, 'name': t.name, 'color': t.color} for t in tags]
    return JsonResponse(data, safe=False)


from openpyxl import Workbook
from openpyxl.styles import Border, Side
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
import json
from django.shortcuts import render


@login_required
def out_car_register(request):
    """出车登记工具页面"""
    route_options = ['稔山线', '遮浪线', '陆丰线', '陆河线', '海城线', '梅陇线', '可塘线']
    driver_options = ['塔', '容', '伦', '达', '远', '武']
    return render(request, 'summary/out_car_register.html', {
        'route_options': route_options,
        'driver_options': driver_options
    })


from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
import json
from django.shortcuts import render
# summary/views.py 中，替换原有的 export_excel 函数
from .models import OutCarRecord
from decimal import Decimal

@login_required
def export_excel(request):
    """导出出车登记数据为 Excel（自动保存记录，包含路线和司机）"""
    if request.method != 'POST':
        return JsonResponse({'code': 0, 'msg': '仅支持POST请求'})

    try:
        data = json.loads(request.body)
        rows = data.get('rows', [])
        date_str = data.get('date', '')
        route = data.get('route', '')
        driver = data.get('driver', '')
        if not rows:
            return JsonResponse({'code': 0, 'msg': '无数据可导出'})
    except:
        return JsonResponse({'code': 0, 'msg': '数据格式错误'})

    # ---------- 辅助函数：提取客户名称（去掉"|"及之前） ----------
    def extract_name(raw):
        if '|' in raw:
            return raw.split('|', 1)[1].strip()
        return raw.strip()

    # ---------- 1. 合并重名，累加金额 ----------
    merged = {}
    order = []
    for row in rows:
        raw_name = row.get('customer_name', '')
        name = extract_name(raw_name)
        amount = row.get('amount', 0)
        if not isinstance(amount, (int, float)):
            amount = 0
        if name not in merged:
            merged[name] = 0
            order.append(name)
        merged[name] += amount

    processed = [{'name': name, 'amount': merged[name]} for name in order]
    total_amount = sum(item['amount'] for item in processed)
    limit = 25

    # ---------- 2. 保存记录到数据库 ----------
    from datetime import datetime
    record_date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else datetime.now().date()
    OutCarRecord.objects.create(
        date=record_date,
        data=processed,
        total_amount=total_amount,
        route=route,
        driver=driver
    )

    # ---------- 3. 生成 Excel ----------
    wb = Workbook()
    ws = wb.active
    ws.title = '出车登记'

    headers = ['序号', '客户名称', '客户金额', '备注']

    # 写入表头（左 A~D，右 F~I）
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    for col_idx, header in enumerate(headers, start=6):
        ws.cell(row=1, column=col_idx, value=header)

    # 写入左侧数据
    left_rows = processed[:limit]
    left_total = 0
    for i, item in enumerate(left_rows, start=1):
        row_num = i + 1
        ws.cell(row=row_num, column=1, value=i)
        ws.cell(row=row_num, column=2, value=item['name'])
        ws.cell(row=row_num, column=3, value=item['amount'])
        ws.cell(row=row_num, column=4, value='')
        left_total += item['amount']

    left_last_row = 1
    if left_rows:
        row_num = len(left_rows) + 2
        ws.cell(row=row_num, column=1, value='总计')
        cell_left_total = ws.cell(row=row_num, column=3, value=left_total)
        cell_left_total.font = Font(color='FF0000')
        left_last_row = row_num

    # 右侧数据
    right_rows = processed[limit:]
    right_total = 0
    right_last_row = 1
    if right_rows:
        for i, item in enumerate(right_rows, start=1):
            row_num = i + 1
            seq = limit + i
            ws.cell(row=row_num, column=6, value=seq)
            ws.cell(row=row_num, column=7, value=item['name'])
            ws.cell(row=row_num, column=8, value=item['amount'])
            ws.cell(row=row_num, column=9, value='')
            right_total += item['amount']

        row_num = len(right_rows) + 2
        ws.cell(row=row_num, column=6, value='总计')
        cell_right_total = ws.cell(row=row_num, column=8, value=right_total)
        cell_right_total.font = Font(color='FF0000')
        right_last_row = row_num

    # 两侧合计
    grand_total = left_total + right_total
    if right_rows:
        grand_row = right_last_row + 2
        ws.cell(row=grand_row, column=6, value='两侧合计')
        cell_grand = ws.cell(row=grand_row, column=8, value=grand_total)
        cell_grand.font = Font(color='FF0000')
        right_last_row = grand_row
    else:
        grand_row = left_last_row + 2
        ws.cell(row=grand_row, column=1, value='两侧合计')
        cell_grand = ws.cell(row=grand_row, column=3, value=grand_total)
        cell_grand.font = Font(color='FF0000')
        left_last_row = grand_row

    bottom_row = max(left_last_row, right_last_row) if right_rows else left_last_row

    # 信息列（K列）：日期、路线（带值）、总金额、退货、司机（带值）、搭档、零钱、实金额、补贴、时间
    date_label = f"日期:{date_str}" if date_str else "日期:"
    route_label = f"路线:{route}" if route else "路线:"
    driver_label = f"司机:{driver}" if driver else "司机:"
    info_labels = [date_label, route_label, '总金额:', '退货:', driver_label, '搭档:', '零钱:200元', '实金额:', '补贴:', '时间:']

    # 计算起始行，使最后一项（时间:）位于 bottom_row
    start_row = bottom_row - (len(info_labels) - 1) * 2
    if start_row < 1:
        start_row = 1
    for idx, label in enumerate(info_labels):
        row = start_row + idx * 2
        ws.cell(row=row, column=11, value=label)

    # 设置列宽
    col_widths = {1:5, 2:12, 3:9, 4:7, 5:2, 6:5, 7:12, 8:9, 9:7, 10:2, 11:16}
    for col, width in col_widths.items():
        ws.column_dimensions[chr(64 + col)].width = width

    # 边框：仅 A~D 和 F~I
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in range(1, bottom_row + 1):
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border
        for col in range(6, 10):
            ws.cell(row=row, column=col).border = thin_border

    # 页面横向
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    # 返回 Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=出车登记.xlsx'
    wb.save(response)
    return response
# summary/views.py 新增
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

@login_required
def out_car_list(request):
    """出车登记历史记录列表"""
    records = OutCarRecord.objects.all().order_by('-date', '-created_at')
    paginator = Paginator(records, 20)  # 每页20条
    page = request.GET.get('page')
    try:
        records_page = paginator.page(page)
    except PageNotAnInteger:
        records_page = paginator.page(1)
    except EmptyPage:
        records_page = paginator.page(paginator.num_pages)

    return render(request, 'summary/out_car_list.html', {'records_page': records_page})