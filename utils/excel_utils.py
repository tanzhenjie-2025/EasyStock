import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def get_column_mapping(headers, expected_map):
    """
    根据表头列表和预期字段映射，返回 {字段名: 列索引} 的字典。
    不会因缺少字段而返回 None，由调用者自行检查关键字段。
    """
    mapping = {}
    if not headers:
        return mapping
    for col_idx, header in enumerate(headers):
        if header is None:
            continue
        header_str = str(header).strip()
        for display_name, field_name in expected_map.items():
            if header_str == display_name:
                mapping[field_name] = col_idx
                break
    return mapping

def export_to_excel_buffer(data, title, headers, selected_fields, custom_fields=None, file_name='导出', total_row=None):
    """
    返回 BytesIO 对象的 Excel 文件
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title

    # 确定最终字段列表（简化，不处理自定义字段位置）
    field_order = selected_fields
    # 若存在自定义字段，简单追加（可根据需求扩展）
    if custom_fields:
        for cf in custom_fields:
            if cf.get('name') not in field_order:
                field_order.append('custom_' + cf['name'])

    # 写入表头
    header_font = Font(bold=True)
    for col_num, field in enumerate(field_order, 1):
        header_text = headers.get(field, field)
        cell = ws.cell(row=1, column=col_num, value=header_text)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # 写入数据
    for row_num, record in enumerate(data, 2):
        for col_num, field in enumerate(field_order, 1):
            value = record.get(field, '')
            ws.cell(row=row_num, column=col_num, value=value)

    # 若有合计行（略）
    if total_row:
        pass

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer