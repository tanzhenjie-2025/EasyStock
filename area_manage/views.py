from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.core.cache import cache
from django.db.models import Q, Count, Prefetch, OuterRef, Subquery, Sum

import logging
import json

from accounts.views import permission_required, create_operation_log
from bill.models import Order, OrderItem
from area_manage.models import Area, AreaGroup
from customer_manage.models import Customer

import openpyxl

from summary.views import export_to_excel
from django.db.models import Sum, Count, Max, Q, F, DecimalField
from django.db.models.functions import Coalesce
# 配置日志
logger = logging.getLogger(__name__)


# ===================== 工具函数（统一优化）=====================
def format_datetime(dt):
    """统一时间格式化 - 先转为上海本地时区再格式化，消除8小时偏差"""
    if not dt:
        return ''
    # 核心修复：转为当前配置的本地时区（Asia/Shanghai）后再格式化
    local_dt = timezone.localtime(dt)
    return local_dt.strftime('%Y-%m-%d %H:%M:%S')


# ===================== 常量定义 =====================
PAGE_SIZE_MAX = 50
AREA_PAGE_SIZE = 15
ALLOW_SORT_AREA = ['name', 'id', 'create_time']
ALLOW_SORT_GROUP = ['name', 'create_time', 'id']

# ===================== 缓存配置 =====================
CACHE_AREA_PAGE = 300
CACHE_GROUP_PAGE = 300
CACHE_DETAIL_PAGE = 300
CACHE_API_LIST = 300
CACHE_API_DETAIL = 300
CACHE_STATISTICS = 600

CACHE_PREFIX = {
    "AREA_LIST": "area_list_",
    "AREA_DETAIL": "area_detail_",
    "GROUP_LIST": "group_list_",
    "GROUP_DETAIL": "group_detail_",
    "AREA_STAT": "area_stats_",
    "GROUP_STAT": "group_stats_",
}

# 🔥 这里必须和 customer_manage/views.py 里定义的完全一致
CACHE_KEY_AREA_LIST_FOR_CUSTOMER = "global:area_list_for_customer"

# ===================== 缓存工具函数 =====================
def generate_cache_key(request, prefix: str, *args) -> str:
    user_id = request.user.id
    params = "_".join([str(arg) for arg in args])
    return f"{prefix}{user_id}_{params}"

def clear_area_cache(area_id: int = None):
    """
    清理区域相关缓存
    """
    if area_id:
        cache.delete(f"{CACHE_PREFIX['AREA_STAT']}{area_id}")
        cache.delete(f"{CACHE_PREFIX['AREA_DETAIL']}{area_id}")

    # 1. 清理区域管理自身的列表缓存
    cache.delete_pattern(f"{CACHE_PREFIX['AREA_LIST']}*")

    # 🔥 2. 关键修改：精准清理客户管理页面的区域列表缓存
    cache.delete(CACHE_KEY_AREA_LIST_FOR_CUSTOMER)
    logger.info(f"已清理区域缓存，包括跨App Key: {CACHE_KEY_AREA_LIST_FOR_CUSTOMER}")

def clear_group_cache(group_id: int = None):
    if group_id:
        cache.delete(f"{CACHE_PREFIX['GROUP_STAT']}{group_id}")
        cache.delete(f"{CACHE_PREFIX['GROUP_DETAIL']}{group_id}")
    cache.delete_pattern(f"{CACHE_PREFIX['GROUP_LIST']}*")


# ===================== 统计函数 =====================
def get_area_statistics(area_id):
    cache_key = f"{CACHE_PREFIX['AREA_STAT']}{area_id}"
    cache_data = cache.get(cache_key)
    if cache_data:
        return cache_data
    try:
        data = {'customer_count': Customer.objects.filter(area_id=area_id).count()}
        cache.set(cache_key, data, CACHE_STATISTICS)
        return data
    except Exception as e:
        logger.error(f"获取区域{area_id}统计数据失败：{str(e)}")
        return {'customer_count': 0}


def get_group_statistics(group_id):
    cache_key = f"{CACHE_PREFIX['GROUP_STAT']}{group_id}"
    cache_data = cache.get(cache_key)
    if cache_data:
        return cache_data
    try:
        # 仅加载必要字段
        group = get_object_or_404(AreaGroup.objects.only('id'), pk=group_id)
        area_ids = group.areas.values_list('id', flat=True)
        customer_count = Customer.objects.filter(area_id__in=area_ids).count()
        data = {'customer_count': customer_count, 'area_count': len(area_ids)}
        cache.set(cache_key, data, CACHE_STATISTICS)
        return data
    except Exception as e:
        logger.error(f"获取区域组{group_id}统计数据失败：{str(e)}")
        return {'customer_count': 0, 'area_count': 0}


# ===================== 区域管理 CRUD =====================

@login_required
@permission_required('area_view')
def area_list(request):
    try:
        keyword = request.GET.get('keyword', '').strip()
        sort_by = request.GET.get('sort', 'name')
        sort_order = request.GET.get('order', 'asc')
        page = max(int(request.GET.get('page', 1)), 1)
        status = request.GET.get('status', 'all').strip()  # 新增：状态筛选参数

        if sort_by not in ALLOW_SORT_AREA:
            sort_by = 'name'

        cache_key = generate_cache_key(request, CACHE_PREFIX['AREA_LIST'], keyword, sort_by, sort_order, page, status)
        cache_data = cache.get(cache_key)
        if cache_data:
            return JsonResponse(cache_data)

        # 🔧 优化：仅加载必要字段，包含is_active
        base_areas = Area.objects.only('id', 'name', 'remark', 'create_time', 'is_active')
        if keyword:
            base_areas = base_areas.filter(Q(name__icontains=keyword) | Q(remark__icontains=keyword))

        # 根据状态过滤
        areas = base_areas
        if status == 'active':
            areas = areas.filter(is_active=True)
        elif status == 'inactive':
            areas = areas.filter(is_active=False)

        areas = areas.order_by(f'-{sort_by}' if sort_order == 'desc' else sort_by)
        total = areas.count()

        # 计算各状态数量
        all_count = base_areas.count()
        active_count = base_areas.filter(is_active=True).count()
        inactive_count = all_count - active_count

        start = (page - 1) * AREA_PAGE_SIZE
        end = start + AREA_PAGE_SIZE
        areas_page = areas[start:end]

        area_ids = [a.id for a in areas_page]
        customer_count_map = dict(
            Customer.objects.filter(area_id__in=area_ids)
            .values('area_id')
            .annotate(count=Count('id'))
            .values_list('area_id', 'count')
        )

        # 🔧 时间格式化已通过 format_datetime 统一转本地时区
        result = [
            {
                'id': a.id,
                'name': a.name,
                'remark': a.remark or '',
                'customer_count': customer_count_map.get(a.id, 0),
                'create_time': format_datetime(a.create_time),
                'is_active': a.is_active
            }
            for a in areas_page
        ]
        response_data = {
            'code': 1, 'data': result,
            'pagination': {
                'total': total, 'page': page, 'page_size': AREA_PAGE_SIZE,
                'total_pages': (total + AREA_PAGE_SIZE - 1) // AREA_PAGE_SIZE
            },
            'counts': {
                'all': all_count,
                'active': active_count,
                'inactive': inactive_count
            }
        }
        cache.set(cache_key, response_data, CACHE_API_LIST)
        return JsonResponse(response_data)

    except Exception as e:
        logger.error(f"查询区域列表失败：{str(e)}", exc_info=True)
        return JsonResponse({'code': 0, 'msg': f'查询失败：{str(e)}'})


@login_required
@permission_required('area_view')
def area_detail_api(request, pk):
    try:
        # 1. 基础区域信息
        area = get_object_or_404(
            Area.objects.only('id', 'name', 'remark', 'create_time', 'update_time'),
            pk=pk
        )
        related_groups = area.areagroup_set.values('id', 'name')

        # 2. 获取分页参数
        page = max(int(request.GET.get('c_page', 1)), 1)
        page_size = 15

        # 3. 查询客户 (分页)
        customers_qs = Customer.objects.filter(area_id=pk).only('id', 'name', 'phone').order_by('-create_time')
        total_customers = customers_qs.count()

        start = (page - 1) * page_size
        end = start + page_size
        customers_page = customers_qs[start:end]

        customers = [{'id': c.id, 'name': c.name, 'phone': c.phone} for c in customers_page]

        # 4. 组装数据 - 时间通过 format_datetime 统一转本地
        data = {
            'id': area.id,
            'name': area.name,
            'remark': area.remark or '',
            'create_time': format_datetime(area.create_time),
            'update_time': format_datetime(area.update_time),
            'customer_count': total_customers,
            'customers': customers,
            'customer_pagination': {
                'page': page,
                'total_pages': (total_customers + page_size - 1) // page_size,
                'total': total_customers
            },
            'related_groups': list(related_groups)
        }
        return JsonResponse({'code': 1, 'data': data})

    except Exception as e:
        logger.error(f"查询区域{pk}详情失败：{str(e)}")
        return JsonResponse({'code': 0, 'msg': f'查询失败：{str(e)}'})


@login_required
@permission_required('area_add')
def area_add(request):
    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            remark = request.POST.get('remark', '').strip()
            if not name:
                return JsonResponse({'code': 0, 'msg': '区域名不能为空'})
            if Area.objects.filter(name=name).exists():
                return JsonResponse({'code': 0, 'msg': '区域已存在'})

            area = Area.objects.create(name=name, remark=remark)
            create_operation_log(request=request, op_type='create', obj_type='area',
                                 obj_id=area.id, obj_name=area.name, detail=f"新增区域：{area.name}")

            # 调用增强版的缓存清理
            clear_area_cache()

            return JsonResponse({'code': 1, 'msg': '添加成功'})
        except Exception as e:
            logger.error(f"新增区域失败：{str(e)}")
            return JsonResponse({'code': 0, 'msg': f'新增失败：{str(e)}'})
    return JsonResponse({'code': 0, 'msg': '仅支持POST请求'})


@login_required
@permission_required('area_edit')
def area_edit(request, pk):
    try:
        # 🔧 优化：only()限制字段
        area = get_object_or_404(Area.objects.only('id', 'name', 'remark'), pk=pk)
        if request.method == 'POST':
            name = request.POST.get('name', '').strip()
            remark = request.POST.get('remark', '').strip()
            if not name:
                return JsonResponse({'code': 0, 'msg': '区域名不能为空'})
            if Area.objects.filter(name=name).exclude(pk=pk).exists():
                return JsonResponse({'code': 0, 'msg': '区域名重复'})

            area.name = name
            area.remark = remark
            area.save()
            create_operation_log(request=request, op_type='update', obj_type='area',
                                 obj_id=area.id, obj_name=area.name, detail=f"编辑区域")

            # 调用增强版的缓存清理
            clear_area_cache(area_id=pk)

            return JsonResponse({'code': 1, 'msg': '修改成功'})
        return JsonResponse({'code': 0, 'msg': '仅支持POST请求'})
    except Exception as e:
        logger.error(f"编辑区域失败：{str(e)}")
        return JsonResponse({'code': 0, 'msg': f'编辑失败：{str(e)}'})


@login_required
@permission_required('area_delete')
def area_delete(request, pk):
    try:
        area = get_object_or_404(Area.objects.only('id', 'name'), pk=pk)
        # 🔥 修改为软删除
        area.is_active = False
        area.save()

        create_operation_log(request=request, op_type='disable', obj_type='area',
                             obj_id=pk, obj_name=area.name, detail=f"禁用区域")

        clear_area_cache(area_id=pk)
        return JsonResponse({'code': 1, 'msg': '禁用成功'})
    except Exception as e:
        logger.error(f"禁用区域失败：{str(e)}")
        return JsonResponse({'code': 0, 'msg': f'禁用失败：{str(e)}'})


# 新增：区域启用功能
@login_required
@permission_required('area_edit')
def area_enable(request, pk):
    try:
        area = get_object_or_404(Area.objects.only('id', 'name'), pk=pk)
        area.is_active = True
        area.save()

        create_operation_log(request=request, op_type='enable', obj_type='area',
                             obj_id=pk, obj_name=area.name, detail=f"启用区域")

        clear_area_cache(area_id=pk)
        return JsonResponse({'code': 1, 'msg': '启用成功'})
    except Exception as e:
        logger.error(f"启用区域失败：{str(e)}")
        return JsonResponse({'code': 0, 'msg': f'启用失败：{str(e)}'})


# ===================== 区域组管理 =====================
@login_required
@permission_required('area_view')
def group_list(request):
    try:
        keyword = request.GET.get('keyword', '').strip()
        sort_by = request.GET.get('sort', 'name')
        sort_order = request.GET.get('order', 'asc')
        page = max(int(request.GET.get('page', 1)), 1)
        page_size = min(int(request.GET.get('page_size', 20)), PAGE_SIZE_MAX)
        status = request.GET.get('status', 'all').strip()

        if sort_by not in ALLOW_SORT_GROUP:
            sort_by = 'name'

        cache_key = generate_cache_key(request, CACHE_PREFIX['GROUP_LIST'], keyword, sort_by, sort_order, page,
                                       page_size, status)
        cache_data = cache.get(cache_key)
        if cache_data:
            return JsonResponse(cache_data)

        # 🔧 优化：简化 Subquery，直接对 area_id 计数
        customer_subquery = Customer.objects.filter(
            area_id=OuterRef('areas__id')
        ).annotate(count=Count('id')).values('count')

        base_groups = AreaGroup.objects.only('id', 'name', 'remark', 'create_time', 'update_time', 'is_active')
        base_groups = base_groups.prefetch_related(
            Prefetch('areas', queryset=Area.objects.only('id', 'name'))
        ).annotate(
            customer_count=Coalesce(Sum(Subquery(customer_subquery)), 0),
            area_count=Count('areas', distinct=True)
        )

        if keyword:
            base_groups = base_groups.filter(
                Q(name__icontains=keyword) | Q(remark__icontains=keyword) | Q(areas__name__icontains=keyword)
            ).distinct()

        groups = base_groups
        if status == 'active':
            groups = groups.filter(is_active=True)
        elif status == 'inactive':
            groups = groups.filter(is_active=False)

        sort_by = f'-{sort_by}' if sort_order == 'desc' else sort_by
        groups = groups.order_by(sort_by)
        total = groups.count()

        all_count = base_groups.count()
        active_count = base_groups.filter(is_active=True).count()
        inactive_count = all_count - active_count

        start = (page - 1) * page_size
        end = start + page_size
        groups_page = groups[start:end]

        # 时间通过 format_datetime 统一转本地时区
        result = [
            {
                'id': g.id, 'name': g.name, 'remark': g.remark or '',
                'area_ids': [a.id for a in g.areas.all()],
                'area_names': [a.name for a in g.areas.all()],
                'customer_count': g.customer_count,
                'area_count': g.area_count,
                'create_time': format_datetime(g.create_time),
                'update_time': format_datetime(g.update_time),
                'is_active': g.is_active
            }
            for g in groups_page
        ]

        response_data = {
            'code': 1, 'data': result,
            'pagination': {'total': total, 'page': page, 'page_size': page_size,
                           'total_pages': (total + page_size - 1) // page_size},
            'counts': {'all': all_count, 'active': active_count, 'inactive': inactive_count}
        }
        cache.set(cache_key, response_data, CACHE_API_LIST)
        return JsonResponse(response_data)

    except Exception as e:
        logger.error(f"查询区域组列表失败：{str(e)}")
        return JsonResponse({'code': 0, 'msg': f'加载失败：{str(e)}'})


@login_required
@permission_required('area_view')
def group_detail_api(request, pk):
    try:
        group = get_object_or_404(
            AreaGroup.objects.only('id', 'name', 'remark', 'create_time', 'update_time'),
            pk=pk
        )

        page = max(int(request.GET.get('a_page', 1)), 1)
        page_size = 10

        # 🔧 优化：先用 values_list 取 area_ids，避免加载所有区域对象
        area_ids = list(group.areas.values_list('id', flat=True))

        # 🔧 优化：批量查询客户数映射
        area_customer_map = dict(
            Customer.objects.filter(area_id__in=area_ids)
            .values('area_id')
            .annotate(count=Count('id'))
            .values_list('area_id', 'count')
        )

        # 再查询分页所需的区域对象
        all_areas_qs = group.areas.only('id', 'name', 'create_time').order_by('name')
        total_areas = len(area_ids)
        start = (page - 1) * page_size
        end = start + page_size
        areas_page = all_areas_qs[start:end]

        areas = [
            {
                'id': a.id,
                'name': a.name,
                'customer_count': area_customer_map.get(a.id, 0)
            } for a in areas_page
        ]

        # 时间通过 format_datetime 统一转本地时区
        data = {
            'id': group.id,
            'name': group.name,
            'remark': group.remark or '',
            'create_time': format_datetime(group.create_time),
            'update_time': format_datetime(group.update_time),
            'customer_count': sum(area_customer_map.values()),
            'areas': areas,
            'area_pagination': {
                'page': page,
                'total_pages': (total_areas + page_size - 1) // page_size,
                'total': total_areas
            }
        }
        return JsonResponse({'code': 1, 'data': data})

    except Exception as e:
        logger.error(f"查询区域组{pk}详情失败：{str(e)}")
        return JsonResponse({'code': 0, 'msg': f'查询失败：{str(e)}'})


@login_required
@permission_required('area_add')
def group_add(request):
    if request.method == 'POST':
        try:
            is_json = request.content_type == 'application/json'
            data = json.loads(request.body) if is_json else request.POST
            name = data.get('name', '').strip()
            remark = data.get('remark', '').strip()
            area_ids = data.get('area_ids', []) if is_json else data.getlist('area_ids[]')

            if not name or AreaGroup.objects.filter(name=name).exists():
                return JsonResponse({'code': 0, 'msg': '组名不能为空/已存在'})

            valid_areas = Area.objects.filter(id__in=area_ids).only('id', 'name')
            g = AreaGroup.objects.create(name=name, remark=remark)
            g.areas.set([a.id for a in valid_areas])

            create_operation_log(request=request, op_type='create', obj_type='area_group',
                                 obj_id=g.id, obj_name=g.name, detail=f"新增区域组：{g.name}")
            clear_group_cache()
            return JsonResponse({'code': 1, 'msg': '创建成功'})
        except Exception as e:
            logger.error(f"新增区域组失败：{str(e)}")
            return JsonResponse({'code': 0, 'msg': f'创建失败：{str(e)}'})
    return JsonResponse({'code': 0, 'msg': '仅支持POST请求'})


@login_required
@permission_required('area_edit')
def group_edit(request, pk):
    try:
        # 🔧 优化：only()限制字段
        g = get_object_or_404(AreaGroup.objects.only('id', 'name', 'remark'), pk=pk)
        if request.method == 'POST':
            is_json = request.content_type == 'application/json'
            data = json.loads(request.body) if is_json else request.POST
            name = data.get('name', '').strip()
            remark = data.get('remark', '').strip()
            area_ids = data.get('area_ids', []) if is_json else data.getlist('area_ids[]')

            if not name or AreaGroup.objects.filter(name=name).exclude(pk=pk).exists():
                return JsonResponse({'code': 0, 'msg': '组名不能为空/重复'})

            valid_area_ids = Area.objects.filter(id__in=area_ids).values_list('id', flat=True)
            g.name = name
            g.remark = remark
            g.save()
            g.areas.set(valid_area_ids)

            create_operation_log(request=request, op_type='update', obj_type='area_group',
                                 obj_id=g.id, obj_name=g.name, detail=f"编辑区域组")
            clear_group_cache(group_id=pk)
            return JsonResponse({'code': 1, 'msg': '修改成功'})
        return JsonResponse({'code': 0, 'msg': '仅支持POST请求'})
    except Exception as e:
        logger.error(f"编辑区域组失败：{str(e)}")
        return JsonResponse({'code': 0, 'msg': f'修改失败：{str(e)}'})


@login_required
@permission_required('area_delete')
def group_delete(request, pk):
    try:
        g = get_object_or_404(AreaGroup.objects.only('id', 'name'), pk=pk)
        # 🔥 修改为软删除
        g.is_active = False
        g.save()

        create_operation_log(request=request, op_type='disable', obj_type='area_group',
                             obj_id=pk, obj_name=g.name, detail=f"禁用区域组")
        clear_group_cache(group_id=pk)
        return JsonResponse({'code': 1, 'msg': '禁用成功'})
    except Exception as e:
        logger.error(f"禁用区域组失败：{str(e)}")
        return JsonResponse({'code': 0, 'msg': f'禁用失败：{str(e)}'})


# 新增：区域组启用功能
@login_required
@permission_required('area_edit')
def group_enable(request, pk):
    try:
        g = get_object_or_404(AreaGroup.objects.only('id', 'name'), pk=pk)
        g.is_active = True
        g.save()

        create_operation_log(request=request, op_type='enable', obj_type='area_group',
                             obj_id=pk, obj_name=g.name, detail=f"启用区域组")
        clear_group_cache(group_id=pk)
        return JsonResponse({'code': 1, 'msg': '启用成功'})
    except Exception as e:
        logger.error(f"启用区域组失败：{str(e)}")
        return JsonResponse({'code': 0, 'msg': f'启用失败：{str(e)}'})


# ===================== 页面入口 =====================
from django.views.decorators.cache import cache_page


@login_required
def area_page(request):
    """根据当前用户权限传递控制变量，禁止缓存全页"""
    can_add = request.user.has_permission('area_add')      # 新增、导入共用此权限
    can_edit = request.user.has_permission('area_edit')    # 编辑、启用
    can_delete = request.user.has_permission('area_delete')# 删除/禁用
    return render(request, 'area_manage/area.html', {
        'can_add': can_add,
        'can_edit': can_edit,
        'can_delete': can_delete,
    })


@login_required
# 移除 @cache_page(CACHE_GROUP_PAGE)   <-- 必须去掉，否则权限控制失效
def group_page(request):
    """区域组管理页面"""
    can_add = request.user.has_permission('area_add')      # 控制新增、导入
    can_edit = request.user.has_permission('area_edit')    # 控制编辑、启用
    can_delete = request.user.has_permission('area_delete')# 控制禁用
    return render(request, 'area_manage/group.html', {
        'can_add': can_add,
        'can_edit': can_edit,
        'can_delete': can_delete,
    })


@login_required
@cache_page(CACHE_DETAIL_PAGE)
def area_detail_page(request, pk):
    # 🔧 优化：only()限制字段 + 修复update_time错误
    area = get_object_or_404(Area.objects.only('id', 'name', 'remark', 'create_time', 'update_time'), pk=pk)
    customer_count = get_area_statistics(pk)['customer_count']
    related_groups = AreaGroup.objects.filter(areas=area)

    update_time = format_datetime(area.update_time) if hasattr(area, 'update_time') else format_datetime(
        area.create_time)
    area_data = {
        'id': area.id, 'name': area.name, 'code': '', 'parent_name': '',
        'remark': area.remark or '', 'customer_count': customer_count,
        'create_time': format_datetime(area.create_time),
        'update_time': update_time
    }
    return render(request, 'area_manage/area_detail.html', {'area': area_data, 'related_groups': related_groups})


@login_required
@cache_page(CACHE_DETAIL_PAGE)
def group_detail_page(request, pk):
    # 🔧 优化：only()限制字段 + 修复update_time错误
    group = get_object_or_404(AreaGroup.objects.only('id', 'name', 'remark', 'create_time', 'update_time'), pk=pk)
    customer_count = get_group_statistics(pk)['customer_count']

    group_data = {
        'id': group.id, 'name': group.name,
        'area_names': [a.name for a in group.areas.all()],
        'remark': group.remark or '', 'customer_count': customer_count,
        'create_time': format_datetime(group.create_time),
        'update_time': format_datetime(group.update_time)
    }
    return render(request, 'area_manage/group_detail.html', {'group': group_data})


# ===================== 区域管理：导入导出新增代码 =====================
@login_required
@permission_required('area_add')
def area_import(request):
    """
    区域批量导入：
    读取Excel，跳过表头和序号列，区域名重复则跳过
    """
    if request.method == 'POST':
        try:
            file = request.FILES.get('file')
            if not file:
                return JsonResponse({'code': 0, 'msg': '请选择文件'})

            wb = openpyxl.load_workbook(file)
            ws = wb.active

            imported_count = 0
            skipped_count = 0
            errors = []

            # 从第2行开始遍历（第1行是表头）
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # row结构: (序号, 区域名, 备注)
                if len(row) < 2:
                    continue

                # 提取数据，忽略第一列序号
                area_name = str(row[1]).strip() if row[1] else ''
                remark = str(row[2]).strip() if len(row) > 2 and row[2] else ''

                if not area_name:
                    continue

                # 检查是否已存在
                if Area.objects.filter(name=area_name).exists():
                    skipped_count += 1
                    continue

                # 创建新区域
                Area.objects.create(name=area_name, remark=remark)
                imported_count += 1

            # 清理缓存
            clear_area_cache()

            # 记录日志
            create_operation_log(
                request=request, op_type='import', obj_type='area',
                obj_id=0, obj_name='批量导入',
                detail=f"导入成功：新增{imported_count}条，跳过{skipped_count}条重复"
            )

            return JsonResponse({
                'code': 1,
                'msg': f'导入完成！新增 {imported_count} 条，跳过 {skipped_count} 条重复数据'
            })

        except Exception as e:
            logger.error(f"导入区域失败：{str(e)}", exc_info=True)
            return JsonResponse({'code': 0, 'msg': f'导入失败：文件格式错误或数据异常'})
    return JsonResponse({'code': 0, 'msg': '仅支持POST请求'})


# ========== 区域导出新逻辑（支持字段选择） ==========
@login_required
@permission_required('area_view')
def area_export(request):
    """
    区域批量导出（支持字段选择和自定义字段）
    """
    try:
        if request.method == 'POST':
            # 1. 获取选中的字段
            selected_fields = request.POST.getlist('fields[]')
            if not selected_fields:
                return JsonResponse({'code': 0, 'msg': '请至少选择一个导出字段'})

            # 2. 获取自定义字段
            custom_fields_json = request.POST.get('custom_fields', '[]')
            try:
                custom_fields = json.loads(custom_fields_json)
            except json.JSONDecodeError:
                custom_fields = []

            # 3. 定义表头映射
            headers = {
                'serial': '序号',
                'id': 'ID',
                'name': '区域名',
                'remark': '备注'
            }

            # 4. 查询并格式化数据
            areas = Area.objects.only('id', 'name', 'remark').order_by('id')
            data = []
            seq = 1
            for area in areas:
                data.append({
                    'serial': seq,
                    'id': area.id,
                    'name': area.name,
                    'remark': area.remark or ''
                })
                seq += 1

            # 5. 生成文件名 - 修复：使用本地日期
            date_str = timezone.localdate().strftime('%Y年%m月%d日')
            file_name = f'{date_str}区域管理导出'

            # 6. 调用通用导出函数（不传 total_row 即无合计）
            response = export_to_excel(
                data=data,
                title='区域列表',
                headers=headers,
                selected_fields=selected_fields,
                custom_fields=custom_fields,
                file_name=file_name,
                total_row=None
            )

            return response
        else:
            return JsonResponse({'code': 0, 'msg': '请求方式错误'})
    except Exception as e:
        logger.error(f"导出区域失败：{str(e)}", exc_info=True)
        return JsonResponse({'code': 0, 'msg': '导出失败'})

@login_required
@permission_required('area_add')
def group_import(request):
    """
    区域组批量导入（支持覆盖更新）：
    读取Excel，格式：[序号, 组名, 包含区域(逗号/中文逗号/空格分隔), 备注]
    组名存在则更新区域列表，不存在则新建；区域名不存在则自动创建。
    """
    if request.method == 'POST':
        try:
            file = request.FILES.get('file')
            if not file:
                return JsonResponse({'code': 0, 'msg': '请选择文件'})

            wb = openpyxl.load_workbook(file)
            ws = wb.active

            created_count = 0      # 新增组数
            updated_count = 0      # 更新组数
            skipped_count = 0      # 跳过行数（无效数据）
            created_area_count = 0 # 新建区域数

            # 预加载所有区域（包括禁用的），便于匹配和新建
            area_map = {a.name: a for a in Area.objects.only('id', 'name')}

            # 从第2行开始遍历
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if len(row) < 2:
                    continue

                group_name = str(row[1]).strip() if row[1] else ''
                area_names_str = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                remark = str(row[3]).strip() if len(row) > 3 and row[3] else ''

                if not group_name:
                    continue

                # 处理分隔符：支持中文逗号、英文逗号、空格（多个空格合并）
                # 先将中文逗号替换为英文逗号，再按英文逗号分割，最后按空格分割并过滤空字符串
                if area_names_str:
                    # 统一替换中文逗号为英文逗号
                    area_names_str = area_names_str.replace('，', ',').replace('、', ',')
                    # 按英文逗号分割，再按空格分割，最后过滤空字符串
                    raw_names = []
                    for part in area_names_str.split(','):
                        for sub in part.split():
                            if sub.strip():
                                raw_names.append(sub.strip())
                    area_names = raw_names
                else:
                    area_names = []

                # 解析区域对象
                valid_areas = []
                for name in area_names:
                    if name in area_map:
                        valid_areas.append(area_map[name])
                    else:
                        # 区域不存在，创建新区域（默认启用）
                        new_area = Area.objects.create(name=name, remark='')
                        area_map[name] = new_area
                        valid_areas.append(new_area)
                        created_area_count += 1

                if not valid_areas:
                    skipped_count += 1
                    continue

                # 查找或创建区域组
                group, created = AreaGroup.objects.get_or_create(name=group_name, defaults={'remark': remark})
                if created:
                    created_count += 1
                else:
                    # 更新备注（若Excel有提供新备注，可覆盖；否则保留原备注）
                    if remark:
                        group.remark = remark
                        group.save()
                    updated_count += 1

                # 更新区域关联（先清空再设置）
                group.areas.set(valid_areas)

            # 清理缓存
            clear_group_cache()

            # 记录操作日志
            create_operation_log(
                request=request, op_type='import', obj_type='area_group',
                obj_id=0, obj_name='批量导入',
                detail=f"导入完成：新增 {created_count} 组，更新 {updated_count} 组，跳过 {skipped_count} 行（无有效区域），新建区域 {created_area_count} 个"
            )

            return JsonResponse({
                'code': 1,
                'msg': f'导入完成！新增 {created_count} 组，更新 {updated_count} 组，跳过 {skipped_count} 行，新建区域 {created_area_count} 个'
            })

        except Exception as e:
            logger.error(f"导入区域组失败：{str(e)}", exc_info=True)
            return JsonResponse({'code': 0, 'msg': f'导入失败：{e}'})
    return JsonResponse({'code': 0, 'msg': '仅支持POST请求'})


@login_required
@permission_required('area_view')
def group_export(request):
    """
    区域组批量导出（支持字段选择和自定义字段）
    """
    try:
        if request.method == 'POST':
            # 1. 获取选中的字段
            selected_fields = request.POST.getlist('fields[]')
            if not selected_fields:
                return JsonResponse({'code': 0, 'msg': '请至少选择一个导出字段'})

            # 2. 获取自定义字段
            custom_fields_json = request.POST.get('custom_fields', '[]')
            try:
                custom_fields = json.loads(custom_fields_json)
            except json.JSONDecodeError:
                custom_fields = []

            # 3. 定义表头映射
            headers = {
                'serial': '序号',
                'id': 'ID',
                'name': '组名',
                'areas': '包含区域',
                'customer_count': '客户数',
                'remark': '备注'
            }

            # 4. 查询并格式化数据（使用之前优化过的统计逻辑）
            customer_subquery = Customer.objects.filter(
                area_id=OuterRef('areas__id')
            ).values('area_id').annotate(count=Count('id')).values('count')

            groups = AreaGroup.objects.only('id', 'name', 'remark')\
                .prefetch_related(Prefetch('areas', queryset=Area.objects.only('id', 'name')))\
                .annotate(
                    customer_count=Coalesce(Sum(Subquery(customer_subquery)), 0),
                    area_count=Count('areas', distinct=True)
                ).order_by('id')

            data = []
            seq = 1
            for g in groups:
                area_names = ', '.join([a.name for a in g.areas.all()])
                data.append({
                    'serial': seq,
                    'id': g.id,
                    'name': g.name,
                    'areas': area_names,
                    'customer_count': g.customer_count,
                    'remark': g.remark or ''
                })
                seq += 1

            # 5. 生成文件名 - 修复：使用本地日期
            date_str = timezone.localdate().strftime('%Y年%m月%d日')
            file_name = f'{date_str}区域组管理导出'

            # 6. 调用通用导出函数
            response = export_to_excel(
                data=data,
                title='区域组列表',
                headers=headers,
                selected_fields=selected_fields,
                custom_fields=custom_fields,
                file_name=file_name,
                total_row=None
            )

            return response
        else:
            return JsonResponse({'code': 0, 'msg': '请求方式错误'})
    except Exception as e:
        logger.error(f"导出区域组失败：{str(e)}", exc_info=True)
        return JsonResponse({'code': 0, 'msg': '导出失败'})

# ===================== 修复版：区域深度统计API =====================
@login_required
@permission_required('area_view')
def area_statistics_api(request, pk):
    """
    按需统计：点击按钮才执行的深度计算
    数据范围：该区域下所有非作废订单
    """
    try:
        # 1. 基础校验
        area = get_object_or_404(Area.objects.only('id'), pk=pk)

        # 2. 锁定订单范围 (利用索引: status, area, is_settled)
        base_orders = Order.objects.filter(
            area_id=pk,
            status__in=['pending', 'printed', 'reopened']  # 排除作废
        ).only('id', 'total_amount', 'is_settled')

        # 3. 统计1：订单总金额 & 欠款 (修复：显式指定 output_field)
        agg_result = base_orders.aggregate(
            total_order_amount=Coalesce(
                Sum('total_amount'),
                0,
                output_field=DecimalField(max_digits=12, decimal_places=2)
            ),
            total_debt=Coalesce(
                Sum('total_amount', filter=Q(is_settled=False)),
                0,
                output_field=DecimalField(max_digits=12, decimal_places=2)
            )
        )

        # 4. 统计2：商品销售汇总 (利用 OrderItem 索引)
        items = OrderItem.objects.filter(
            order__area_id=pk,
            order__status__in=['pending', 'printed', 'reopened']
        ).select_related('product').only(
            'product__id', 'product__name', 'quantity', 'amount'
        )

        product_summary = {}
        for item in items:
            pid = item.product.id if item.product else 0
            pname = item.product.name if item.product else "未知商品"
            if pid not in product_summary:
                product_summary[pid] = {
                    'product_id': pid,
                    'product_name': pname,
                    'total_quantity': 0,
                    'total_amount': 0.0
                }
            product_summary[pid]['total_quantity'] += item.quantity or 0
            product_summary[pid]['total_amount'] += float(item.amount) if item.amount else 0.0

        # 转换为列表并排序
        product_list = sorted(product_summary.values(), key=lambda x: -x['total_amount'])

        # 5. 组装返回数据 - 修复：计算时间转为本地时区
        data = {
            'total_order_amount': float(agg_result['total_order_amount']),
            'total_debt': float(agg_result['total_debt']),
            'product_summary': product_list,
            'calculated_at': timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M:%S')
        }

        return JsonResponse({'code': 1, 'data': data})

    except Exception as e:
        logger.error(f"区域{pk}统计失败：{str(e)}", exc_info=True)
        return JsonResponse({'code': 0, 'msg': f'统计失败：{str(e)}'})

# ===================== 新增：区域组深度统计API (模仿区域统计) =====================
@login_required
@permission_required('area_view')
def group_statistics_api(request, pk):
    """
    按需统计：点击按钮才执行的深度计算
    数据范围：该区域组下所有非作废订单
    """
    try:
        # 1. 基础校验 & 获取区域ID列表
        group = get_object_or_404(AreaGroup.objects.only('id'), pk=pk)
        area_ids = group.areas.values_list('id', flat=True)

        if not area_ids:
            return JsonResponse({'code': 1, 'data': {
                'total_order_amount': 0.0,
                'total_debt': 0.0,
                'product_summary': [],
                'calculated_at': timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M:%S')
            }})

        # 2. 锁定订单范围 (利用索引: status, area)
        base_orders = Order.objects.filter(
            area_id__in=area_ids,
            status__in=['pending', 'printed', 'reopened']
        ).only('id', 'total_amount', 'is_settled')

        # 3. 统计1：订单总金额 & 欠款
        agg_result = base_orders.aggregate(
            total_order_amount=Coalesce(
                Sum('total_amount'),
                0,
                output_field=DecimalField(max_digits=12, decimal_places=2)
            ),
            total_debt=Coalesce(
                Sum('total_amount', filter=Q(is_settled=False)),
                0,
                output_field=DecimalField(max_digits=12, decimal_places=2)
            )
        )

        # 4. 统计2：商品销售汇总
        items = OrderItem.objects.filter(
            order__area_id__in=area_ids,
            order__status__in=['pending', 'printed', 'reopened']
        ).select_related('product').only(
            'product__id', 'product__name', 'quantity', 'amount'
        )

        product_summary = {}
        for item in items:
            pid = item.product.id if item.product else 0
            pname = item.product.name if item.product else "未知商品"
            if pid not in product_summary:
                product_summary[pid] = {
                    'product_id': pid,
                    'product_name': pname,
                    'total_quantity': 0,
                    'total_amount': 0.0
                }
            product_summary[pid]['total_quantity'] += item.quantity or 0
            product_summary[pid]['total_amount'] += float(item.amount) if item.amount else 0.0

        product_list = sorted(product_summary.values(), key=lambda x: -x['total_amount'])

        # 修复：计算时间转为本地时区
        data = {
            'total_order_amount': float(agg_result['total_order_amount']),
            'total_debt': float(agg_result['total_debt']),
            'product_summary': product_list,
            'calculated_at': timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M:%S')
        }
        return JsonResponse({'code': 1, 'data': data})

    except Exception as e:
        logger.error(f"区域组{pk}统计失败：{str(e)}", exc_info=True)
        return JsonResponse({'code': 0, 'msg': f'统计失败：{str(e)}'})


# 工具函数：解析时间范围
def parse_time_range(time_range, start_date_str, end_date_str):
    from datetime import datetime, timedelta
    # 修复：使用上海本地日期，避免UTC日期偏差1天
    today = timezone.localdate()

    if time_range == 'today':
        return today, today
    elif time_range == 'week':
        start = today - timedelta(days=today.weekday())
        return start, today
    elif time_range == 'month':
        return today.replace(day=1), today
    elif time_range == 'custom' and start_date_str and end_date_str:
        try:
            start = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            return start, end
        except:
            pass
    # 默认：最近30天
    return today - timedelta(days=30), today


# 1. 区域统计页面入口
@login_required
@permission_required('area_view')
def area_stats_page(request):
    return render(request, 'area_manage/area_stats.html')


# 2. 核心：区域统计数据接口（统一入口）
@login_required
@permission_required('area_view')
def calculate_area_stats(request):
    try:
        # 获取参数
        time_range = request.GET.get('time_range', '30days')
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')
        top_type = request.GET.get('top_type', 'amount')  # amount/order

        # 1. 解析时间
        start_dt, end_dt = parse_time_range(time_range, start_date, end_date)

        # 2. 基础订单QuerySet（利用索引：status, area, create_time）
        base_orders = Order.objects.filter(
            status__in=['pending', 'printed', 'reopened'],
            create_time__date__gte=start_dt,
            create_time__date__lte=end_dt
        ).select_related('area')

        # 3. 全局核心指标
        global_stats = base_orders.aggregate(
            total_sales=Coalesce(Sum('total_amount'), 0, output_field=DecimalField(max_digits=12, decimal_places=2)),
            total_orders=Count('id')
        )

        # 4. 区域明细列表（按区域分组统计）
        area_details = base_orders.values(
            'area_id', 'area__name'
        ).annotate(
            sales=Coalesce(Sum('total_amount'), 0, output_field=DecimalField(max_digits=12, decimal_places=2)),
            order_count=Count('id'),
            last_order_time=Max('create_time')
        ).order_by('-sales')

        # 计算总销售额用于占比
        total_sales_val = float(global_stats['total_sales']) if global_stats['total_sales'] else 0

        # 组装区域明细数据
        area_list = []
        for item in area_details:
            contribution = 0.0
            if total_sales_val > 0:
                contribution = (float(item['sales']) / total_sales_val) * 100

            # 修复：最后下单时间转为上海本地时区再格式化
            last_order_str = ''
            if item['last_order_time']:
                last_order_local = timezone.localtime(item['last_order_time'])
                last_order_str = last_order_local.strftime('%Y-%m-%d %H:%M')

            area_list.append({
                'area_id': item['area_id'],
                'area_name': item['area__name'] or '未分配区域',
                'sales': float(item['sales']),
                'order_count': item['order_count'],
                'contribution': round(contribution, 2),
                'last_order_time': last_order_str if last_order_str else '无'
            })

        # 5. TOP 排行（直接复用 area_list 的排序结果）
        top_list = area_list[:30]
        if top_type == 'order':
            top_list = sorted(area_list, key=lambda x: -x['order_count'])[:30]

        return JsonResponse({
            'code': 1,
            'global_stats': {
                'total_sales': float(global_stats['total_sales']),
                'total_orders': global_stats['total_orders']
            },
            'area_list': area_list,
            'top_list': top_list,
            'date_range': {
                'start': start_dt.strftime('%Y-%m-%d'),
                'end': end_dt.strftime('%Y-%m-%d')
            }
        })

    except Exception as e:
        logger.error(f"区域统计计算失败：{str(e)}", exc_info=True)
        return JsonResponse({'code': 0, 'msg': f'统计失败：{str(e)}'})


# 3. 区域画像详情接口（点击查看详情时调用）
@login_required
@permission_required('area_view')
def area_portrait_api(request, area_id):
    try:
        area = get_object_or_404(Area.objects.only('id', 'name', 'remark', 'create_time'), pk=area_id)

        # 时间范围 - 修复：使用上海本地日期
        end_date = timezone.localdate()
        start_date = end_date - timezone.timedelta(days=30)

        # 该区域订单基础数据
        base_orders = Order.objects.filter(
            area_id=area_id,
            status__in=['pending', 'printed', 'reopened'],
            create_time__date__gte=start_date,
            create_time__date__lte=end_date
        )

        # 1. 区域核心指标
        area_stats = base_orders.aggregate(
            total_sales=Coalesce(Sum('total_amount'), 0, output_field=DecimalField(max_digits=12, decimal_places=2)),
            total_orders=Count('id'),
            total_debt=Coalesce(
                Sum('total_amount', filter=Q(is_settled=False)),
                0,
                output_field=DecimalField(max_digits=12, decimal_places=2)
            )
        )

        # 2. 该区域客户 TOP 10
        customer_top = base_orders.filter(
            customer__isnull=False
        ).values(
            'customer_id', 'customer__name'
        ).annotate(
            sales=Sum('total_amount'),
            order_count=Count('id')
        ).order_by('-sales')[:10]

        customer_list = [
            {
                'customer_id': item['customer_id'],
                'customer_name': item['customer__name'],
                'sales': float(item['sales']),
                'order_count': item['order_count']
            } for item in customer_top
        ]

        # 修复：区域创建时间转为本地时区再格式化
        return JsonResponse({
            'code': 1,
            'area_info': {
                'id': area.id,
                'name': area.name,
                'remark': area.remark or '无',
                'create_time': timezone.localtime(area.create_time).strftime('%Y-%m-%d')
            },
            'stats': {
                'total_sales': float(area_stats['total_sales']),
                'total_orders': area_stats['total_orders'],
                'total_debt': float(area_stats['total_debt'])
            },
            'customer_top': customer_list
        })

    except Exception as e:
        logger.error(f"获取区域画像失败：{str(e)}", exc_info=True)
        return JsonResponse({'code': 0, 'msg': f'获取失败：{str(e)}'})


# ===================== 新增：区域组统计页面入口 =====================
@login_required
@permission_required('area_view')
def group_stats_page(request):
    return render(request, 'area_manage/group_stats.html')


# ===================== 新增：核心区域组统计数据接口 =====================
@login_required
@permission_required('area_view')
def calculate_group_stats(request):
    try:
        time_range = request.GET.get('time_range', '30days')
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')
        sort_by = request.GET.get('sort_by', 'sales')
        keyword = request.GET.get('keyword', '').strip()

        start_dt, end_dt = parse_time_range(time_range, start_date, end_date)

        # 1. 预加载所有区域组及其包含的区域映射
        group_area_map = {}
        all_groups = AreaGroup.objects.only('id', 'name').filter(is_active=True)
        for g in all_groups:
            group_area_map[g.id] = {
                'name': g.name,
                'area_ids': set(g.areas.values_list('id', flat=True))
            }

        # 2. 🔧 核心优化：先在数据库层按 Area 聚合订单数据，避免加载全量 Order
        area_order_stats = Order.objects.filter(
            status__in=['pending', 'printed', 'reopened'],
            create_time__date__gte=start_dt,
            create_time__date__lte=end_dt,
            area_id__isnull=False
        ).values('area_id').annotate(
            sales=Coalesce(Sum('total_amount'), 0, output_field=DecimalField(max_digits=12, decimal_places=2)),
            order_count=Count('id'),
            debt=Coalesce(
                Sum('total_amount', filter=Q(is_settled=False)),
                0,
                output_field=DecimalField(max_digits=12, decimal_places=2)
            )
        )

        # 3. 初始化组统计数据
        group_stats_data = {}
        global_total_sales = 0.0
        global_total_orders = 0
        global_total_debt = 0.0

        for g_id, g_info in group_area_map.items():
            group_stats_data[g_id] = {
                'group_id': g_id,
                'group_name': g_info['name'],
                'sales': 0.0,
                'order_count': 0,
                'debt': 0.0,
                'area_ids': g_info['area_ids']
            }

        # 4. 遍历按 Area 聚合的结果，累加到对应的 Group
        for area_stat in area_order_stats:
            area_id = area_stat['area_id']
            sales = float(area_stat['sales'])
            order_count = area_stat['order_count']
            debt = float(area_stat['debt'])

            # 全局累加
            global_total_sales += sales
            global_total_orders += order_count
            global_total_debt += debt

            # 匹配区域组并累加
            for g_id, g_data in group_stats_data.items():
                if area_id in g_data['area_ids']:
                    g_data['sales'] += sales
                    g_data['order_count'] += order_count
                    g_data['debt'] += debt

        # 5. 后续处理（计算占比、过滤、排序）保持不变
        group_list = list(group_stats_data.values())
        for g in group_list:
            g['contribution'] = round((g['sales'] / global_total_sales) * 100, 2) if global_total_sales > 0 else 0.0
            del g['area_ids']

        if keyword:
            group_list = [g for g in group_list if keyword.lower() in g['group_name'].lower()]

        if sort_by == 'order_count':
            group_list.sort(key=lambda x: (-x['order_count'], -x['sales']))
        elif sort_by == 'debt':
            group_list.sort(key=lambda x: (-x['debt'], -x['sales']))
        else:
            group_list.sort(key=lambda x: (-x['sales'], -x['order_count']))

        for idx, g in enumerate(group_list):
            g['rank'] = idx + 1

        return JsonResponse({
            'code': 1,
            'global_stats': {
                'total_sales': round(global_total_sales, 2),
                'total_orders': global_total_orders,
                'total_debt': round(global_total_debt, 2)
            },
            'group_list': group_list,
            'date_range': {
                'start': start_dt.strftime('%Y-%m-%d'),
                'end': end_dt.strftime('%Y-%m-%d')
            }
        })

    except Exception as e:
        logger.error(f"区域组统计计算失败：{str(e)}", exc_info=True)
        return JsonResponse({'code': 0, 'msg': f'统计失败：{str(e)}'})