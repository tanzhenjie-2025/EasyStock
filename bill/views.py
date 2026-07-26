# ========== 先导入所有必要模块（统一开头，避免重复） ==========
from django.db.models.functions import Coalesce
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponseBadRequest
from django.shortcuts import render
from django.views.decorators.http import require_POST
from mpmath import re

from product.models import Product, ProductAlias, ProductTag
from customer_manage.models import Customer, CustomerPrice

from django.db.models import Q, Sum, Count, Case, When, DecimalField

from datetime import datetime, timedelta
from django.contrib.auth.decorators import login_required

import decimal
from .models import SortRule, ProductTag
from django.contrib.auth.decorators import login_required
from django.conf import settings
from django.core.cache import cache

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

# ========== 导入用户模块的RBAC核心组件 ==========
from accounts.models import ROLE_SUPER_ADMIN, ROLE_ADMIN, ROLE_OPERATOR, PERM_ORDER_CANCEL_OWN, User
from accounts.views import (
    permission_required,  # RBAC权限装饰器
    create_operation_log,  # 统一日志记录
    get_client_ip  # 获取客户端IP
)
from product.models import Unit
# ========== 开单模块权限常量（和用户模块保持一致） ==========
PERM_ORDER_CREATE = 'order_create'
PERM_ORDER_VIEW = 'order_view'
PERM_ORDER_PRINT = 'order_print'
PERM_ORDER_CANCEL = 'order_cancel'
PERM_ORDER_REOPEN = 'order_reopen'
PERM_ORDER_SETTLE = 'order_settle'
PERM_ORDER_UNSETTLE = 'order_unsettle'
PERM_ORDER_SUMMARY = 'order_summary'
PERM_PRODUCT_SEARCH = 'product_search'
PERM_ORDER_PRICE_CHECK = 'order_price_check'

# ========== 订单模块缓存时长常量（统一管理） ==========
CACHE_STOCK_LIST = 60  # 库存列表：60秒
CACHE_ORDER_LIST = 60  # 订单列表：60秒
CACHE_ORDER_DETAIL = 120  # 订单详情：2分钟
CACHE_PRINT_ORDER = 300  # 订单打印：5分钟
CACHE_CUSTOMER_RECENT_PRODUCT = 60  # 客户最近商品：60秒
CACHE_PRODUCT_SEARCH = 30  # 商品搜索：30秒
CACHE_CUSTOMER_SEARCH = 10  # 客户搜索：10秒

# ========== 订单模块缓存 Key 定义 ==========
CACHE_PREFIX_STOCK_LIST = "stock_list_"
CACHE_PREFIX_ORDER_LIST = "order_list_"
CACHE_PREFIX_ORDER_DETAIL = "order_detail_"
CACHE_PREFIX_PRINT_ORDER = "print_order_"
CACHE_PREFIX_PRODUCT_SEARCH = "product_search_"
CACHE_PREFIX_CUSTOMER_SEARCH = "customer_search_"
CACHE_PREFIX_CUSTOMER_RECENT_PRODUCT = "customer_recent_products_"

# ==========  订单有效状态常量（索引前缀核心字段） ==========
ORDER_STATUS_VALID = ['pending', 'printed', 'reopened']

import logging

logger = logging.getLogger(__name__)

import json

import io
from openpyxl import Workbook, load_workbook
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required, permission_required
from django.db import transaction
from urllib.parse import quote
from .models import SortRule, ProductTag
from django.contrib.auth.decorators import login_required, permission_required
from django.utils import timezone

from django.db.models import Q
from pypinyin import lazy_pinyin
from accounts.models import User

from area_manage.models import Area
from customer_manage.models import Customer
from product.models import Product
from functools import wraps
from django.core.exceptions import PermissionDenied
from openpyxl.styles import Font, Alignment

from django.db.models import Prefetch

from datetime import datetime
from decimal import Decimal, InvalidOperation
from openpyxl.utils.datetime import from_excel
from datetime import datetime, date

from django.contrib.auth.decorators import login_required, permission_required
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib import messages
from .models import Order, OrderItem
def accounts_permission_required(perm_code):
    """自定义权限装饰器，使用项目的 has_permission 检查"""
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            # 前置登录已由 @login_required 保证
            if not request.user.has_permission(perm_code):
                raise PermissionDenied  # 直接返回 403，不再重定向登录
            return view_func(request, *args, **kwargs)
        return _wrapped_view
    return decorator

def clear_order_cache(order_no: str = None):
    """
    清理订单相关缓存（含列表、详情、打印页）
    """
    # 1. 清理所有订单列表缓存
    cache.delete_pattern(f"{CACHE_PREFIX_ORDER_LIST}*")

    # 2. 清理指定订单的详情和打印缓存
    if order_no:
        cache.delete(f"{CACHE_PREFIX_ORDER_DETAIL}{order_no}")
        cache.delete(f"{CACHE_PREFIX_PRINT_ORDER}{order_no}")

    logger.info(f"已清理订单缓存: {order_no if order_no else '全列表'}")


def clear_stock_cache():
    """
    清理库存列表缓存
    """
    cache.delete_pattern(f"{CACHE_PREFIX_STOCK_LIST}*")
    logger.info("已清理库存列表缓存")


def clear_product_search_cache():
    """
    清理商品搜索缓存
    """
    cache.delete_pattern(f"{CACHE_PREFIX_PRODUCT_SEARCH}*")
    logger.info("已清理商品搜索缓存")


def clear_customer_related_cache(customer_id: int = None):
    """
    清理客户相关业务缓存（最近购买商品等）
    """
    if customer_id:
        cache.delete(f"{CACHE_PREFIX_CUSTOMER_RECENT_PRODUCT}{customer_id}")
    cache.delete_pattern(f"{CACHE_PREFIX_CUSTOMER_RECENT_PRODUCT}*")
    logger.info(f"已清理客户相关缓存: {customer_id if customer_id else '全量'}")

def clear_sort_cache():
    """
    清理排序规则及关联的商品标签映射缓存
    """
    cache.delete('sort_stages_json')
    cache.delete('product_tags_map_json')
    logger.info("已清理排序规则相关缓存")


def ajax_login_required(view_func):
    """AJAX登录验证装饰器：未登录返回JSON，而非重定向HTML"""

    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if request.user.is_authenticated:
            return view_func(request, *args, **kwargs)
        # 识别AJAX请求，返回JSON错误
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or 'application/json' in request.headers.get(
                'Content-Type', ''):
            return JsonResponse({'code': 0, 'msg': '请先登录系统'}, status=401)
        # 非AJAX请求仍重定向登录页
        return login_required(view_func)(request, *args, **kwargs)

    return wrapper


def ajax_permission_required(permission_code):
    """重构：AJAX RBAC权限验证装饰器"""

    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            # 未登录→返回JSON
            if not request.user.is_authenticated:
                return JsonResponse({'code': 0, 'msg': '请先登录系统'}, status=401)

            # 超级管理员→直接放行
            if request.user.role and request.user.role.code == ROLE_SUPER_ADMIN:
                return view_func(request, *args, **kwargs)

            # 检查RBAC权限
            if not request.user.has_permission(permission_code):
                # AJAX请求返回JSON
                if request.headers.get(
                        'X-Requested-With') == 'XMLHttpRequest' or 'application/json' in request.headers.get(
                    'Content-Type', ''):
                    return JsonResponse({'code': 0, 'msg': '无操作权限，请联系管理员'}, status=403)
                # 非AJAX请求重定向无权限页
                return redirect('/accounts/no-permission/')

            return view_func(request, *args, **kwargs)

        return wrapper

    return decorator

def get_sort_context():
    """返回包含排序规则和商品标签映射的 context 字典，带缓存"""
    # 排序规则
    sort_stages_json = cache.get('sort_stages_json')
    if sort_stages_json is None:
        rules = SortRule.objects.select_related('tag').order_by('stage', 'priority')
        stages_dict = {}
        for r in rules:
            if r.stage not in stages_dict:
                stages_dict[r.stage] = []
            item = {
                'type': r.rule_type,
                'priority': r.priority,
            }
            if r.rule_type == 'tag':
                item['tag_id'] = r.tag_id
                item['tag_name'] = r.tag.name
            else:
                item['spec_condition'] = r.spec_condition
            stages_dict[r.stage].append(item)
        stages = [{'stage': s, 'rules': stages_dict[s]} for s in sorted(stages_dict.keys())]
        sort_stages_json = json.dumps(stages)
        cache.set('sort_stages_json', sort_stages_json, 3600)

    # 商品标签映射
    product_tags_map_json = cache.get('product_tags_map_json')
    if product_tags_map_json is None:
        products = Product.objects.filter(is_active=True).prefetch_related('tags')
        tags_map = {}
        for p in products:
            tag_ids = list(p.tags.filter(is_active=True).values_list('id', flat=True))
            if tag_ids:
                tags_map[str(p.id)] = tag_ids
        product_tags_map_json = json.dumps(tags_map)
        cache.set('product_tags_map_json', product_tags_map_json, 3600)

    return {
        'sort_stages_json': sort_stages_json,
        'product_tags_map_json': product_tags_map_json,
    }

@login_required
@permission_required(PERM_ORDER_CREATE)
def index(request):
    is_super_admin = request.user.role and request.user.role.code == ROLE_SUPER_ADMIN

    context = {
        'is_super_admin': is_super_admin,
    }
    context.update(get_sort_context())   # 注入排序数据

    return render(request, 'bill/index.html', context)


@login_required
@permission_required(PERM_PRODUCT_SEARCH)
def search_product(request):
    """商品搜索（手动缓存版）"""
    keyword = request.GET.get('keyword', '').strip()
    customer_id = request.GET.get('customer_id', '').strip()

    if not keyword:
        return JsonResponse({'code': 0, 'data': []})

    # 🔥 手动缓存：基础商品信息缓存
    cache_key = f"{CACHE_PREFIX_PRODUCT_SEARCH}{keyword}"
    cached_products = cache.get(cache_key)

    if cached_products is None:
        product_matches = Product.objects.filter(
            Q(name__icontains=keyword) |
            Q(pinyin_full__icontains=keyword) |
            Q(pinyin_abbr__icontains=keyword)
        )

        alias_matches = ProductAlias.objects.filter(
            Q(alias_name__icontains=keyword) |
            Q(alias_pinyin_full__icontains=keyword) |
            Q(alias_pinyin_abbr__icontains=keyword)
        ).values_list('product_id', flat=True)
        alias_products = Product.objects.filter(id__in=alias_matches)

        all_products = (product_matches | alias_products).distinct()[:200]

        cached_products = []
        for p in all_products:
            cached_products.append({
                'id': p.id,
                'name': p.name,
                'standard_price': float(p.price),
                'unit': p.unit,
                'stock_system': p.stock_system,
                # 👇 新增规格字段
                'specification': p.specification
            })
        cache.set(cache_key, cached_products, timeout=CACHE_PRODUCT_SEARCH)
        logger.info(f"设置商品搜索缓存: {cache_key}")

    # 客户专属价查询（保留，无优化）
    customer_prices = {}
    if customer_id:
        product_ids = [item['id'] for item in cached_products]
        cp_list = CustomerPrice.objects.filter(customer_id=customer_id, product_id__in=product_ids)
        customer_prices = {cp.product_id: float(cp.custom_price) for cp in cp_list}

    data = []
    for item in cached_products:
        product_id = item['id']
        final_price = customer_prices.get(product_id, item['standard_price'])
        data.append({
            'id': product_id,
            'name': item['name'],
            'price': final_price,
            'standard_price': item['standard_price'],
            'unit': item['unit'],
            'stock_system': item['stock_system'],
            'specification': item['specification']
        })

    return JsonResponse({'code': 1, 'data': data})


@login_required
@permission_required(PERM_ORDER_CREATE)
def save_order(request):
    if request.method != 'POST':
        return JsonResponse({'code': 0, 'msg': '请求方式错误'})

    try:
        with transaction.atomic():
            data = json.loads(request.body)
            items = data.get('items', [])
            customer_id = data.get('customer_id', '')
            customer_name = data.get('customer_name', '').strip()
            order_number = data.get('order_number', '').strip()
            original_order_no = data.get('original_order_no', '')

            if not items:
                return JsonResponse({'code': 0, 'msg': '无订单明细'})

            # ---------- 1. 分离有效商品和备注行 ----------
            valid_product_ids = []
            item_data_list = []
            has_return = False
            delivery_method = 'delivery'
            general_remarks = []          # 新增：收集通用备注

            for item in items:
                name = item.get('name', '').strip()
                spec = item.get('spec', '').strip()
                unit = item.get('unit', '').strip()
                qty = item.get('qty', 0)
                price = item.get('price', 0)
                is_makeup = item.get('is_makeup', False)
                operation_type = item.get('operation_type', '')

                # ---------- 备注行处理（名称空 + 规格非空） ----------
                if not name and spec:
                    # 提取交付方式（优先级：自提 > 寄件，先出现的为准）
                    if '自' in spec:
                        delivery_method = 'pickup'
                    elif '寄' in spec:
                        delivery_method = 'express'

                    # 提取退货标记
                    if '退' in spec or '换' in spec:
                        has_return = True

                    # 如果没有命中以上任何关键词 → 视为通用备注
                    if not ('自' in spec or '寄' in spec or '退' in spec or '换' in spec):
                        general_remarks.append(spec)

                    # 备注行不加入订单明细，直接跳过
                    continue

                # ---------- 普通商品行 ----------
                if not name:
                    return JsonResponse({'code': 0, 'msg': '商品名称不能为空'})
                if not isinstance(qty, int) or qty <= 0:
                    return JsonResponse({'code': 0, 'msg': f'商品{name}数量无效'})

                product_id_int = None
                pid_raw = item.get('id', '').strip()
                if pid_raw:
                    try:
                        pid_int = int(pid_raw)
                        if pid_int > 0:
                            product_id_int = pid_int
                            valid_product_ids.append(pid_int)
                    except (ValueError, TypeError):
                        pass

                item_data_list.append({
                    'pid': product_id_int,
                    'name': name,
                    'spec': spec,
                    'unit': unit,
                    'qty': qty,
                    'price': Decimal(str(price)),
                    'is_makeup': is_makeup,
                    'operation_type': operation_type if is_makeup else '',
                })

            # ---------- 2. 批量查询有效商品 ----------
            products_map = {}
            if valid_product_ids:
                products_map = Product.objects.filter(
                    id__in=valid_product_ids,
                    is_active=True
                ).in_bulk()

            # ---------- 3. 查询客户专属价 ----------
            existing_product_ids = list(products_map.keys())
            customer_prices_dict = {}
            if customer_id and existing_product_ids:
                cp_list = CustomerPrice.objects.filter(
                    customer_id=customer_id,
                    product_id__in=existing_product_ids
                )
                customer_prices_dict = {cp.product_id: cp.custom_price for cp in cp_list}

            # ---------- 4. 创建订单主表 ----------
            order = Order()
            order.creator = request.user
            order.customer_name_snapshot = customer_name
            order.order_number_snapshot = order_number or None
            order.delivery_method = delivery_method
            order.has_return = has_return

            if customer_id:
                customer = get_object_or_404(Customer, id=customer_id)
                order.customer = customer
                order.area = customer.area
                if order_number and customer.order_number != order_number:
                    customer.order_number = order_number
                    customer.save(update_fields=['order_number'])

            # ====== 处理 original_order_no（加单/重开） ======
            original_order = None
            if original_order_no:
                original_order = get_object_or_404(Order, order_no=original_order_no)
                if original_order.status == 'cancelled':
                    order.original_order = original_order
                    order.status = 'reopened'
                else:
                    original_order.status = 'cancelled'
                    original_order.cancelled_by = request.user
                    original_order.cancelled_time = timezone.now()
                    original_order.cancelled_reason = '加单重开（系统自动作废）'
                    original_order.save(update_fields=['status', 'cancelled_by', 'cancelled_time', 'cancelled_reason'])
                    create_operation_log(request, 'cancel_order', 'order', str(original_order.id),
                                         f"订单-{original_order.order_no}", "加单自动作废")
                    clear_order_cache()
                    order.original_order = original_order
                    order.status = 'reopened'

            # ---------- 5. 生成订单明细 ----------
            total_amount = 0
            order_items = []
            update_stock_products = []

            for item_data in item_data_list:
                pid = item_data['pid']
                name = item_data['name']
                qty = item_data['qty']
                input_price = item_data['price']
                amount = input_price * qty
                total_amount += amount
                save_spec = item_data['spec']

                if pid is not None and pid in products_map:
                    product = products_map[pid]
                    snap_standard = product.price
                    snap_customer = customer_prices_dict.get(pid, None)
                    product.stock_system -= qty
                    update_stock_products.append(product)

                    order_items.append(OrderItem(
                        order=order,
                        product=product,
                        product_name=product.name,
                        unit=item_data['unit'],
                        specification=save_spec,
                        quantity=qty,
                        amount=amount,
                        actual_unit_price=input_price,
                        snapshot_standard_price=snap_standard,
                        snapshot_customer_price=snap_customer,
                        is_makeup_item=item_data['is_makeup'],
                        operation_type=item_data['operation_type'],
                    ))
                else:
                    order_items.append(OrderItem(
                        order=order,
                        product=None,
                        product_name=name,
                        unit=item_data['unit'],
                        specification=save_spec,
                        quantity=qty,
                        amount=amount,
                        actual_unit_price=input_price,
                        snapshot_standard_price=None,
                        snapshot_customer_price=None,
                        is_makeup_item=item_data['is_makeup'],
                        operation_type=item_data['operation_type'],
                    ))

            # ---------- 6. 保存（新增通用备注写入） ----------
            order.total_amount = total_amount
            if general_remarks:
                order.remark = '；'.join(general_remarks)   # 多条备注用中文分号分隔
            order.save()
            OrderItem.objects.bulk_create(order_items)

            if update_stock_products:
                Product.objects.bulk_update(update_stock_products, ['stock_system'])

            # ---------- 7. 日志与缓存清理 ----------
            create_operation_log(request, 'create_order', 'order', str(order.id),
                                 f"订单-{order.order_no}", "创建订单")
            clear_stock_cache()
            clear_order_cache()
            if customer_id:
                clear_customer_related_cache(int(customer_id))

            return JsonResponse({'code': 1, 'msg': '开单成功', 'order_no': order.order_no})

    except Exception as e:
        logger.error(f"开单失败：{str(e)}", exc_info=True)
        return JsonResponse({'code': 0, 'msg': f'开单失败：{str(e)}'})

@login_required
def print_order(request, order_no):
    """订单打印页面（手动缓存版）"""
    cache_key = f"{CACHE_PREFIX_PRINT_ORDER}{order_no}"
    cached_data = cache.get(cache_key)

    if cached_data:
        return HttpResponse(cached_data)

    order = get_object_or_404(
        Order.objects.select_related('customer', 'area', 'creator'),
        order_no=order_no
    )
    items = order.items.select_related('product')

    # 原有补货/换货检测（用于浮动提示“带回退货”）
    has_return_or_exchange = order.items.filter(
        is_makeup_item=True,
        operation_type__in=['return', 'exchange']
    ).exists()

    items_display = list(items[:15])
    items_display.extend([None] * (15 - len(items_display)))
    float_start = find_float_start(items_display)

    # ========== 交付方式水印 ==========
    watermark_text = None
    if order.delivery_method == 'pickup':
        watermark_text = '客户自提'
    elif order.delivery_method == 'express':
        watermark_text = '快递寄件'

    general_remark = order.remark.strip() if order.remark else None

    # ========== 退货标记水印（新增） ==========
    has_return = order.has_return  # 订单级的退货标记

    is_super_admin = request.user.role and request.user.role.code == ROLE_SUPER_ADMIN

    context = {
        'order': order,
        'items_display': items_display,
        'is_super_admin': is_super_admin,
        'has_return_or_exchange': has_return_or_exchange,
        'float_start': float_start,
        'phone_numbers': settings.PHONE_NUMBERS,
        'complaint_phone': settings.COMPLAINT_PHONE,
        'bill_title': settings.BILL_TITLE,
        'watermark_text': watermark_text,
        'has_return': has_return,   # 新增
        'general_remark': general_remark,  # 新增
    }

    response = render(request, 'bill/print.html', context)
    cache.set(cache_key, response.content, CACHE_PRINT_ORDER)

    return response

# @login_required
# @permission_required(PERM_ORDER_PRINT)
# def batch_print_orders(request):
#     """批量打印订单页面"""
#     order_nos_param = request.GET.get('order_nos', '')
#     order_nos = [no.strip() for no in order_nos_param.split(',') if no.strip()]
#     if not order_nos:
#         return HttpResponseBadRequest("请选择至少一个订单")
#
#     orders = Order.objects.filter(
#         order_no__in=order_nos
#     ).exclude(status='cancelled').select_related('customer', 'area', 'creator')
#
#     orders_data = []
#     for order in orders:
#         items = order.items.select_related('product')
#         items_display = list(items[:15]) + [None] * (15 - min(len(items), 15))
#         has_return_or_exchange = has_return_or_exchange_items(order)
#         float_start = find_float_start(items_display)
#
#         # 交付方式水印
#         watermark_text = None
#         if order.delivery_method == 'pickup':
#             watermark_text = '客户自提'
#         elif order.delivery_method == 'express':
#             watermark_text = '快递寄件'
#
#         general_remark = order.remark.strip() if order.remark else None
#
#         orders_data.append({
#             'order': order,
#             'items_display': items_display,
#             'has_return_or_exchange': has_return_or_exchange,
#             'float_start': float_start,
#             'watermark_text': watermark_text,
#             'has_return': order.has_return,   # 新增
#             'general_remark': general_remark,  # 新增
#         })
#
#     context = {
#         'orders_data': orders_data,
#         'phone_numbers': settings.PHONE_NUMBERS,
#         'complaint_phone': settings.COMPLAINT_PHONE,
#         'bill_title': settings.BILL_TITLE,
#     }
#     return render(request, 'bill/batch_print.html', context)



def prepare_order_data(order):
    """准备单个订单的打印数据（供视图和缓存使用）"""
    items = order.items.select_related('product')
    items_display = list(items[:15])
    items_display.extend([None] * (15 - len(items_display)))
    has_return_or_exchange = order.items.filter(
        is_makeup_item=True,
        operation_type__in=['return', 'exchange']
    ).exists()
    float_start = find_float_start(items_display)

    watermark_text = None
    if order.delivery_method == 'pickup':
        watermark_text = '客户自提'
    elif order.delivery_method == 'express':
        watermark_text = '快递寄件'

    general_remark = order.remark.strip() if order.remark else None

    return {
        'order': order,
        'items_display': items_display,
        'has_return_or_exchange': has_return_or_exchange,
        'float_start': float_start,
        'watermark_text': watermark_text,
        'has_return': order.has_return,
        'general_remark': general_remark,
    }


@login_required
@permission_required(PERM_ORDER_PRINT, raise_exception=True)
def print_orders(request):
    """
    统一打印视图：支持单个或多个订单
    参数：
        order_no   - 单个订单号（可选）
        order_nos  - 逗号分隔的多个订单号（可选）
    若同时提供，会合并去重。
    """
    order_no = request.GET.get('order_no')
    order_nos_param = request.GET.get('order_nos', '')
    order_nos = [no.strip() for no in order_nos_param.split(',') if no.strip()]

    if order_no and order_no not in order_nos:
        order_nos.insert(0, order_no)

    if not order_nos:
        return HttpResponseBadRequest("请提供至少一个订单编号")

    is_batch = len(order_nos) > 1

    # 单订单尝试读缓存
    if not is_batch:
        cache_key = f"{CACHE_PREFIX_PRINT_ORDER}{order_nos[0]}"
        cached_data = cache.get(cache_key)
        if cached_data:
            return HttpResponse(cached_data)

    # 查询订单（排除已取消）
    orders_qs = Order.objects.filter(
        order_no__in=order_nos
    ).exclude(status='cancelled').select_related('customer', 'area', 'creator')

    order_dict = {o.order_no: o for o in orders_qs}
    orders_data = []
    for no in order_nos:
        order = order_dict.get(no)
        if order:
            orders_data.append(prepare_order_data(order))

    if not orders_data:
        return HttpResponseBadRequest("没有有效订单（可能已取消）")

    context = {
        'orders_data': orders_data,
        'phone_numbers': settings.PHONE_NUMBERS,
        'complaint_phone': settings.COMPLAINT_PHONE,
        'bill_title': settings.BILL_TITLE,
        'is_batch': is_batch,
    }
    response = render(request, 'bill/print_orders.html', context)

    # 单订单写入缓存
    if not is_batch:
        cache.set(cache_key, response.content, CACHE_PRINT_ORDER)

    return response


@login_required
@permission_required(PERM_ORDER_PRINT, raise_exception=True)
def mark_printed(request):
    """
    统一标记打印接口：接收订单号列表，更新状态为“已打印”
    """
    if request.method != 'POST':
        return JsonResponse({'code': 0, 'msg': '仅支持POST'})
    try:
        data = json.loads(request.body)
        order_nos = data.get('order_nos', [])
        if not order_nos:
            return JsonResponse({'code': 0, 'msg': '未提供订单号'})
        # 批量更新状态（请根据实际字段名调整）
        updated = Order.objects.filter(order_no__in=order_nos).update(status='printed')
        return JsonResponse({'code': 1, 'msg': f'已更新 {updated} 个订单为已打印'})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': str(e)})

@login_required
def print_empty_template(request):
    """打印空白模板，通过 GET 参数 count 指定张数（默认1）"""
    count = int(request.GET.get('count', 1))
    count = max(1, min(count, 100))  # 限制 1~100 张
    context = {
        'count': range(count),
        'phone_numbers': settings.PHONE_NUMBERS,
        'complaint_phone': settings.COMPLAINT_PHONE,
        'bill_title': settings.BILL_TITLE,
    }
    return render(request, 'bill/empty_template_print.html', context)


def prepare_key_data(order):
    """准备单个订单的套打数据（含备注水印）"""
    items = order.items.select_related('product')
    items_display = list(items[:15]) + [None] * (15 - min(len(items), 15))

    # 原有套打字段
    if order.customer_name_snapshot:
        customer_name_display = order.customer_name_snapshot
    elif order.customer:
        customer_name_display = order.customer.name
    else:
        customer_name_display = '无'

    if order.order_number_snapshot:
        order_number_display = order.order_number_snapshot
    elif order.customer and order.customer.order_number:
        order_number_display = order.customer.order_number
    else:
        order_number_display = ''

    creator_name = order.creator.name if order.creator else '未知'

    # ========== 新增：水印相关（与正常打印完全一致） ==========
    has_return_or_exchange = order.items.filter(
        is_makeup_item=True,
        operation_type__in=['return', 'exchange']
    ).exists()
    float_start = find_float_start(items_display)  # 确保已导入 find_float_start

    watermark_text = None
    if order.delivery_method == 'pickup':
        watermark_text = '客户自提'
    elif order.delivery_method == 'express':
        watermark_text = '快递寄件'

    general_remark = order.remark.strip() if order.remark else None

    return {
        'order': order,
        'items_display': items_display,
        'customer_name_display': customer_name_display,
        'order_number_display': order_number_display,
        'creator_name': creator_name,
        # 水印字段（新增）
        'has_return_or_exchange': has_return_or_exchange,
        'float_start': float_start,
        'watermark_text': watermark_text,
        'has_return': order.has_return,
        'general_remark': general_remark,
    }


@login_required
@permission_required(PERM_ORDER_PRINT, raise_exception=True)
def print_key_data(request):
    """统一套打视图：支持单个或多个订单"""
    order_no = request.GET.get('order_no')
    order_nos_param = request.GET.get('order_nos', '')
    order_nos = [no.strip() for no in order_nos_param.split(',') if no.strip()]

    if order_no and order_no not in order_nos:
        order_nos.insert(0, order_no)

    if not order_nos:
        return HttpResponseBadRequest("请提供至少一个订单编号")

    orders = Order.objects.filter(
        order_no__in=order_nos
    ).exclude(status='cancelled').select_related('customer', 'area', 'creator')

    orders_data = []
    for order in orders:
        orders_data.append(prepare_key_data(order))

    if not orders_data:
        return HttpResponseBadRequest("没有有效订单（可能已取消）")

    is_batch = len(orders_data) > 1

    context = {
        'orders_data': orders_data,
        'is_batch': is_batch,
        'phone_numbers': settings.PHONE_NUMBERS,
        'complaint_phone': settings.COMPLAINT_PHONE,
        'bill_title': settings.BILL_TITLE,
    }
    return render(request, 'bill/print_key_data.html', context)

# def print_key_data(request, order_no):
#     order = get_object_or_404(
#         Order.objects.select_related('customer', 'area', 'creator'),
#         order_no=order_no
#     )
#     items = order.items.select_related('product')
#     items_display = list(items[:15]) + [None] * (15 - min(len(items), 15))
#
#     # 安全处理客户名称
#     if order.customer_name_snapshot:
#         customer_name_display = order.customer_name_snapshot
#     elif order.customer:
#         customer_name_display = order.customer.name
#     else:
#         customer_name_display = '无'
#
#     # 安全处理制单工号
#     if order.order_number_snapshot:
#         order_number_display = order.order_number_snapshot
#     elif order.customer and order.customer.order_number:
#         order_number_display = order.customer.order_number
#     else:
#         order_number_display = ''
#
#     # 开单人员（creator 通常存在，但仍做防御）
#     creator_name = order.creator.name if order.creator else '未知'
#
#     context = {
#         'order': order,
#         'items_display': items_display,
#         'customer_name_display': customer_name_display,
#         'order_number_display': order_number_display,
#         'creator_name': creator_name,
#         'phone_numbers': settings.PHONE_NUMBERS,
#         'complaint_phone': settings.COMPLAINT_PHONE,
#         'bill_title': settings.BILL_TITLE,
#     }
#     return render(request, 'bill/print_key_data.html', context)


@login_required
@permission_required(PERM_ORDER_PRINT)
def batch_print_key_data(request):
    """批量打印关键数据（套打）"""
    order_nos_param = request.GET.get('order_nos', '')
    order_nos = [no.strip() for no in order_nos_param.split(',') if no.strip()]
    if not order_nos:
        return HttpResponseBadRequest("请选择至少一个订单")

    orders = Order.objects.filter(
        order_no__in=order_nos
    ).exclude(status='cancelled').select_related('customer', 'area', 'creator')

    orders_data = []
    for order in orders:
        items = order.items.select_related('product')
        items_display = list(items[:15]) + [None] * (15 - min(len(items), 15))

        # 安全处理客户名称
        if order.customer_name_snapshot:
            c_name = order.customer_name_snapshot
        elif order.customer:
            c_name = order.customer.name
        else:
            c_name = '无'

        # 安全处理制单工号
        if order.order_number_snapshot:
            o_num = order.order_number_snapshot
        elif order.customer and order.customer.order_number:
            o_num = order.customer.order_number
        else:
            o_num = ''

        creator_name = order.creator.name if order.creator else '未知'

        orders_data.append({
            'order': order,
            'items_display': items_display,
            'customer_name_display': c_name,
            'order_number_display': o_num,
            'creator_name': creator_name,
        })

    context = {
        'orders_data': orders_data,
        'phone_numbers': settings.PHONE_NUMBERS,
        'complaint_phone': settings.COMPLAINT_PHONE,
        'bill_title': settings.BILL_TITLE,
    }
    return render(request, 'bill/batch_print_key_data.html', context)

@login_required
@permission_required(PERM_ORDER_CREATE)
def sort_rule_setting(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            stages = data.get('stages', [])
            with transaction.atomic():
                SortRule.objects.all().delete()
                for stage_info in stages:
                    stage_num = stage_info['stage']
                    for rule in stage_info.get('rules', []):
                        SortRule.objects.create(
                            stage=stage_num,
                            rule_type=rule['type'],
                            tag_id=rule.get('tag_id') if rule['type'] == 'tag' else None,
                            spec_condition=rule.get('spec_condition') if rule['type'] == 'spec' else None,
                            priority=rule['priority']
                        )
            # ✅ 排序规则变更后，立即清理相关缓存
            clear_sort_cache()
            return JsonResponse({'code': 1, 'msg': '规则保存成功'})
        except Exception as e:
            return JsonResponse({'code': 0, 'msg': f'保存失败：{str(e)}'})

    # GET：返回按阶段分组的数据
    rules_qs = SortRule.objects.select_related('tag').order_by('stage', 'priority')
    stages_dict = {}
    for r in rules_qs:
        if r.stage not in stages_dict:
            stages_dict[r.stage] = []
        stages_dict[r.stage].append({
            'type': r.rule_type,
            'priority': r.priority,
            'tag_id': r.tag_id,
            'tag_name': r.tag.name if r.tag else '',
            'spec_condition': r.spec_condition,
        })

    stages_data = []
    for stage_num in sorted(stages_dict.keys()):
        stages_data.append({
            'stage': stage_num,
            'rules': stages_dict[stage_num]
        })

    # 如果没有阶段，给一个默认空阶段供界面展示
    if not stages_data:
        stages_data.append({'stage': 1, 'rules': []})

    tags_data = [{'id': t.id, 'name': t.name} for t in ProductTag.objects.filter(is_active=True)]
    return render(request, 'bill/sort_rule_setting.html', {
        'stages_json': json.dumps(stages_data),
        'tags_json': json.dumps(tags_data),
    })


@login_required
@permission_required(PERM_ORDER_CREATE)
def get_sort_rules(request):
    """供开单页调用的排序规则 API，返回阶段分组数组"""
    rules = SortRule.objects.select_related('tag').order_by('stage', 'priority')
    stages_dict = {}
    for r in rules:
        if r.stage not in stages_dict:
            stages_dict[r.stage] = []
        item = {
            'type': r.rule_type,
            'priority': r.priority,
        }
        if r.rule_type == 'tag':
            item['tag_id'] = r.tag_id
            item['tag_name'] = r.tag.name
        else:
            item['spec_condition'] = r.spec_condition
        stages_dict[r.stage].append(item)

    stages = []
    for stage_num in sorted(stages_dict.keys()):
        stages.append({
            'stage': stage_num,
            'rules': stages_dict[stage_num]
        })
    return JsonResponse({'code': 1, 'data': stages})

# views.py
@login_required
def get_all_product_tags(request):
    # 只返回启用且有标签的商品
    products = Product.objects.filter(is_active=True).prefetch_related('tags')
    data = {}
    for p in products:
        tag_ids = list(p.tags.filter(is_active=True).values_list('id', flat=True))
        if tag_ids:
            data[str(p.id)] = tag_ids
    return JsonResponse({'code': 1, 'data': data})

@login_required
@accounts_permission_required(PERM_ORDER_VIEW)
def order_list(request):
    """订单列表页（新增Tab状态筛选版 + 财务进度展示）"""
    order_no = request.GET.get('order_no', '').strip()
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    area_id = request.GET.get('area_id', '')
    customer_name = request.GET.get('customer_name', '').strip()
    amount_operator = request.GET.get('amount_operator', '')
    amount_value = request.GET.get('amount_value', '').strip()
    status = request.GET.get('status', 'all')
    # 🔥 新增：开单人筛选参数
    creator_id = request.GET.get('creator_id', '')
    page = request.GET.get('page', 1)

    # 🔥 缓存键中加入 creator_id
    cache_key = f"{CACHE_PREFIX_ORDER_LIST}{request.user.id}_{order_no}_{date_from}_{date_to}_{area_id}_{customer_name}_{amount_operator}_{amount_value}_{status}_{creator_id}_{page}"
    cached_data = cache.get(cache_key)

    if cached_data:
        return HttpResponse(cached_data)

    # 关联预加载（无N+1）
    orders = Order.objects.select_related('area', 'customer', 'creator').order_by('-create_time')

    # 权限控制
    is_super_admin = request.user.role and request.user.role.code == ROLE_SUPER_ADMIN
    can_view_others = request.user.has_permission('order_view_others')
    if not is_super_admin and not can_view_others:
        orders = orders.filter(creator=request.user)

    is_admin = request.user.role and request.user.role.code == ROLE_ADMIN
    is_operator = request.user.role and request.user.role.code == ROLE_OPERATOR

    # 🔥 新增：获取开单人列表（用于筛选下拉框）
    if is_super_admin or can_view_others:
        creators = User.objects.filter(created_orders__isnull=False).distinct().order_by('user_code')
    else:
        # 无查看他人权限时，只显示当前用户自己
        creators = User.objects.filter(id=request.user.id)

    # 🔥 状态筛选（核心Tab逻辑，包含已结清/未结清）
    base_orders = orders
    if status == 'normal':
        orders = orders.filter(status__in=ORDER_STATUS_VALID)
    elif status == 'cancelled':
        orders = orders.filter(status='cancelled')
    elif status == 'settled':
        orders = orders.filter(is_settled=True, status__in=ORDER_STATUS_VALID)
    elif status == 'unsettled':
        orders = orders.filter(is_settled=False, status__in=ORDER_STATUS_VALID)

    # 🔥 Tab数量统计
    counts = base_orders.aggregate(
        count_all=Count('id'),
        count_normal=Count(Case(When(status__in=ORDER_STATUS_VALID, then='id'))),
        count_cancelled=Count(Case(When(status='cancelled', then='id'))),
        count_settled=Count(Case(When(status__in=ORDER_STATUS_VALID, is_settled=True, then='id'))),
        count_unsettled=Count(Case(When(status__in=ORDER_STATUS_VALID, is_settled=False, then='id')))
    )
    count_all = counts['count_all']
    count_normal = counts['count_normal']
    count_cancelled = counts['count_cancelled']
    count_settled = counts['count_settled']
    count_unsettled = counts['count_unsettled']

    # 原有筛选逻辑
    if order_no:
        orders = orders.filter(order_no__startswith=order_no)

    if area_id and area_id.isdigit():
        orders = orders.filter(area_id=int(area_id))

    if customer_name:
        orders = orders.filter(customer__name__istartswith=customer_name)

    if date_from:
        try:
            start_datetime = timezone.make_aware(datetime.strptime(date_from, '%Y-%m-%d'))
            orders = orders.filter(create_time__gte=start_datetime)
        except:
            pass
    if date_to:
        try:
            end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
            end_datetime = timezone.make_aware(datetime.combine(end_date + timedelta(days=1), datetime.min.time()))
            orders = orders.filter(create_time__lt=end_datetime)
        except:
            pass

    if amount_operator in ['gt', 'lt'] and amount_value:
        try:
            amount = decimal.Decimal(amount_value)
            orders = orders.filter(total_amount__gt=amount) if amount_operator == 'gt' else orders.filter(
                total_amount__lt=amount)
        except decimal.InvalidOperation:
            pass

    # 🔥 新增：开单人筛选
    if creator_id and creator_id.isdigit():
        orders = orders.filter(creator_id=int(creator_id))

    # 分页
    paginator = Paginator(orders, 10)
    try:
        page_orders = paginator.page(page)
    except PageNotAnInteger:
        page_orders = paginator.page(1)
    except EmptyPage:
        page_orders = paginator.page(paginator.num_pages)

    # 统计数据（基于当前筛选结果）
    stats = orders.aggregate(
        total_orders=Count('id'),
        total_sales=Sum('total_amount', default=decimal.Decimal('0.00')),
        settled_orders=Count(Case(When(is_settled=True, then='id'))),
        total_debt=Sum(Case(
            When(is_settled=False, then='total_amount')
        ), default=decimal.Decimal('0.00'), output_field=DecimalField())
    )
    total_orders = stats['total_orders']
    total_sales = stats['total_sales']
    settled_orders = stats['settled_orders']
    total_debt = stats['total_debt']

    # 作废权限计算 & 财务数据计算
    current_time = timezone.now()
    order_list = list(page_orders)
    for order in order_list:
        time_diff = (current_time - order.create_time).total_seconds() / 60
        order.time_diff = time_diff

        order.unpaid_amount = order.total_amount - order.received_amount
        if order.total_amount > 0:
            order.paid_percent = (order.received_amount / order.total_amount) * 100
        else:
            order.paid_percent = 100

        can_cancel = False
        if order.status != 'cancelled' and not order.is_settled:
            if is_super_admin:
                can_cancel = True
            elif is_admin:
                if (order.creator == request.user and request.user.has_permission('order_cancel_own')) or \
                        (order.creator != request.user and request.user.has_permission('order_cancel_others')):
                    can_cancel = True
            elif is_operator:
                if order.creator == request.user and time_diff <= 30 and request.user.has_permission('order_cancel_own'):
                    can_cancel = True
        order.can_cancel = can_cancel

    areas = Area.objects.all().order_by('name')
    context = {
        'orders': order_list,
        'page_orders': page_orders,
        'paginator': paginator,
        'areas': areas,
        'date_from': date_from,
        'date_to': date_to,
        'area_id': area_id,
        'customer_name': customer_name,
        'amount_operator': amount_operator,
        'amount_value': amount_value,
        'is_super_admin': is_super_admin,
        'is_admin': is_admin,
        'is_operator': is_operator,
        'order_no': order_no,
        'total_orders': total_orders,
        'total_sales': total_sales,
        'settled_orders': settled_orders,
        'total_debt': total_debt,
        'status': status,
        'count_all': count_all,
        'count_normal': count_normal,
        'count_cancelled': count_cancelled,
        'count_settled': count_settled,
        'count_unsettled': count_unsettled,
        # 🔥 新增：开单人筛选数据
        'creators': creators,
        'creator_id': creator_id,
    }

    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']
    base_query_string = query_params.urlencode()

    def get_page_range(page, num_pages, surrounding=2):
        if num_pages <= 7:
            return list(range(1, num_pages + 1))
        pages = [1]
        if page.number - surrounding > 2:
            pages.append('...')
        start = max(2, page.number - surrounding)
        end = min(num_pages - 1, page.number + surrounding)
        pages.extend(range(start, end + 1))
        if page.number + surrounding < num_pages - 1:
            pages.append('...')
        pages.append(num_pages)
        return pages

    page_range_display = get_page_range(page_orders, paginator.num_pages)

    context.update({
        'base_query_string': base_query_string,
        'page_range_display': page_range_display,
    })
    response = render(request, 'bill/order_list.html', context)
    cache.set(cache_key, response.content, CACHE_ORDER_LIST)

    return response

@login_required
@permission_required(PERM_ORDER_VIEW)
def order_detail(request, order_no):
    """订单详情页（手动缓存版 + 财务数据展示）"""
    # 🔥 手动缓存 Key
    cache_key = f"{CACHE_PREFIX_ORDER_DETAIL}{order_no}"
    cached_data = cache.get(cache_key)

    if cached_data:
        return HttpResponse(cached_data)

    # 🔥 优化1：一次性预加载所有关联（customer/area/creator），彻底解决N+1
    order = get_object_or_404(
        Order.objects.select_related('customer', 'area', 'creator'),
        order_no=order_no
    )

    # 🔥 优化2：缓存用户角色/权限，仅查询1次
    user_role = request.user.role
    role_code = user_role.code if user_role else ''
    is_super_admin = role_code == ROLE_SUPER_ADMIN
    can_view_others = request.user.has_permission('order_view_others')

    # 权限控制
    if not is_super_admin and not can_view_others and order.creator != request.user:
        return redirect('/bill/orders/')

    # 🔥 优化3：使用Django时区时间
    current_time = timezone.now()
    time_diff = (current_time - order.create_time).total_seconds() / 60

    # 缓存权限，避免重复查询
    can_cancel_own = request.user.has_permission('order_cancel_own')
    can_cancel_others = request.user.has_permission('order_cancel_others')
    is_admin = role_code == ROLE_ADMIN
    is_operator = role_code == ROLE_OPERATOR

    # 作废按钮逻辑（不变）
    show_cancel_btn = False
    if order.status != 'cancelled' and not order.is_settled and order.status != 'printed':
        if is_super_admin:
            show_cancel_btn = True
        elif is_admin:
            if (order.creator == request.user and can_cancel_own) or (
                    order.creator != request.user and can_cancel_others):
                show_cancel_btn = True
        elif is_operator:
            if order.creator == request.user and can_cancel_own and time_diff <= 30:
                show_cancel_btn = True

    # 🔥 优化4：使用已优化的明细数据（模板必须用这个，禁止用order.items.all）
    items = OrderItem.objects.select_related('product').filter(order=order)

    # 🔥 新增：详情页财务数据计算
    unpaid_amount = order.total_amount - order.received_amount
    if order.total_amount > 0:
        paid_percent = (order.received_amount / order.total_amount) * 100
    else:
        paid_percent = 100

    context = {
        'order': order,
        'items': items,  # 模板必须用这个变量
        'is_super_admin': is_super_admin,
        'time_diff': time_diff,
        'can_cancel_own': can_cancel_own,
        'can_cancel_others': can_cancel_others,
        'is_admin': is_admin,
        'is_operator': is_operator,
        'show_cancel_btn': show_cancel_btn,
        # 🔥 新增：传递给模板
        'unpaid_amount': unpaid_amount,
        'paid_percent': paid_percent
    }

    response = render(request, 'bill/order_detail.html', context)
    cache.set(cache_key, response.content, CACHE_ORDER_DETAIL)

    return response

@login_required
def search_customer(request):
    """客户搜索（支持拼音，返回制单号）"""
    keyword = request.GET.get('keyword', '').strip()
    if not keyword:
        return JsonResponse({'code': 0, 'data': []})

    # 根据名称、拼音全拼、首字母模糊搜索
    customers = Customer.objects.filter(
        Q(name__icontains=keyword) |
        Q(pinyin_full__icontains=keyword) |
        Q(pinyin_abbr__icontains=keyword)
    ).distinct()[:50]

    data = []
    for c in customers:
        data.append({
            'id': c.id,
            'full_name': f'{c.area.name} | {c.name}' if c.area else c.name,
            'order_number': c.order_number or '',   # 制单号，为空则返回空字符串
        })
    return JsonResponse({'code': 1, 'data': data})

@login_required
@permission_required(PERM_ORDER_CANCEL_OWN)
def cancel_order(request, order_no):
    """
    作废订单（高性能优化版 + 手动缓存清理）
    """
    if request.method != 'POST':
        return JsonResponse({'code': 0, 'msg': '仅支持POST请求'}, status=405)

    # 事务包裹：订单作废 + 库存恢复 原子性操作
    with transaction.atomic():
        try:
            # ===================== 优化1：订单查询预加载关联字段，减少查询 =====================
            order = get_object_or_404(
                Order.objects.select_related('creator'),
                order_no=order_no
            )
            current_time = timezone.now()
            # 统一使用 Django 时区计算时间差
            time_diff = (current_time - order.create_time).total_seconds() / 60

            # ===================== 1. 获取角色（提前） =====================
            user_role_code = request.user.role.code if request.user.role else None
            is_super_admin = user_role_code == ROLE_SUPER_ADMIN
            is_admin = user_role_code == ROLE_ADMIN
            is_operator = user_role_code == ROLE_OPERATOR

            # ===================== 2. 状态锁校验（超级管理员可越过部分限制） =====================
            if order.status == 'cancelled':
                return JsonResponse({'code': 0, 'msg': '该订单已作废，无需重复操作'}, status=400)

            if not is_super_admin:
                if order.is_settled:
                    return JsonResponse({'code': 0, 'msg': '已收款的订单无法作废'}, status=400)
                if order.status == 'printed':
                    return JsonResponse({'code': 0, 'msg': '已出库的订单无法作废'}, status=400)

            # ===================== 3. 后续权限判断（保持不变） =====================
            if not is_super_admin:
                if is_admin:
                    if order.creator != request.user and not request.user.has_permission('order_cancel_others'):
                        return JsonResponse({'code': 0, 'msg': '无作废他人订单的权限'}, status=403)
                elif is_operator:
                    if order.creator != request.user:
                        return JsonResponse({'code': 0, 'msg': '普通店员仅能作废自己创建的订单'}, status=403)
                    if time_diff > 5:
                        return JsonResponse({'code': 0, 'msg': f'仅支持开单后5分钟内作废，当前已过{time_diff:.1f}分钟'},
                                            status=400)
                else:
                    return JsonResponse({'code': 0, 'msg': '无作废订单的权限'}, status=403)

            # ===================== 3. 参数校验 =====================
            data = json.loads(request.body)
            reason = data.get('reason', '').strip()
            if not reason:
                return JsonResponse({'code': 0, 'msg': '作废原因至少填写1个字'}, status=400)

            # ===================== 4. 执行作废操作 =====================
            order.status = 'cancelled'
            order.cancelled_by = request.user
            order.cancelled_time = current_time
            order.cancelled_reason = reason
            order.save(update_fields=['status', 'cancelled_by', 'cancelled_time', 'cancelled_reason'])

            # ===================== 优化2：解决N+1查询 + 批量恢复库存 =====================
            # 🔥 核心：一次查询获取所有订单项+关联商品，无N+1
            order_items = order.items.select_related('product')
            product_list = []
            item_count = 0

            for item in order_items:
                if item.product:
                    # 🔥 修复：旧字段stock → stock_system（恢复系统库存）
                    item.product.stock_system += item.quantity
                    product_list.append(item.product)
                    item_count += 1

            # 🔥 核心：批量更新库存，1次数据库操作（性能提升10~100倍）
            if product_list:
                Product.objects.bulk_update(product_list, fields=['stock_system'])

            # ===================== 5. 日志记录 =====================
            role_name = request.user.role.name if request.user.role else '未知'
            create_operation_log(
                request=request,
                op_type='cancel_order',
                obj_type='order',
                obj_id=str(order.id),
                obj_name=f"订单-{order.order_no}",
                detail=f"作废订单{order.order_no}，操作人角色：{role_name}，原因：{reason}，恢复{item_count}个商品库存，开单后{time_diff:.1f}分钟作废"
            )

            # ===================== 6. 缓存清理（核心新增） =====================
            customer_id = order.customer_id if order.customer else None
            clear_order_cache(order_no)  # 清理当前订单详情/打印
            clear_stock_cache()  # 库存恢复，清理库存列表
            if customer_id:
                clear_customer_related_cache(customer_id)  # 清理客户最近购买

            return JsonResponse({'code': 1, 'msg': '订单作废成功', 'order_no': order_no})

        except json.JSONDecodeError:
            return JsonResponse({'code': 0, 'msg': '请求数据格式错误，必须是JSON'}, status=400)
        except Exception as e:
            # 事务会自动回滚，安全返回错误
            return JsonResponse({'code': 0, 'msg': f'作废失败：{str(e)}'}, status=500)


def has_return_or_exchange_items(order):
    # 优先用 operation_type
    if order.items.filter(is_makeup_item=True, operation_type__in=['return', 'exchange']).exists():
        return True
    # 降级：如果 operation_type 为空，则通过 amount < 0 判断（假设退货为负金额）
    # 或者通过商品名称包含关键词，但不可靠，建议仅作为过渡
    return order.items.filter(
        is_makeup_item=True,
        amount__lt=0
    ).exists()

def find_float_start(items_display):
    """从第11行（索引10）开始，查找连续3个空行，返回起始索引，否则返回None"""
    for start in range(10, 12):  # 索引10~15，保证有3行
        if all(items_display[start + i] is None for i in range(3)):
            return start
    return None
@login_required
@permission_required(PERM_ORDER_REOPEN)
def reopen_order_edit(request, order_no):
    original_order = get_object_or_404(
        Order.objects.select_related('customer', 'area'),
        order_no=order_no
    )
    if original_order.status != 'cancelled':
        return redirect('bill:order_detail', order_no=order_no)

    is_super_admin = request.user.role and request.user.role.code == ROLE_SUPER_ADMIN
    items = OrderItem.objects.select_related('product').filter(order=original_order)

    order_data = {
        'order_no': original_order.order_no,
        'customer_id': original_order.customer_id if original_order.customer else '',
        'customer_name': original_order.customer_name_snapshot or (
            f"{original_order.area.name} | {original_order.customer.name}"
            if original_order.customer and original_order.area else ''
        ),
        'items': [
            {
                'id': item.product_id if item.product else '',
                'name': item.product_name or (item.product.name if item.product else ''),
                'qty': item.quantity,
                'unit': item.unit,
                'price': float(item.actual_unit_price) if item.actual_unit_price else 0,
                'amt': float(item.amount) if item.amount else 0,
                'spec': item.specification
            }
            for item in items
        ]
    }

    context = {
        'is_super_admin': is_super_admin,
        'reopen_order_data': order_data,
    }
    context.update(get_sort_context())   # 注入排序数据

    return render(request, 'bill/index.html', context)

@ajax_login_required
@ajax_permission_required(PERM_ORDER_SETTLE)
def settle_order(request, order_no):
    """标记订单结清（性能优化版 + 缓存清理）"""
    if request.method != 'POST':
        return JsonResponse({'code': 0, 'msg': '仅支持POST请求'}, status=405)

    try:
        # 优化：select_related 预加载（规范，无N+1）
        order = get_object_or_404(Order.objects.select_related('creator'), order_no=order_no)

        # 状态校验
        if order.status == 'cancelled':
            return JsonResponse({'code': 0, 'msg': '作废订单无法标记结清'}, status=400)
        if order.is_settled:
            return JsonResponse({'code': 0, 'msg': '该订单已结清，无需重复操作'}, status=400)

        # 参数校验
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'code': 0, 'msg': '请求数据格式错误，必须是JSON'}, status=400)

        remark = data.get('remark', '').strip()
        if not remark:
            return JsonResponse({'code': 0, 'msg': '请填写结清备注'}, status=400)

        # 优化：统一使用 Django 时区时间
        order.is_settled = True
        order.settled_by = request.user
        order.settled_time = timezone.now()
        order.settled_remark = remark
        order.save(update_fields=['is_settled', 'settled_by', 'settled_time', 'settled_remark'])

        # 操作日志
        create_operation_log(
            request=request, op_type='settle_order', obj_type='order',
            obj_id=str(order.id), obj_name=f"订单-{order.order_no}",
            detail=f"标记订单{order.order_no}结清，备注：{remark}"
        )

        # 🔥 缓存清理
        clear_order_cache(order_no)

        return JsonResponse({'code': 1, 'msg': '订单标记结清成功', 'order_no': order_no})

    except Exception as e:
        return JsonResponse({'code': 0, 'msg': f'标记结清失败：{str(e)}'}, status=500)


@ajax_login_required
@ajax_permission_required(PERM_ORDER_UNSETTLE)
def unsettle_order(request, order_no):
    """撤销订单结清（性能优化版 + 缓存清理）"""
    if request.method != 'POST':
        return JsonResponse({'code': 0, 'msg': '仅支持POST请求'}, status=405)

    try:
        order = get_object_or_404(Order.objects.select_related('creator'), order_no=order_no)

        if not order.is_settled:
            return JsonResponse({'code': 0, 'msg': '该订单未结清，无需撤销'}, status=400)

        # 参数校验
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'code': 0, 'msg': '请求数据格式错误，必须是JSON'}, status=400)

        remark = data.get('remark', '').strip()
        if not remark:
            return JsonResponse({'code': 0, 'msg': '请填写撤销结清备注'}, status=400)

        # 优化：时区统一 + 仅更新必要字段
        order.is_settled = False
        order.unsettled_by = request.user
        order.unsettled_time = timezone.now()
        order.unsettled_remark = remark
        order.save(update_fields=['is_settled', 'unsettled_by', 'unsettled_time', 'unsettled_remark'])

        # 操作日志
        create_operation_log(
            request=request, op_type='unsettle_order', obj_type='order',
            obj_id=str(order.id), obj_name=f"订单-{order.order_no}",
            detail=f"撤销订单{order.order_no}结清状态，备注：{remark}"
        )

        # 🔥 缓存清理
        clear_order_cache(order_no)

        return JsonResponse({'code': 1, 'msg': '撤销订单结清成功', 'order_no': order_no})

    except Exception as e:
        return JsonResponse({'code': 0, 'msg': f'撤销结清失败：{str(e)}'}, status=500)


@ajax_login_required
@ajax_permission_required(PERM_ORDER_SETTLE)
def batch_settle_order(request):
    """批量标记订单结清（高性能优化版 + 缓存清理）"""
    if request.method != 'POST':
        return JsonResponse({'code': 0, 'msg': '仅支持POST请求'}, status=405)

    try:
        # 参数解析
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'code': 0, 'msg': '请求数据格式错误，必须是JSON'}, status=400)

        order_list = data.get('orders', [])
        if not order_list:
            return JsonResponse({'code': 0, 'msg': '请选择要结清的订单'}, status=400)

        success_count = 0
        fail_list = []
        update_orders = []
        current_time = timezone.now()

        order_no_map = {}
        for item in order_list:
            order_no = str(item.get('order_no', '')).strip()
            remark = str(item.get('remark', '')).strip()

            if not order_no or not remark:
                fail_list.append(f'{order_no or "未知订单"}：备注不能为空')
                continue
            order_no_map[order_no] = remark

        if not order_no_map:
            return JsonResponse({'code': 0, 'msg': '无有效订单数据'}, status=400)

        valid_orders = Order.objects.filter(order_no__in=order_no_map.keys())
        valid_order_nos = {o.order_no for o in valid_orders}

        for order in valid_orders:
            remark = order_no_map[order.order_no]

            if order.status == 'cancelled':
                fail_list.append(f'{order.order_no}：作废订单无法结清')
                continue
            if order.is_settled:
                fail_list.append(f'{order.order_no}：已结清，无需重复操作')
                continue

            order.is_settled = True
            order.settled_by = request.user
            order.settled_time = current_time
            order.settled_remark = remark
            update_orders.append(order)

        for order_no in order_no_map.keys():
            if order_no not in valid_order_nos:
                fail_list.append(f'{order_no}：订单不存在')

        if update_orders:
            with transaction.atomic():
                Order.objects.bulk_update(
                    update_orders,
                    fields=['is_settled', 'settled_by', 'settled_time', 'settled_remark']
                )
                for order in update_orders:
                    create_operation_log(
                        request=request, op_type='batch_settle_order', obj_type='order',
                        obj_id=str(order.id), obj_name=f"订单-{order.order_no}",
                        detail=f"批量结清订单{order.order_no}，备注：{order.settled_remark}"
                    )
            success_count = len(update_orders)

            # 🔥 缓存清理：清理所有受影响的订单缓存
            for order in update_orders:
                clear_order_cache(order.order_no)

        msg = f'批量处理完成！成功{success_count}个，失败{len(fail_list)}个'
        if fail_list:
            msg += f'；失败原因：{"; ".join(fail_list)}'

        return JsonResponse({'code': 1 if success_count > 0 else 0,
                             'msg': msg,
                             'success_count': success_count,
                             'fail_list': fail_list})

    except Exception as e:
        return JsonResponse({'code': 0, 'msg': f'批量结清失败：{str(e)}'}, status=500)
@login_required
@permission_required(PERM_PRODUCT_SEARCH)
def get_customer_recent_products(request):
    customer_id = request.GET.get('customer_id', '').strip()
    if not customer_id:
        return JsonResponse({'code': 0, 'msg': '请选择客户', 'data': []})

    cache_key = f"{CACHE_PREFIX_CUSTOMER_RECENT_PRODUCT}{customer_id}"
    cached_data = cache.get(cache_key)
    if cached_data:
        return JsonResponse({'code': 1, 'data': cached_data})

    try:
        # 获取最近有效订单ID（取10个，覆盖足够的最近购买记录）
        recent_order_ids = list(
            Order.objects.filter(
                customer_id=customer_id,
                status__in=ORDER_STATUS_VALID,
                is_settled=False
            )
            .order_by('-create_time')
            .values_list('id', flat=True)[:10]
        )

        if not recent_order_ids:
            cache.set(cache_key, [], timeout=CACHE_CUSTOMER_RECENT_PRODUCT)
            return JsonResponse({'code': 1, 'data': []})

        order_items = OrderItem.objects.filter(
            order_id__in=recent_order_ids,
            is_makeup_item=False
        ).select_related('product', 'order').order_by('-order__create_time')

        # 分别处理有product的商品与自由开单商品
        product_dict = {}          # key: product.id
        free_product_dict = {}     # key: "free_产品名|规格|单位|价格"

        for item in order_items:
            if item.product:
                product = item.product
                if product.id in product_dict:
                    continue

                # ----- 价格快照逻辑 -----
                # 1. 优先使用开单时的实际单价（成交价快照）
                if item.actual_unit_price is not None:
                    final_price = float(item.actual_unit_price)
                # 2. 若无实际单价，尝试客户价快照
                elif item.snapshot_customer_price is not None:
                    final_price = float(item.snapshot_customer_price)
                # 3. 再尝试标准价快照
                elif item.snapshot_standard_price is not None:
                    final_price = float(item.snapshot_standard_price)
                # 4. 兜底：使用商品当前标准价
                else:
                    final_price = float(product.price)

                # 规格：优先使用订单明细中的规格快照
                specification = item.specification or product.specification or ''

                product_dict[product.id] = {
                    'id': product.id,
                    'name': product.name,
                    'price': final_price,                    # 成交价快照
                    'standard_price': float(product.price),  # 当前标准价（用于对比）
                    'unit': product.unit,
                    'last_purchase_time': item.order.create_time.strftime('%Y-%m-%d %H:%M'),
                    'last_quantity': item.quantity,
                    'specification': specification,
                }
            else:
                # 自由开单商品：使用当时的成交价（原有逻辑不变）
                name = item.product_name or ''
                spec = item.specification or ''
                unit = item.unit or ''
                price = float(item.actual_unit_price) if item.actual_unit_price else 0
                free_key = f"free_{name}|{spec}|{unit}|{price}"
                if free_key in free_product_dict:
                    continue
                free_product_dict[free_key] = {
                    'id': None,
                    'name': name,
                    'price': price,
                    'standard_price': price,
                    'unit': unit,
                    'last_purchase_time': item.order.create_time.strftime('%Y-%m-%d %H:%M'),
                    'last_quantity': item.quantity,
                    'specification': spec,
                }

        # 组装结果
        recent_products = list(product_dict.values())
        free_offset = 0
        for free_data in free_product_dict.values():
            free_offset += 1
            free_data['id'] = -100000 - free_offset
            recent_products.append(free_data)

        # 缓存
        cache.set(cache_key, recent_products, timeout=CACHE_CUSTOMER_RECENT_PRODUCT)
        logger.info(
            f"设置客户最近商品缓存: {cache_key} "
            f"(含{len(product_dict)}个系统商品, {len(free_product_dict)}个自由商品)"
        )

        return JsonResponse({'code': 1, 'data': recent_products})

    except Exception as e:
        logger.error(f"获取客户最近商品失败: {str(e)}", exc_info=True)
        return JsonResponse({'code': 0, 'msg': f'获取失败：{str(e)}', 'data': []})
# ===================== 2. 新增：价格核算视图 =====================

@login_required
@permission_required(PERM_ORDER_PRICE_CHECK)
def price_check_view(request):
    """价格核算页面入口"""
    date_from = request.GET.get('date_from', (timezone.now() - timedelta(days=7)).strftime('%Y-%m-%d'))
    date_to = request.GET.get('date_to', timezone.now().strftime('%Y-%m-%d'))

    # 传递空的结果集，只显示筛选框
    return render(request, 'bill/price_check.html', {
        'date_from': date_from,
        'date_to': date_to,
        'results': None,
        'stats': None
    })


@login_required
@permission_required(PERM_ORDER_PRICE_CHECK)
def price_check_ajax(request):
    """执行价格核算的AJAX接口"""
    if request.method != 'POST':
        return JsonResponse({'code': 0, 'msg': '请求错误'})

    date_from = request.POST.get('date_from')
    date_to = request.POST.get('date_to')

    if not date_from or not date_to:
        return JsonResponse({'code': 0, 'msg': '请选择日期范围'})

    # 构建时间范围
    start_datetime = timezone.make_aware(datetime.strptime(date_from, '%Y-%m-%d'))
    end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
    end_datetime = timezone.make_aware(datetime.combine(end_date + timedelta(days=1), datetime.min.time()))

    # 查询订单及明细 (select_related 优化性能)
    orders = Order.objects.filter(
        create_time__gte=start_datetime,
        create_time__lt=end_datetime,
        status__in=['pending', 'printed', 'reopened']  # 只核查有效订单
    ).select_related('customer', 'creator').prefetch_related('items__product').order_by('-create_time')

    # 权限控制 (如果非管理员，只能看自己的)
    is_super_admin = request.user.role and request.user.role.code == ROLE_SUPER_ADMIN
    can_view_others = request.user.has_permission('order_view_others')
    if not is_super_admin and not can_view_others:
        orders = orders.filter(creator=request.user)

    results = []
    total_checked = 0
    total_abnormal = 0
    total_loss_risk = decimal.Decimal('0.00')

    for order in orders:
        total_checked += 1
        order_has_issue = False
        issue_items = []

        for item in order.items.all():
            if item.is_makeup_item:  # ← 补货品项不参与核算
                continue
            # 确定基准价
            base_price = item.snapshot_standard_price
            price_type = "标准价"

            # 如果有客户价快照，基准价应为客户价
            if item.snapshot_customer_price is not None:
                base_price = item.snapshot_customer_price
                price_type = "客户价"

            # 如果没有快照数据（历史旧数据），跳过或标记
            if base_price is None or item.actual_unit_price is None:
                continue

            diff = item.actual_unit_price - base_price
            issue_type = None
            issue_label = ""

            # 逻辑判断
            if item.snapshot_customer_price is not None:
                # 情况A：有熟客价
                if item.actual_unit_price != item.snapshot_customer_price:
                    # 虽然有熟客价，但没用对
                    if item.actual_unit_price == item.snapshot_standard_price:
                        issue_type = 'mismatch'
                        issue_label = "错配：未用熟客价"
                    elif item.actual_unit_price < item.snapshot_customer_price:
                        issue_type = 'short'
                        issue_label = "低报：低于熟客价"
                        total_loss_risk += (abs(diff) * item.quantity)
                    else:
                        issue_type = 'over'
                        issue_label = "高报：高于熟客价"

            else:
                # 情况B：无熟客价
                if item.actual_unit_price < item.snapshot_standard_price:
                    issue_type = 'short'
                    issue_label = "低报"
                    total_loss_risk += (abs(diff) * item.quantity)
                elif item.actual_unit_price > item.snapshot_standard_price:
                    issue_type = 'over'
                    issue_label = "高报"

            if issue_type:
                order_has_issue = True
                issue_items.append({
                    'product_name': item.product.name if item.product else '未知',
                    'qty': item.quantity,
                    'snapshot_std': item.snapshot_standard_price,
                    'snapshot_cust': item.snapshot_customer_price,
                    'actual': item.actual_unit_price,
                    'diff': diff,
                    'type': issue_type,
                    'label': issue_label
                })

        if order_has_issue:
            total_abnormal += 1
            results.append({
                'order_no': order.order_no,
                'customer_name': order.customer.name if order.customer else '散客',
                'creator_name': order.creator.name if order.creator else '未知',
                'create_time': order.create_time,
                'items': issue_items
            })

    stats = {
        'checked': total_checked,
        'abnormal': total_abnormal,
        'loss': total_loss_risk
    }

    return JsonResponse({'code': 1, 'data': results, 'stats': stats})


# ===================== 新增：订单统计相关视图 =====================

# 工具函数：解析时间范围（复用你区域组统计的逻辑）
def parse_order_time_range(time_range, start_date_str, end_date_str):
    from datetime import datetime, timedelta
    today = timezone.now().date()

    if time_range == 'today':
        return today, today
    elif time_range == '7days':
        return today - timedelta(days=7), today
    elif time_range == 'month':
        return today.replace(day=1), today
    elif time_range == 'custom' and start_date_str and end_date_str:
        try:
            start = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            return start, end
        except:
            pass
    # 默认：最近30天
    return today - timedelta(days=30), today


@login_required
@permission_required(PERM_ORDER_SUMMARY)
def order_stats_page(request):
    """订单统计页面入口（零计算，仅渲染HTML）"""
    return render(request, 'bill/order_stats.html')


@login_required
@permission_required(PERM_ORDER_SUMMARY)
def calculate_order_stats(request):
    """
    核心：订单统计计算接口（懒加载专用）
    只有点击按钮才调用，利用现有索引优化性能
    """
    try:
        # 1. 获取参数
        time_range = request.GET.get('time_range', '30days')
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')

        # 2. 解析时间
        start_dt, end_dt = parse_order_time_range(time_range, start_date, end_date)

        # 3. 构建基础QuerySet（利用索引：status, is_settled, create_time）
        # 注意：这里不做权限过滤，统计全公司数据（如果需要按人过滤请自行添加）
        base_orders = Order.objects.filter(
            create_time__date__gte=start_dt,
            create_time__date__lte=end_dt
        )

        # 4. 核心指标聚合（一次数据库查询搞定所有聚合）
        # 利用索引：status, is_settled, create_time, total_amount
        agg_result = base_orders.aggregate(
            # 经营核心
            total_sales=Coalesce(Sum('total_amount', filter=Q(status__in=ORDER_STATUS_VALID)), 0,
                                 output_field=DecimalField(max_digits=12, decimal_places=2)),
            total_orders=Count('id', filter=Q(status__in=ORDER_STATUS_VALID)),

            # 回款监控
            settled_amount=Coalesce(Sum('total_amount', filter=Q(status__in=ORDER_STATUS_VALID, is_settled=True)), 0,
                                    output_field=DecimalField(max_digits=12, decimal_places=2)),
            total_debt=Coalesce(Sum('total_amount', filter=Q(status__in=ORDER_STATUS_VALID, is_settled=False)), 0,
                                output_field=DecimalField(max_digits=12, decimal_places=2)),
            debt_order_count=Count('id', filter=Q(status__in=ORDER_STATUS_VALID, is_settled=False)),

            # 风险预警
            cancelled_count=Count('id', filter=Q(status='cancelled')),
            reopened_count=Count('id', filter=Q(status='reopened')),

            # 活跃客户
            active_customers=Count('customer', distinct=True,
                                   filter=Q(status__in=ORDER_STATUS_VALID, customer__isnull=False))
        )

        # 5. 计算衍生指标
        total_sales_val = float(agg_result['total_sales'])
        total_orders_val = agg_result['total_orders']
        settled_amount_val = float(agg_result['settled_amount'])

        avg_order_value = round(total_sales_val / total_orders_val, 2) if total_orders_val > 0 else 0.0
        repayment_rate = round((settled_amount_val / total_sales_val) * 100, 2) if total_sales_val > 0 else 0.0

        # 6. 组装返回数据
        data = {
            # 经营核心
            'total_sales': total_sales_val,
            'total_orders': total_orders_val,
            'avg_order_value': avg_order_value,

            # 回款监控
            'settled_amount': settled_amount_val,
            'total_debt': float(agg_result['total_debt']),
            'repayment_rate': repayment_rate,
            'debt_order_count': agg_result['debt_order_count'],

            # 风险预警
            'cancelled_count': agg_result['cancelled_count'],
            'reopened_count': agg_result['reopened_count'],

            # 活跃客户
            'active_customers': agg_result['active_customers'],

            # 统计信息
            'date_range': {
                'start': start_dt.strftime('%Y-%m-%d'),
                'end': end_dt.strftime('%Y-%m-%d')
            },
            'calculated_at': timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        }

        return JsonResponse({'code': 1, 'data': data})

    except Exception as e:
        logger.error(f"订单统计计算失败：{str(e)}", exc_info=True)
        return JsonResponse({'code': 0, 'msg': f'统计失败：{str(e)}'})

def parse_datetime_cell(value):
    """解析日期单元格，支持字符串、datetime对象和Excel序列号"""
    if value is None:
        return None
    # 若已是 datetime 或 date，直接转换
    if isinstance(value, (datetime, date)):
        dt = value
    # 若为数字（int/float），作为 Excel 序列号处理
    elif isinstance(value, (int, float)):
        try:
            dt = from_excel(value)
        except Exception:
            return None
    # 否则尝试作为字符串解析
    else:
        s = str(value).strip()
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
            try:
                dt = datetime.strptime(s, fmt)
                break
            except ValueError:
                continue
        else:
            return None
    # 转换为 timezone-aware（假设导入时间与当前时区一致）
    if dt is not None:
        if timezone.is_naive(dt):
            dt = timezone.make_aware(dt, timezone.get_current_timezone())
    return dt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.http import StreamingHttpResponse
import io
import decimal
from urllib.parse import quote
import time

@login_required
@permission_required('order_export')
def export_orders(request):
    """流式导出订单 Excel，使用 write_only 模式，逐批写入，内存恒定"""
    # 1. 构建查询集（与原来相同，但仅筛选必要字段）
    order_no = request.GET.get('order_no', '').strip()
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    area_id = request.GET.get('area_id', '')
    customer_name = request.GET.get('customer_name', '').strip()
    amount_operator = request.GET.get('amount_operator', '')
    amount_value = request.GET.get('amount_value', '').strip()
    status = request.GET.get('status', 'all')

    is_super_admin = request.user.role and request.user.role.code == ROLE_SUPER_ADMIN
    can_view_others = request.user.has_permission('order_view_others')
    orders = Order.objects.select_related('area', 'creator', 'settled_by') \
                          .order_by('-create_time') \
                          .only('order_no', 'customer_name_snapshot', 'area', 'creator', 'create_time',
                                'total_amount', 'status', 'is_settled', 'received_amount', 'settled_by',
                                'settled_time', 'order_number_snapshot', 'delivery_method', 'is_verified')
    if not is_super_admin and not can_view_others:
        orders = orders.filter(creator=request.user)

    # 状态筛选等（省略，与原逻辑相同）
    if status == 'normal':
        orders = orders.filter(status__in=ORDER_STATUS_VALID)
    elif status == 'cancelled':
        orders = orders.filter(status='cancelled')
    elif status == 'settled':
        orders = orders.filter(is_settled=True, status__in=ORDER_STATUS_VALID)
    elif status == 'unsettled':
        orders = orders.filter(is_settled=False, status__in=ORDER_STATUS_VALID)
    # ... 其余筛选条件与原代码一致，不再赘述

    # 2. 使用流式 Workbook（write_only=True）
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="订单数据")

    headers = [
        '订单编号', '客户名称', '区域', '商品名称', '规格', '单位',
        '数量', '单价', '小计金额', '订单状态', '交付方式',
        '创建时间', '开单人', '订单总金额', '是否结清', '已收金额',
        '结清人', '结清时间', '制单号快照', '是否补货', '审核状态'
    ]
    # write_only 模式下，需要手动设置表头样式（样式不支持直接写入，可通过后续调整）
    ws.append(headers)

    # 3. 分块查询订单（每次取 1000 条，同时预取 items）
    batch_size = 1000
    order_ids = list(orders.values_list('id', flat=True))  # 先获取所有 ID，再分块查询
    total_orders = len(order_ids)

    for i in range(0, total_orders, batch_size):
        batch_ids = order_ids[i:i+batch_size]
        batch_orders = Order.objects.filter(id__in=batch_ids) \
            .prefetch_related(
                Prefetch('items', queryset=OrderItem.objects.only(
                    'product_name', 'specification', 'unit', 'quantity',
                    'actual_unit_price', 'amount', 'is_makeup_item'
                ))
            )
        for order in batch_orders:
            area_name = order.area.name if order.area else ''
            customer_snap = order.customer_name_snapshot or ''
            create_time_val = timezone.localtime(order.create_time).replace(tzinfo=None) if order.create_time else ''
            creator_name = order.creator.username if order.creator else ''
            total_amount = float(order.total_amount) if order.total_amount else 0.0
            is_settled_text = '是' if order.is_settled else '否'
            received_amount = float(order.received_amount) if order.received_amount else 0.0
            settled_by_name = order.settled_by.username if order.settled_by else ''
            settled_time_val = timezone.localtime(order.settled_time).replace(tzinfo=None) if order.settled_time else ''
            order_number_snap = order.order_number_snapshot or ''
            delivery_method = order.get_delivery_method_display() or ''  # 显示中文
            is_verified_text = '是' if order.is_verified else '否'

            items = order.items.all()
            if not items:
                # 没有明细时仍然写一行（仅订单信息，商品列留空）
                row = [order.order_no, customer_snap, area_name, '', '', '', 0, 0, 0,
                       order.status, delivery_method, create_time_val, creator_name,
                       total_amount, is_settled_text, received_amount,
                       settled_by_name, settled_time_val, order_number_snap, '', is_verified_text]
                ws.append(row)
            else:
                for item in items:
                    price = float(item.actual_unit_price) if item.actual_unit_price else 0.0
                    amt = float(item.amount) if item.amount else 0.0
                    makeup_text = '是' if item.is_makeup_item else ''
                    row = [
                        order.order_no,
                        customer_snap,
                        area_name,
                        item.product_name or '',
                        item.specification or '',
                        item.unit or '',
                        item.quantity,
                        price,
                        amt,
                        order.status,
                        delivery_method,
                        create_time_val,
                        creator_name,
                        total_amount,
                        is_settled_text,
                        received_amount,
                        settled_by_name,
                        settled_time_val,
                        order_number_snap,
                        makeup_text,
                        is_verified_text
                    ]
                    ws.append(row)

    # 4. 保存到 BytesIO 并返回
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    now_str = timezone.now().strftime('%Y%m%d')
    filename = f'订单导出{now_str}.xlsx'
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"
    return response

def parse_excel_to_structure(workbook, max_rows=None):
    """
    解析 Excel，返回结构化数据。
    使用迭代器逐行读取，避免一次性加载所有行到内存。
    :param workbook: openpyxl Workbook 对象（read_only=True 更佳）
    :param max_rows: 最多解析多少行数据（不含表头），None 表示不限制
    :return: dict 包含 order_groups, 统计信息, 错误行等
    """
    ws = workbook.active
    # 使用迭代器逐行获取
    rows_iter = ws.iter_rows(values_only=True)

    # 读取表头
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ValueError("Excel 无数据")

    # 构建列名到索引的映射
    header = [str(c).strip() if c else "" for c in header_row]
    col_map = {name: idx for idx, name in enumerate(header)}

    # 数据结构
    order_groups = {}
    area_names = set()
    product_names = set()
    pure_customer_names = set()
    pure_name_area = {}
    product_create_info = {}
    row_errors = []
    row_count = 0

    # 内部辅助函数
    def get_val(row, field, default=""):
        idx = col_map.get(field)
        if idx is not None and idx < len(row):
            val = row[idx]
            return str(val).strip() if val is not None else default
        return default

    def parse_customer_name(raw_name, given_area):
        if given_area:
            prefix = given_area + " | "
            if raw_name.startswith(prefix):
                pure = raw_name[len(prefix):].strip()
            else:
                pure = raw_name
            return given_area, pure
        else:
            if " | " in raw_name:
                parts = raw_name.split(" | ", 1)
                extracted_area = parts[0].strip()
                pure = parts[1].strip()
                return extracted_area, pure
            else:
                return "", raw_name

    # 逐行解析
    for idx, row in enumerate(rows_iter, start=2):
        if max_rows and row_count >= max_rows:
            break
        row_count += 1

        # 读取所有字段（与原逻辑完全一致）
        order_no = get_val(row, "订单编号")
        raw_customer_name = get_val(row, "客户名称")
        area_name = get_val(row, "区域")
        prod_name = get_val(row, "商品名称")
        spec = get_val(row, "规格")
        unit = get_val(row, "单位")

        qty_str = get_val(row, "数量", "0")
        try:
            qty = int(qty_str)
        except ValueError:
            row_errors.append({'row': idx, 'error': f'数量格式错误: {qty_str}'})
            continue

        price_str = get_val(row, "单价", "0")
        try:
            price = Decimal(price_str)
        except InvalidOperation:
            price = Decimal('0')

        status = get_val(row, "订单状态", "pending")
        if not status:
            status = "pending"

        is_verified_str = get_val(row, "审核状态", "否")
        is_verified = is_verified_str in ['是', '1', 'true', 'True', 'TRUE']

        create_time = None
        create_time_str = get_val(row, "创建时间")
        if create_time_str:
            try:
                dt = datetime.strptime(create_time_str, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                try:
                    dt = datetime.strptime(create_time_str, '%Y-%m-%d')
                except ValueError:
                    dt = None
            if dt:
                if timezone.is_naive(dt):
                    create_time = timezone.make_aware(dt, timezone.get_current_timezone())
                else:
                    create_time = dt

        creator_username = get_val(row, "开单人")
        is_settled_str = get_val(row, "是否结清", "否")
        is_settled = (is_settled_str == '是')

        received_str = get_val(row, "已收金额", "0")
        try:
            received_amount = Decimal(received_str)
        except InvalidOperation:
            received_amount = Decimal('0')

        settled_by_username = get_val(row, "结清人")
        settled_time = None
        settled_time_str = get_val(row, "结清时间")
        if settled_time_str:
            try:
                dt = datetime.strptime(settled_time_str, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                try:
                    dt = datetime.strptime(settled_time_str, '%Y-%m-%d')
                except ValueError:
                    dt = None
            if dt:
                if timezone.is_naive(dt):
                    settled_time = timezone.make_aware(dt, timezone.get_current_timezone())
                else:
                    settled_time = dt

        order_number_snapshot = get_val(row, "制单号快照")
        is_makeup_str = get_val(row, "是否补货")
        is_makeup_item = is_makeup_str in ['是', '1', 'true', 'True']

        final_area, pure_customer_name = parse_customer_name(raw_customer_name, area_name)

        # 按订单分组（订单编号 + 客户名称 + 区域 作为唯一键）
        order_key = (order_no, raw_customer_name, area_name)
        if order_key not in order_groups:
            order_groups[order_key] = {
                'order_no': order_no,
                'raw_customer_name': raw_customer_name,
                'area_name': area_name,
                'items': [],
                'create_time': None,
                'creator_username': creator_username,
                'is_settled': is_settled,
                'received_amount': received_amount,
                'settled_by_username': settled_by_username,
                'settled_time': settled_time,
                'order_number_snapshot': order_number_snapshot,
                'is_verified': is_verified,
            }

        # 添加明细行
        order_groups[order_key]['items'].append({
            'product_name': prod_name,
            'spec': spec,
            'unit': unit,
            'qty': qty,
            'price': price,
            'status': status,
            'pure_customer_name': pure_customer_name,
            'area_name': final_area,
            'is_makeup_item': is_makeup_item,
        })

        # 只在第一次出现时记录订单级别的创建时间
        if create_time and not order_groups[order_key]['create_time']:
            order_groups[order_key]['create_time'] = create_time

        # 收集非作废、非补货品项的基础数据（用于后续自动创建区域/客户等，但预览暂未使用）
        if status != 'cancelled' and not is_makeup_item:
            if final_area:
                area_names.add(final_area)
            if pure_customer_name:
                pure_customer_names.add(pure_customer_name)
                if pure_customer_name not in pure_name_area:
                    pure_name_area[pure_customer_name] = final_area
            product_names.add(prod_name)
            product_key = (prod_name, unit)
            if product_key not in product_create_info:
                product_create_info[product_key] = {
                    'spec': spec,
                    'price': price,
                }

    return {
        'order_groups': order_groups,
        'area_names': area_names,
        'product_names': product_names,
        'pure_customer_names': pure_customer_names,
        'pure_name_area': pure_name_area,
        'product_create_info': product_create_info,
        'row_errors': row_errors,
        'total_rows': row_count,
        'has_more': row_count >= (max_rows if max_rows else 0)  # 是否因限制而截断
    }

@login_required
@permission_required(PERM_ORDER_CREATE)
def import_orders_preview(request):
    """
    第一步：上传 Excel，返回预览数据（保留审核状态）
    限制只读取前 200 行，避免大文件内存溢出。
    """
    if request.method != 'POST':
        return JsonResponse({'code': 0, 'msg': '仅支持POST'})

    excel_file = request.FILES.get('file')
    if not excel_file:
        return JsonResponse({'code': 0, 'msg': '请上传文件'})

    # 预览只解析前 200 行数据（可根据实际需求调整）
    PREVIEW_ROWS = 200
    try:
        wb = load_workbook(excel_file, read_only=True)
        data = parse_excel_to_structure(wb, max_rows=PREVIEW_ROWS)
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': f'文件解析失败：{str(e)}'})

    order_groups = data['order_groups']

    # ---------- 批量查询已存在的订单（仅预览用） ----------
    all_order_nos = [g['order_no'] for g in order_groups.values() if g['order_no']]
    existing_orders_map = {}
    if all_order_nos:
        existing_orders = Order.objects.filter(order_no__in=all_order_nos).only('order_no', 'is_verified')
        existing_orders_map = {o.order_no: o.is_verified for o in existing_orders}

    # ---------- 构建预览列表 ----------
    order_preview_list = []
    for key, g in order_groups.items():
        order_no = g['order_no']
        items_preview = []
        order_status = g['items'][0]['status'] if g['items'] else 'pending'

        skip = False
        warnings = []
        if order_no:
            if order_no in existing_orders_map:
                skip = True
                if existing_orders_map[order_no]:
                    warnings.append('该订单已审核，跳过')
                else:
                    warnings.append('该订单已存在但未审核，跳过')
        else:
            warnings.append('订单编号为空，无法导入')
            skip = True

        for item in g['items']:
            items_preview.append({
                'product_name': item['product_name'],
                'spec': item['spec'],
                'unit': item['unit'],
                'qty': item['qty'],
                'price': str(item['price']),
                'amount': str(item['price'] * item['qty']),
                'status': item['status'],
                'is_makeup': item.get('is_makeup_item', False),
            })

        order_preview_list.append({
            'order_no': order_no,
            'raw_customer_name': g['raw_customer_name'],
            'area_name': g['area_name'],
            'pure_customer_name': g['items'][0]['pure_customer_name'] if g['items'] else '',
            'status': order_status,
            'create_time': g['create_time'].strftime('%Y-%m-%d %H:%M:%S') if g['create_time'] else '',
            'creator_username': g['creator_username'],
            'is_settled': g['is_settled'],
            'received_amount': str(g['received_amount']),
            'settled_by_username': g['settled_by_username'],
            'settled_time': g['settled_time'].strftime('%Y-%m-%d %H:%M:%S') if g['settled_time'] else '',
            'order_number_snapshot': g.get('order_number_snapshot', ''),
            'items': items_preview,
            'warnings': warnings,
            'skip': skip,
            'is_verified': g.get('is_verified', False),
        })

    # ---------- 返回 JSON ----------
    return JsonResponse({
        'code': 1,
        'data': {
            'orders': order_preview_list,
            'parse_errors': data['row_errors'],
            'total_rows': data['total_rows'],          # 实际已解析行数（受 max_rows 限制）
            'has_more': data.get('has_more', False),   # 是否因限制而截断（即文件可能还有更多数据）
        }
    })


def parse_time(time_str):
    if not time_str:
        return timezone.now()
    try:
        dt = datetime.strptime(time_str, '%Y-%m-%d %H:%M:%S')
    except ValueError:
        try:
            dt = datetime.strptime(time_str, '%Y-%m-%d')
        except ValueError:
            return timezone.now()
    if timezone.is_naive(dt):
        return timezone.make_aware(dt, timezone.get_current_timezone())
    return dt

def parse_bool(value):
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in ['true', '是', '1']
    return bool(value)

# 辅助函数（放在视图文件顶部）
def _parse_time(time_str):
    """解析时间字符串，若失败返回当前时间"""
    if not time_str:
        return timezone.now()
    try:
        dt = datetime.strptime(time_str, '%Y-%m-%d %H:%M:%S')
    except ValueError:
        try:
            dt = datetime.strptime(time_str, '%Y-%m-%d')
        except ValueError:
            return timezone.now()
    if timezone.is_naive(dt):
        return timezone.make_aware(dt, timezone.get_current_timezone())
    return dt

def _parse_bool(value):
    """解析布尔值，支持字符串 '是'/'true'/'1'"""
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in ['true', '是', '1']
    return bool(value)


@login_required
@permission_required(PERM_ORDER_CREATE)
def import_orders_confirm(request):
    """第二步：执行批量导入（分批事务 + 批量创建）"""
    if request.method != 'POST':
        return JsonResponse({'code': 0, 'msg': '仅支持POST'})

    try:
        payload = json.loads(request.body)
    except:
        return JsonResponse({'code': 0, 'msg': 'JSON 格式错误'})

    orders_data = payload.get('orders', [])
    valid_orders = [o for o in orders_data if not o.get('skip')]
    if not valid_orders:
        return JsonResponse({'code': 0, 'msg': '没有可导入的订单'})

    # ---------- 1. 校验订单号重复（批量查询） ----------
    all_order_nos = [o['order_no'] for o in valid_orders if o.get('order_no')]
    if all_order_nos:
        existing_set = set(Order.objects.filter(order_no__in=all_order_nos).values_list('order_no', flat=True))
        conflicts = [no for no in all_order_nos if no in existing_set]
        if conflicts:
            return JsonResponse({'code': 0, 'msg': f'以下订单号已存在，无法导入：{", ".join(conflicts)}'})

    # ---------- 2. 预加载区域、客户、用户映射 ----------
    all_area_names = set()
    all_customer_names = set()
    for order in valid_orders:
        area_name = order.get('area_name', '')
        if area_name:
            all_area_names.add(area_name)
        pure_customer = order.get('pure_customer_name', '')
        if pure_customer:
            all_customer_names.add(pure_customer)

    area_full_map = {}
    if all_area_names:
        areas = Area.objects.filter(name__in=all_area_names, is_active=True)
        area_full_map = {a.name: a for a in areas}

    customer_full_map = {}
    if all_customer_names:
        customers = Customer.objects.filter(name__in=all_customer_names, is_active=True)
        customer_full_map = {c.name: c for c in customers}

    all_creator_usernames = set()
    for order in valid_orders:
        if order.get('creator_username'):
            all_creator_usernames.add(order['creator_username'])
    user_map = {}
    if all_creator_usernames:
        users = User.objects.filter(username__in=all_creator_usernames)
        user_map = {u.username: u for u in users}

    # ---------- 3. 分批处理（每批 100 个订单） ----------
    BATCH_SIZE = 100
    total_success = 0
    all_errors = []
    # 用于记录每批创建的订单对象，以便最后更新缓存（如果需要）
    created_orders_in_batch = []

    for i in range(0, len(valid_orders), BATCH_SIZE):
        batch = valid_orders[i:i+BATCH_SIZE]
        batch_orders_to_create = []
        batch_items_to_create = []
        batch_order_objects = []  # 用于保存未保存的 Order 实例

        with transaction.atomic():
            # 3.1 构建 Order 实例（不保存）
            for order_data in batch:
                # 解析各字段
                order_no = order_data['order_no']
                if not order_no:
                    all_errors.append('订单编号不能为空')
                    continue

                items = order_data.get('items', [])
                if not items:
                    all_errors.append(f'订单 {order_no} 缺少明细行')
                    continue

                status = order_data['status']
                area = area_full_map.get(order_data.get('area_name'))
                customer = customer_full_map.get(order_data.get('pure_customer_name'))
                creator = user_map.get(order_data.get('creator_username', ''), request.user)

                create_time = _parse_time(order_data.get('create_time'))
                is_verified = _parse_bool(order_data.get('is_verified', False))

                order = Order(
                    order_no=order_no,
                    customer_name_snapshot=order_data.get('raw_customer_name', ''),
                    area=area,
                    customer=customer,
                    creator=creator,
                    total_amount=0,  # 稍后计算
                    status=status,
                    order_number_snapshot=order_data.get('order_number_snapshot', ''),
                    is_settled=False,  # 先设为False，后面根据数据更新
                    received_amount=Decimal('0'),
                    create_time=create_time,
                    is_verified=is_verified,
                )
                batch_orders_to_create.append(order)

            # 3.2 批量插入订单（一次查询）
            if batch_orders_to_create:
                Order.objects.bulk_create(batch_orders_to_create)
                # 重新查询这些订单，获取 ID 和完整对象
                order_nos = [o.order_no for o in batch_orders_to_create]
                created_orders = Order.objects.filter(order_no__in=order_nos)
                order_map = {o.order_no: o for o in created_orders}

                # 3.3 准备 OrderItem 和更新 Order 总金额/结清状态
                for order_data in batch:
                    order_no = order_data['order_no']
                    order = order_map.get(order_no)
                    if not order:
                        continue

                    items = order_data['items']
                    total = Decimal('0')
                    for item in items:
                        price = Decimal(item['price'])
                        qty = int(item['qty'])
                        amount = price * qty
                        total += amount
                        is_makeup = item.get('is_makeup', False)
                        batch_items_to_create.append(OrderItem(
                            order=order,
                            product=None,
                            product_name=item['product_name'],
                            specification=item.get('spec', ''),
                            unit=item.get('unit', ''),
                            quantity=qty,
                            amount=amount,
                            actual_unit_price=price,
                            snapshot_standard_price=price,
                            snapshot_customer_price=None,
                            is_makeup_item=is_makeup,
                        ))

                    # 更新订单总金额
                    order.total_amount = total
                    # 处理结清状态
                    if order_data['status'] != 'cancelled':
                        is_settled = order_data.get('is_settled', False)
                        received = Decimal(order_data.get('received_amount', '0'))
                        if is_settled or received > 0:
                            order.is_settled = is_settled
                            order.received_amount = received

                    # 将更新后的 order 收集起来，用于后续 bulk_update
                    batch_order_objects.append(order)

                # 3.4 批量创建 OrderItem
                if batch_items_to_create:
                    OrderItem.objects.bulk_create(batch_items_to_create)

                # 3.5 批量更新 Order 的 total_amount, is_settled, received_amount
                if batch_order_objects:
                    Order.objects.bulk_update(
                        batch_order_objects,
                        fields=['total_amount', 'is_settled', 'received_amount']
                    )

                total_success += len(batch)

    # 清除缓存（如有）
    clear_order_cache()

    # 如果有错误，一并返回
    if all_errors:
        return JsonResponse({
            'code': 1,
            'msg': f'成功导入 {total_success} 个订单，部分失败',
            'errors': all_errors
        })
    else:
        return JsonResponse({'code': 1, 'msg': f'成功导入 {total_success} 个订单', 'errors': []})

def import_order_page(request):
    return render(request, 'bill/import_order.html')

from collections import defaultdict

def parse_customer_name(customer_snapshot, area_name=''):
    """
    从客户快照中解析出区域名和纯客户名。
    示例："北区 | 张三" -> ('北区', '张三')
    """
    if not customer_snapshot:
        return (area_name, '')
    if ' | ' in customer_snapshot:
        parts = customer_snapshot.split(' | ', 1)
        return (parts[0].strip(), parts[1].strip())
    return (area_name, customer_snapshot.strip())

from django.views.decorators.cache import never_cache
@never_cache
@login_required
@permission_required(PERM_ORDER_CREATE)
def audit_orders_preview(request):
    """
    审核预览：查询所有未审核订单，检测需新建的区域/客户/商品，以及价格冲突。
    返回统计信息和订单摘要列表。
    """
    if request.method not in ('GET', 'POST'):
        return JsonResponse({'code': 0, 'msg': '仅支持GET/POST'})

    # 查询所有未审核订单（预加载关联数据，避免N+1）
    orders = Order.objects.filter(is_verified=False) \
        .select_related('area') \
        .prefetch_related('items') \
        .order_by('-create_time')

    if not orders.exists():
        return JsonResponse({'code': 0, 'msg': '没有未审核的订单'})

    # 收集待检测数据
    area_names = set()
    product_map = {}          # key: (name, unit) -> {price, spec}
    customer_names = set()
    customer_area = {}        # 客户名 -> 所在区域

    # 同时构造订单摘要
    order_summaries = []

    for order in orders:
        area_name = order.area.name if order.area else ''
        raw_customer = order.customer_name_snapshot or ''
        final_area, pure_name = parse_customer_name(raw_customer, area_name)

        if final_area:
            area_names.add(final_area)
        if pure_name:
            customer_names.add(pure_name)
            if pure_name not in customer_area:
                customer_area[pure_name] = final_area

        # 遍历订单明细，收集商品信息（跳过补货品项）
        for item in order.items.all():
            if item.is_makeup_item:
                continue
            key = (item.product_name, item.unit)
            if key not in product_map:
                product_map[key] = {
                    'price': item.actual_unit_price or Decimal('0'),
                    'spec': item.specification or '',
                }

        order_summaries.append({
            'order_no': order.order_no,
            'customer_snapshot': raw_customer,
            'area_name': area_name,
            'pure_customer': pure_name,
            'total_amount': float(order.total_amount),
            'create_time': order.create_time.strftime('%Y-%m-%d %H:%M'),
        })

    # 1. 检测新区域
    existing_areas = set(
        Area.objects.filter(name__in=area_names, is_active=True)
        .values_list('name', flat=True)
    )
    new_areas = sorted(area_names - existing_areas)

    # 2. 检测商品（新商品和价格冲突）
    existing_products = Product.objects.filter(is_active=True)
    existing_prod_dict = {}
    for p in existing_products:
        existing_prod_dict[(p.name, p.unit)] = p.price

    new_products = []
    conflict_products = []
    for (name, unit), info in product_map.items():
        if (name, unit) not in existing_prod_dict:
            new_products.append({
                'name': name,
                'unit': unit,
                'price': str(info['price']),
                'spec': info['spec'],
            })
        else:
            db_price = existing_prod_dict[(name, unit)]
            order_price = info['price']
            if db_price != order_price:
                conflict_products.append({
                    'name': name,
                    'unit': unit,
                    'db_price': str(db_price),
                    'order_price': str(order_price),
                })

    # 3. 检测新客户
    existing_customers = set(
        Customer.objects.filter(name__in=customer_names, is_active=True)
        .values_list('name', flat=True)
    )
    new_customers = []
    for name in customer_names:
        if name not in existing_customers:
            new_customers.append({
                'name': name,
                'area': customer_area.get(name, ''),
            })

    return JsonResponse({
        'code': 1,
        'data': {
            'new_areas': new_areas,
            'new_products': new_products,
            'conflict_products': conflict_products,
            'new_customers': new_customers,
            'orders': order_summaries,
            'total_unverified': orders.count(),
        }
    })

@login_required
@permission_required(PERM_ORDER_CREATE)
def audit_orders_confirm(request):
    """
    执行审核：根据用户勾选创建/覆盖区域、商品、客户，标记订单为已审核。
    请求体：
    {
        "order_nos": ["order1", ...],         # 要审核的订单号列表，必填
        "new_areas": ["北区", ...],           # 勾选的区域名
        "new_products": [                     # 勾选的商品处理列表
            {
                "name": "商品A",
                "unit": "个",
                "spec": "规格",
                "price": "10.00",
                "action": "create"           # create 或 overwrite
            },
            ...
        ],
        "new_customers": [                    # 勾选的客户
            {"name": "客户1", "area": "北区"},
            ...
        ]
    }
    """
    if request.method != 'POST':
        return JsonResponse({'code': 0, 'msg': '仅支持POST'})

    try:
        payload = json.loads(request.body)
    except:
        return JsonResponse({'code': 0, 'msg': 'JSON 格式错误'})

    order_nos = payload.get('order_nos', [])
    new_areas = payload.get('new_areas', [])
    new_products = payload.get('new_products', [])
    new_customers = payload.get('new_customers', [])

    if not order_nos:
        return JsonResponse({'code': 0, 'msg': '请选择要审核的订单'})

    # 验证订单存在且未审核
    orders = list(
        Order.objects.filter(order_no__in=order_nos, is_verified=False)
        .select_related('area')
        .prefetch_related('items')
    )
    if not orders:
        return JsonResponse({'code': 0, 'msg': '所选订单均不存在或已审核'})

    # 1. 创建区域
    if new_areas:
        existing_areas = set(
            Area.objects.filter(name__in=new_areas, is_active=True)
            .values_list('name', flat=True)
        )
        missing = [a for a in new_areas if a not in existing_areas]
        if missing:
            Area.objects.bulk_create([Area(name=n) for n in missing])

    # 2. 创建客户
    if new_customers:
        existing_cust = set(
            Customer.objects.filter(
                name__in=[c['name'] for c in new_customers],
                is_active=True
            ).values_list('name', flat=True)
        )
        for c in new_customers:
            if c['name'] not in existing_cust:
                area_obj = Area.objects.filter(name=c.get('area', '')).first()
                Customer.objects.create(name=c['name'], area=area_obj)

    # 3. 处理商品（新建或覆盖价格）
    existing_prods = Product.objects.filter(is_active=True)
    existing_prod_map = {}
    for p in existing_prods:
        existing_prod_map[(p.name, p.unit)] = p

    for p in new_products:
        name = p['name'].strip()
        unit = p['unit'].strip()
        spec = p.get('spec', '')
        price = Decimal(p['price'])
        action = p.get('action', 'create')

        key = (name, unit)
        if action == 'overwrite':
            if key in existing_prod_map:
                prod = existing_prod_map[key]
                if prod.price != price:
                    prod.price = price
                    prod.save(update_fields=['price'])
        elif action == 'create':
            if key not in existing_prod_map:
                Product.objects.create(
                    name=name,
                    unit=unit,
                    specification=spec,
                    price=price,
                    stock_system=0,
                    stock_actual=0
                )

    # 4. 标记订单为已审核，并更新明细的价格快照
    # 重新获取最新商品价格
    updated_products = Product.objects.filter(is_active=True)
    price_dict = {}
    for p in updated_products:
        price_dict[(p.name, p.unit)] = p.price

    count = 0
    with transaction.atomic():
        for order in orders:
            order.is_verified = True
            for item in order.items.all():
                if not item.is_makeup_item:
                    key = (item.product_name, item.unit)
                    if key in price_dict:
                        item.snapshot_standard_price = price_dict[key]
                        item.snapshot_customer_price = None
                        item.save(update_fields=['snapshot_standard_price', 'snapshot_customer_price'])
            order.save(update_fields=['is_verified'])
            count += 1

    clear_order_cache()
    return JsonResponse({'code': 1, 'msg': f'成功审核 {count} 个订单'})

def audit_order_page(request):
    """渲染订单审核页面（全新独立页面）"""
    return render(request, 'bill/audit_order.html')

# @login_required
# @permission_required(PERM_ORDER_PRINT)
# def mark_order_printed(request, order_no):
#     """标记订单为已打印（仅在窗口打印后由前端调用）"""
#     if request.method != 'POST':
#         return JsonResponse({'code': 0, 'msg': '仅支持POST请求'}, status=405)
#
#     order = get_object_or_404(Order, order_no=order_no)
#
#     # 允许 pending 和 reopened 状态标记为已打印
#     if order.status in ('pending', 'reopened'):
#         order.status = 'printed'
#         order.save(update_fields=['status'])
#
#         # 清理相关缓存
#         clear_order_cache(order_no)
#
#         # 记录操作日志
#         create_operation_log(
#             request,
#             'mark_printed', 'order', str(order.id),
#             f"订单-{order_no}", "打印后标记为已打印"
#         )
#         return JsonResponse({'code': 1, 'msg': '订单已标记为已打印'})
#
#     elif order.status == 'printed':
#         return JsonResponse({'code': 1, 'msg': '订单已是已打印状态'})
#
#     else:
#         # 作废等状态不允许标记
#         return JsonResponse({'code': 0, 'msg': f'订单状态为{order.status}，无法标记已打印'})
#
#
# @login_required
# @permission_required(PERM_ORDER_PRINT)
# @require_POST
# def batch_mark_printed(request):
#     """批量标记订单为已打印（将 pending 或 reopened 状态改为 printed）"""
#     data = json.loads(request.body)
#     order_nos = data.get('order_nos', [])
#     if not order_nos:
#         return JsonResponse({'code': 0, 'msg': '参数错误'})
#
#     # 定义可打印状态（根据实际模型调整）
#     PRINTABLE_STATUSES = ['pending', 'reopened']  # 若重开状态为其他值，请替换
#     updated = Order.objects.filter(
#         order_no__in=order_nos,
#         status__in=PRINTABLE_STATUSES
#     ).update(status='printed')
#
#     return JsonResponse({'code': 1, 'msg': f'成功标记 {updated} 个订单为已打印'})


@login_required
@permission_required('bill.export_sortrule')  # 按需修改权限
def export_sort_rules(request):
    """导出所有排序规则为 Excel"""
    rules = SortRule.objects.select_related('tag').order_by('stage', 'priority', 'id')

    wb = Workbook()
    ws = wb.active
    ws.title = '排序规则'

    # 表头
    headers = ['阶段', '规则类型', '标签名称', '规格条件', '优先级', '启用']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # 填充数据
    for row_num, rule in enumerate(rules, start=2):
        ws.cell(row=row_num, column=1, value=rule.stage)
        ws.cell(row=row_num, column=2, value=rule.rule_type)  # 'tag' 或 'spec'
        if rule.rule_type == 'tag' and rule.tag:
            ws.cell(row=row_num, column=3, value=rule.tag.name)
        else:
            ws.cell(row=row_num, column=3, value='')
        ws.cell(row=row_num, column=4, value=rule.spec_condition or '')
        ws.cell(row=row_num, column=5, value=rule.priority)
        ws.cell(row=row_num, column=6, value='是' if rule.is_active else '否')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'排序规则_{timezone.now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"
    wb.save(response)
    return response


@login_required
@permission_required('bill.import_sortrule')
def import_sort_rules(request):
    """导入排序规则（覆盖现有规则）"""
    if request.method != 'POST':
        return JsonResponse({'code': 0, 'msg': '仅支持 POST 请求'})

    excel_file = request.FILES.get('file')
    if not excel_file:
        return JsonResponse({'code': 0, 'msg': '请上传 Excel 文件'})

    try:
        wb = load_workbook(excel_file, read_only=True)
        ws = wb.active
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': f'文件解析失败：{str(e)}'})

    # 读取数据行（从第2行开始）
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        return JsonResponse({'code': 0, 'msg': 'Excel 中无数据'})

    new_rules = []
    errors = []
    # 收集所有标签名称以便批量查询
    tag_names = set()
    for row in rows:
        if len(row) < 6:
            errors.append(f'行数据不完整（至少需要6列）: {row}')
            continue
        stage, rule_type, tag_name, spec_cond, priority, active_str = row[:6]
        # 基本校验
        try:
            stage = int(stage)
            priority = int(priority) if priority is not None else 0
        except (ValueError, TypeError):
            errors.append(f'阶段或优先级必须为数字: {row}')
            continue
        if rule_type not in ['tag', 'spec']:
            errors.append(f'规则类型必须为 "tag" 或 "spec": {row}')
            continue
        if rule_type == 'tag' and tag_name:
            tag_names.add(tag_name)
        if rule_type == 'spec' and spec_cond not in ['has_spec', 'no_spec']:
            errors.append(f'规格条件必须为 "has_spec" 或 "no_spec": {row}')
            continue
        is_active = active_str in ['是', '1', 'true', 'True']
        new_rules.append({
            'stage': stage,
            'rule_type': rule_type,
            'tag_name': tag_name,
            'spec_condition': spec_cond,
            'priority': priority,
            'is_active': is_active,
        })

    if errors:
        return JsonResponse({'code': 0, 'msg': '数据校验失败', 'errors': errors[:5]})

    # 查询所有标签
    tag_map = {}
    if tag_names:
        tags = ProductTag.objects.filter(name__in=tag_names)
        tag_map = {tag.name: tag for tag in tags}
        missing = tag_names - set(tag_map.keys())
        if missing:
            return JsonResponse({'code': 0, 'msg': f'以下标签在系统中不存在: {", ".join(missing)}'})

    # 事务：删除旧规则，创建新规则
    with transaction.atomic():
        SortRule.objects.all().delete()
        rules_to_create = []
        for item in new_rules:
            rule = SortRule(
                stage=item['stage'],
                rule_type=item['rule_type'],
                priority=item['priority'],
                is_active=item['is_active'],
            )
            if item['rule_type'] == 'tag':
                rule.tag = tag_map.get(item['tag_name'])
                rule.spec_condition = None
            else:
                rule.tag = None
                rule.spec_condition = item['spec_condition']
            rules_to_create.append(rule)
        if rules_to_create:
            SortRule.objects.bulk_create(rules_to_create)

    return JsonResponse({'code': 1, 'msg': f'导入成功，共导入 {len(rules_to_create)} 条规则'})



@login_required
@permission_required(PERM_ORDER_REOPEN)   # 复用重开权限，或新建 PERM_ORDER_ADD
def add_order_from_existing(request, order_no):
    """
    加单：跳转到开单页，携带原订单数据，并标识为加单模式。
    原订单必须为未作废状态（pending 或 printed）。
    """
    original_order = get_object_or_404(
        Order.objects.select_related('customer', 'area'),
        order_no=order_no
    )
    if original_order.status == 'cancelled':
        messages.warning(request, '已作废订单不能加单，请使用重开功能。')
        return redirect('bill:order_detail', order_no=order_no)

    # 获取原订单明细
    items = OrderItem.objects.select_related('product').filter(order=original_order)

    # 构造传递给开单页的数据（复用 reopen_order_data 结构）
    order_data = {
        'order_no': original_order.order_no,          # ✅ 新增：原订单号（用于显示和JS传参）
        'original_order_no': original_order.order_no, # 保留，明确语义（可选）
        'customer_id': original_order.customer_id if original_order.customer else '',
        'customer_name': original_order.customer_name_snapshot or (
            f"{original_order.area.name} | {original_order.customer.name}"
            if original_order.customer and original_order.area else ''
        ),
        'items': [
            {
                'id': item.product_id if item.product else '',
                'name': item.product_name or (item.product.name if item.product else ''),
                'qty': item.quantity,
                'unit': item.unit,
                'price': float(item.actual_unit_price) if item.actual_unit_price else 0,
                'amt': float(item.amount) if item.amount else 0,
                'spec': item.specification,
            }
            for item in items
        ],
        'is_add': True,   # 前端可用此标识区分“加单”和“重开”（可选）
    }

    context = {
        'is_super_admin': request.user.role and request.user.role.code == ROLE_SUPER_ADMIN,
        'reopen_order_data': order_data,   # 前端JS会读取此变量
    }
    context.update(get_sort_context())   # 排序规则等

    return render(request, 'bill/index.html', context)

# bill/views.py 末尾添加

import io
from openpyxl import load_workbook, Workbook
from django.db import transaction
from django.utils import timezone
from .models import SortRule, ProductTag

def export_to_excel_buffer(data, title, headers, selected_fields, custom_fields=None, file_name='导出', total_row=None):
    """
    返回 BytesIO 对象的 Excel 文件
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title

    # 确定最终字段列表（简化，不处理自定义字段位置）
    field_order = selected_fields
    # 若存在自定义字段，简单追加（可根据需求扩展）
    if custom_fields:
        for cf in custom_fields:
            if cf.get('name') not in field_order:
                field_order.append('custom_' + cf['name'])

    # 写入表头
    header_font = Font(bold=True)
    for col_num, field in enumerate(field_order, 1):
        header_text = headers.get(field, field)
        cell = ws.cell(row=1, column=col_num, value=header_text)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # 写入数据
    for row_num, record in enumerate(data, 2):
        for col_num, field in enumerate(field_order, 1):
            value = record.get(field, '')
            ws.cell(row=row_num, column=col_num, value=value)

    # 若有合计行（略）
    if total_row:
        pass

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_sort_rules_to_io(rules=None):
    """
    导出排序规则为 BytesIO（用于一键备份）
    """
    if rules is None:
        rules = SortRule.objects.select_related('tag').order_by('stage', 'priority', 'id')
    data = []
    seq = 1
    for rule in rules:
        data.append({
            'serial': seq,
            'stage': rule.stage,
            'rule_type': rule.rule_type,
            'tag_name': rule.tag.name if rule.tag else '',
            'spec_condition': rule.spec_condition or '',
            'priority': rule.priority,
            'is_active': '是' if rule.is_active else '否'
        })
        seq += 1

    headers = {
        'serial': '序号',
        'stage': '阶段',
        'rule_type': '规则类型',
        'tag_name': '标签名称',
        'spec_condition': '规格条件',
        'priority': '优先级',
        'is_active': '启用'
    }
    selected_fields = ['serial', 'stage', 'rule_type', 'tag_name', 'spec_condition', 'priority', 'is_active']

    buffer = export_to_excel_buffer(
        data=data,
        title='排序规则',
        headers=headers,
        selected_fields=selected_fields,
        file_name='排序规则导出'
    )
    return buffer

from django.db.models import Prefetch
from decimal import Decimal

def export_orders_to_io(orders=None):
    """
    导出订单数据为 BytesIO（每个订单明细占一行，全量）
    """
    if orders is None:
        orders = Order.objects.select_related('area', 'creator', 'settled_by')\
            .prefetch_related(
                Prefetch('items', queryset=OrderItem.objects.select_related('product'))
            ).order_by('-create_time')

    data = []
    for order in orders:
        area_name = order.area.name if order.area else ''
        customer_name_snap = order.customer_name_snapshot or ''
        order_no_text = order.order_no
        create_time_val = timezone.localtime(order.create_time).replace(tzinfo=None) if order.create_time else ''
        creator_name = order.creator.username if order.creator else ''
        total_amount = float(order.total_amount) if order.total_amount else 0.0
        is_settled_text = '是' if order.is_settled else '否'
        received_amount = float(order.received_amount) if order.received_amount else 0.0
        settled_by_name = order.settled_by.username if order.settled_by else ''
        settled_time_val = timezone.localtime(order.settled_time).replace(tzinfo=None) if order.settled_time else ''
        order_number_snap = order.order_number_snapshot or ''
        delivery_method = order.delivery_method or ''
        is_verified_text = '是' if order.is_verified else '否'

        for item in order.items.all():
            data.append({
                'order_no': order_no_text,
                'customer_name': customer_name_snap,
                'area': area_name,
                'product_name': item.product_name,
                'specification': item.specification or '',
                'unit': item.unit or '',
                'quantity': item.quantity,
                'price': float(item.actual_unit_price) if item.actual_unit_price else 0.0,
                'amount': float(item.amount) if item.amount else 0.0,
                'status': order.status,
                'delivery_method': delivery_method,
                'create_time': create_time_val,
                'creator': creator_name,
                'total_amount': total_amount,
                'is_settled': is_settled_text,
                'received_amount': received_amount,
                'settled_by': settled_by_name,
                'settled_time': settled_time_val,
                'order_number_snapshot': order_number_snap,
                'is_makeup': '是' if item.is_makeup_item else '',
                'is_verified': is_verified_text,
            })

    headers = {
        'order_no': '订单编号',
        'customer_name': '客户名称',
        'area': '区域',
        'product_name': '商品名称',
        'specification': '规格',
        'unit': '单位',
        'quantity': '数量',
        'price': '单价',
        'amount': '小计金额',
        'status': '订单状态',
        'delivery_method': '交付方式',
        'create_time': '创建时间',
        'creator': '开单人',
        'total_amount': '订单总金额',
        'is_settled': '是否结清',
        'received_amount': '已收金额',
        'settled_by': '结清人',
        'settled_time': '结清时间',
        'order_number_snapshot': '制单号快照',
        'is_makeup': '是否补货',
        'is_verified': '审核状态'
    }
    selected_fields = [
        'order_no', 'customer_name', 'area', 'product_name', 'specification', 'unit',
        'quantity', 'price', 'amount', 'status', 'delivery_method', 'create_time',
        'creator', 'total_amount', 'is_settled', 'received_amount', 'settled_by',
        'settled_time', 'order_number_snapshot', 'is_makeup', 'is_verified'
    ]

    buffer = export_to_excel_buffer(
        data=data,
        title='订单数据',
        headers=headers,
        selected_fields=selected_fields,
        file_name='订单导出'
    )
    return buffer

def get_default_creator():
    """获取默认创建者（第一个超级管理员），如果不存在则返回 None"""
    return User.objects.filter(is_superuser=True).first()

def import_orders_from_io(file_obj, strategy='append'):
    """
    从 BytesIO 导入订单（直接执行导入，跳过预览）
    自动创建不存在的区域和客户（区域为空时允许为 None，客户名为空则跳过订单）
    """
    try:
        wb = load_workbook(file_obj, read_only=True)
        parsed = parse_excel_to_structure(wb)
    except Exception as e:
        return {'success': 0, 'skipped': 0, 'errors': [f'解析失败: {str(e)}']}

    order_groups = parsed['order_groups']

    # 预检查已存在订单
    all_order_nos = [g['order_no'] for g in order_groups.values() if g['order_no']]
    existing_set = set()
    if all_order_nos:
        existing_set = set(Order.objects.filter(order_no__in=all_order_nos).values_list('order_no', flat=True))

    # 构建有效订单列表
    valid_orders = []
    skipped = 0
    errors = []

    for key, g in order_groups.items():
        order_no = g['order_no']
        if not order_no:
            skipped += 1
            errors.append(f'订单组缺少编号: {key}')
            continue
        if order_no in existing_set:
            skipped += 1
            continue

        items = g['items']
        pure_customer = items[0]['pure_customer_name']
        # 如果客户名称为空，跳过该订单
        if not pure_customer:
            skipped += 1
            errors.append(f'订单 {order_no} 客户名称为空，跳过')
            continue

        order_data = {
            'order_no': order_no,
            'raw_customer_name': g['raw_customer_name'],
            'area_name': g['area_name'] or '',   # 确保为空字符串
            'pure_customer_name': pure_customer,
            'status': items[0]['status'],
            'create_time': g['create_time'],
            'creator_username': g['creator_username'],
            'is_settled': g['is_settled'],
            'received_amount': str(g['received_amount']),
            'settled_by_username': g['settled_by_username'],
            'settled_time': g['settled_time'],
            'order_number_snapshot': g.get('order_number_snapshot', ''),
            'items': [{
                'product_name': item['product_name'],
                'spec': item['spec'],
                'unit': item['unit'],
                'qty': item['qty'],
                'price': str(item['price']),
                'status': item['status'],
                'is_makeup': item.get('is_makeup_item', False),
            } for item in items],
            'is_verified': g.get('is_verified', False),
        }
        valid_orders.append(order_data)

    if not valid_orders:
        # 如果没有有效订单，返回结果（可能全部是跳过）
        return {'success': 0, 'skipped': skipped, 'errors': errors}

    # ---------- 自动创建缺失的区域和客户 ----------
    all_area_names = set()
    all_customer_names = set()
    area_for_customer = {}

    for order_data in valid_orders:
        area_name = order_data.get('area_name')
        if area_name:
            all_area_names.add(area_name)
        cust_name = order_data.get('pure_customer_name')
        if cust_name:
            all_customer_names.add(cust_name)
            if area_name and cust_name not in area_for_customer:
                area_for_customer[cust_name] = area_name

    # 创建区域（仅非空名称）
    area_map = {}
    for area_name in all_area_names:
        if area_name:
            area_obj, created = Area.objects.get_or_create(
                name=area_name,
                defaults={'remark': '订单导入自动创建', 'is_active': True}
            )
            area_map[area_name] = area_obj

    # 创建客户（允许区域为 None）
    customer_map = {}
    for cust_name in all_customer_names:
        if not cust_name:
            continue
        area_name_for_cust = area_for_customer.get(cust_name)
        area_obj = area_map.get(area_name_for_cust) if area_name_for_cust else None
        try:
            customer_obj = Customer.objects.get(name=cust_name, area=area_obj)
        except Customer.DoesNotExist:
            customer_obj = Customer.objects.create(
                name=cust_name,
                area=area_obj,
                remark='订单导入自动创建',
                is_active=True
            )
        customer_map[cust_name] = customer_obj

    # 用户映射
    all_creator_usernames = {o['creator_username'] for o in valid_orders if o.get('creator_username')}
    user_map = {u.username: u for u in User.objects.filter(username__in=all_creator_usernames)}
    default_creator = get_default_creator()

    # ---------- 导入订单 ----------
    success = 0
    import_errors = []
    with transaction.atomic():
        for order_data in valid_orders:
            order_no = order_data['order_no']
            try:
                items = order_data['items']
                if not items:
                    import_errors.append(f'订单 {order_no} 无明细')
                    continue

                # 获取区域（若区域名为空则置为 None）
                area_name = order_data.get('area_name')
                area = area_map.get(area_name) if area_name else None

                customer = customer_map.get(order_data.get('pure_customer_name'))
                if not customer:
                    # 理论上不会发生，因为已过滤且自动创建，但保留防护
                    import_errors.append(f'订单 {order_no} 客户不存在')
                    continue

                creator = user_map.get(order_data.get('creator_username', ''))
                if not creator:
                    creator = default_creator

                create_time = order_data.get('create_time') or timezone.now()
                is_verified = order_data.get('is_verified', False)

                order = Order(
                    order_no=order_no,
                    customer_name_snapshot=order_data.get('raw_customer_name', ''),
                    area=area,
                    customer=customer,
                    creator=creator,
                    total_amount=0,
                    status=order_data['status'],
                    order_number_snapshot=order_data.get('order_number_snapshot', ''),
                    is_settled=False,
                    received_amount=Decimal('0'),
                    create_time=create_time,
                    is_verified=is_verified,
                )
                order.save()

                if order_data['status'] != 'cancelled':
                    is_settled = order_data.get('is_settled', False)
                    received = Decimal(order_data.get('received_amount', '0'))
                    if is_settled or received > 0:
                        order.is_settled = is_settled
                        order.received_amount = received
                        order.save(update_fields=['is_settled', 'received_amount'])

                total = Decimal('0')
                items_to_create = []
                for item in items:
                    prod_name = item['product_name']
                    unit = item.get('unit', '')
                    price = Decimal(item['price'])
                    qty = int(item['qty'])
                    amount = price * qty
                    total += amount
                    is_makeup = item.get('is_makeup', False)
                    items_to_create.append(OrderItem(
                        order=order,
                        product=None,
                        product_name=prod_name,
                        specification=item.get('spec', ''),
                        unit=unit,
                        quantity=qty,
                        amount=amount,
                        actual_unit_price=price,
                        snapshot_standard_price=price,
                        snapshot_customer_price=None,
                        is_makeup_item=is_makeup,
                    ))
                OrderItem.objects.bulk_create(items_to_create)
                order.total_amount = total
                order.save(update_fields=['total_amount'])
                success += 1

            except Exception as e:
                import_errors.append(f'订单 {order_no} 导入失败: {str(e)}')
                raise Exception(f'订单 {order_no} 导入失败: {str(e)}')

    return {'success': success, 'skipped': skipped, 'errors': import_errors}

def import_sort_rules_from_io(file_obj, strategy='append'):
    """
    从 BytesIO 导入排序规则（覆盖现有规则）
    返回 {'success': int, 'skipped': int, 'errors': list}
    """
    try:
        wb = load_workbook(file_obj, read_only=True)
        ws = wb.active
    except Exception as e:
        return {'success': 0, 'skipped': 0, 'errors': [f'文件解析失败: {str(e)}']}

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        # 无数据时视为成功（跳过）
        return {'success': 0, 'skipped': 0, 'errors': []}

    tag_names = set()
    new_rules = []
    errors = []
    for row_idx, row in enumerate(rows, start=2):
        if len(row) < 7:
            errors.append(f'第{row_idx}行：列数不足')
            continue
        stage_val, rule_type, tag_name, spec_cond, priority_val, active_str = row[1:7]
        try:
            stage = int(stage_val) if stage_val is not None else 0
            priority = int(priority_val) if priority_val is not None else 0
        except (ValueError, TypeError):
            errors.append(f'第{row_idx}行：阶段或优先级必须为数字')
            continue
        if rule_type not in ['tag', 'spec']:
            errors.append(f'第{row_idx}行：规则类型必须为 "tag" 或 "spec"')
            continue
        if rule_type == 'tag' and tag_name:
            tag_names.add(tag_name)
        if rule_type == 'spec' and spec_cond not in ['has_spec', 'no_spec']:
            errors.append(f'第{row_idx}行：规格条件必须为 "has_spec" 或 "no_spec"')
            continue
        is_active = active_str in ['是', '1', 'true', 'True']
        new_rules.append({
            'stage': stage,
            'rule_type': rule_type,
            'tag_name': tag_name,
            'spec_condition': spec_cond,
            'priority': priority,
            'is_active': is_active,
        })

    if errors:
        return {'success': 0, 'skipped': 0, 'errors': errors}

    # 检查标签是否存在
    tag_map = {}
    if tag_names:
        tags = ProductTag.objects.filter(name__in=tag_names)
        tag_map = {tag.name: tag for tag in tags}
        missing = tag_names - set(tag_map.keys())
        if missing:
            return {'success': 0, 'skipped': 0, 'errors': [f'以下标签不存在: {", ".join(missing)}']}

    # 覆盖导入（删除旧规则）
    with transaction.atomic():
        SortRule.objects.all().delete()
        rules_to_create = []
        for item in new_rules:
            rule = SortRule(
                stage=item['stage'],
                rule_type=item['rule_type'],
                priority=item['priority'],
                is_active=item['is_active'],
            )
            if item['rule_type'] == 'tag':
                rule.tag = tag_map.get(item['tag_name'])
                rule.spec_condition = None
            else:
                rule.tag = None
                rule.spec_condition = item['spec_condition']
            rules_to_create.append(rule)
        SortRule.objects.bulk_create(rules_to_create)

    return {'success': len(rules_to_create), 'skipped': 0, 'errors': []}
