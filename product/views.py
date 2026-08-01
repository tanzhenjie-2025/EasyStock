from django.db import IntegrityError, transaction
from django.db.models.functions import Coalesce
from django.views.decorators.http import require_POST
import io
import openpyxl
import xlrd
from datetime import datetime, timedelta
from django.db.models import Sum, Count, Q, F, Prefetch, Case, When, DateTimeField, DecimalField

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

from django.core.cache import cache

from django.db import transaction
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Sum, Count

from accounts.models import ROLE_SUPER_ADMIN, ROLE_ADMIN, ROLE_OPERATOR, PERM_PRODUCT_AUDIT
from product.models import Product, StockIn, StockInItem, ProductPriceHistory, ProductTag
from bill.views import (
    # 复用缓存工具
    clear_stock_cache, clear_product_search_cache,
)
import json
import decimal
import logging

logger = logging.getLogger(__name__)

from .models import Unit

# ========== RBAC权限组件 ==========
from accounts.views import permission_required, create_operation_log
from accounts.models import (
    PERM_PRODUCT_VIEW, PERM_PRODUCT_ADD, PERM_PRODUCT_EDIT,
    PERM_PRODUCT_DELETE, PERM_PRODUCT_ALIAS_ADD, PERM_PRODUCT_ALIAS_DELETE,
    PERM_PRODUCT_IMPORT, PERM_PRODUCT_STOCK_OP, PERM_PRODUCT_DETAIL,
PERM_CUSTOMER_IMPORT,PERM_CUSTOMER_EXPORT
)


from pypinyin import lazy_pinyin    # 确保已导入

# 业务模型
from bill.models import Order, OrderItem
from product.models import Product, ProductAlias
from area_manage.models import Area

# ====================== 缓存常量配置 ======================
CACHE_AREA = 3600
CACHE_COMMON = 60
CACHE_SALES_RANK = 10
CACHE_PAGINATION_COUNT = 120

CACHE_PREFIX_PRODUCT_LIST = "product:list:"
CACHE_PREFIX_PRODUCT_DETAIL = "product:detail:"
CACHE_PREFIX_SALES_RANK = "product:sales_rank:"
CACHE_PREFIX_PRODUCT_COUNT = "product:count:"
KEY_AREA = "area:data"
KEY_PRODUCT_ALIAS = "product:alias"

# ========== 入库模块权限常量 ==========
PERM_STOCK_IN_CREATE = 'stock_in_create'  # 新建入库
PERM_STOCK_IN_VIEW = 'stock_in_view'  # 查看入库
PERM_STOCK_IN_CANCEL = 'stock_in_cancel'  # 作废入库


# ====================== 缓存工具函数 ======================
def clear_product_all_cache():
    # 删除精确 key
    cache.delete_many([KEY_AREA, KEY_PRODUCT_ALIAS])

    # 改用 delete_pattern 批量删除（与订单模块保持一致）
    cache.delete_pattern(f"{CACHE_PREFIX_PRODUCT_LIST}*")
    cache.delete_pattern(f"{CACHE_PREFIX_PRODUCT_DETAIL}*")
    cache.delete_pattern(f"{CACHE_PREFIX_SALES_RANK}*")
    cache.delete_pattern(f"{CACHE_PREFIX_PRODUCT_COUNT}*")

    # 清理排序相关缓存
    cache.delete('sort_stages_json')
    cache.delete('product_tags_map_json')

    logger.info("已清理全部商品及排序相关缓存")


@login_required
def search_unit(request):
    """单位模糊搜索，keyword为空时返回所有启用的单位"""
    keyword = request.GET.get('keyword', '').strip()
    if keyword:
        units = Unit.objects.filter(
            Q(name__icontains=keyword) |
            Q(pinyin_full__icontains=keyword) |
            Q(pinyin_abbr__icontains=keyword),
            is_active=True
        )[:50]  # 搜索时最多返回50条
    else:
        # 返回全部启用单位，按排序字段
        units = Unit.objects.filter(is_active=True).order_by('sort_order', 'id')[:100]

    data = [{'name': unit.name} for unit in units]
    return JsonResponse({'code': 1, 'data': data})

@login_required
@permission_required(PERM_PRODUCT_VIEW)
def product_manage(request):
    no_tag = request.GET.get('no_tag', '0').strip() == '1'
    page = request.GET.get('page', 1)
    keyword = request.GET.get('keyword', '').strip()
    status = request.GET.get('status', 'all')
    # 🔥 修复：清洗 tag_ids，过滤空字符串和非数字
    tag_ids = [tid for tid in request.GET.getlist('tag_ids', []) if tid.strip().isdigit()]
    active_tab = request.GET.get('active_tab', 'list')

    # ====================== POST逻辑保持不变 ======================
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add_tag':
            tag_name = request.POST.get('tag_name', '').strip()
            tag_color = request.POST.get('tag_color', '#3498db')
            if tag_name:
                ProductTag.objects.create(name=tag_name, color=tag_color)
                create_operation_log(request, 'create', 'product_tag', 0, tag_name, '新增标签')
        elif action == 'toggle_tag':
            tag_id = request.POST.get('tag_id')
            tag = get_object_or_404(ProductTag.all_objects, id=tag_id)
            tag.is_active = not tag.is_active
            tag.save(update_fields=['is_active'])
            create_operation_log(request, 'update', 'product_tag', tag.id, tag.name, '切换标签状态')
        elif action == 'batch_add_tag':
            target_tag_id = request.POST.get('target_tag_id')
            product_ids_str = request.POST.get('product_ids', '').strip()
            product_ids = [int(pid) for pid in product_ids_str.split(',') if pid.strip().isdigit()]
            if target_tag_id and product_ids:
                tag = get_object_or_404(ProductTag, id=target_tag_id)
                tag.products.add(*product_ids)
                create_operation_log(request, 'update', 'product', 0, f'批量为{len(product_ids)}个商品添加标签', '批量打标')
        elif action == 'batch_remove_tag':
            target_tag_id = request.POST.get('target_tag_id')
            product_ids_str = request.POST.get('product_ids', '').strip()
            product_ids = [int(pid) for pid in product_ids_str.split(',') if pid.strip().isdigit()]
            if target_tag_id and product_ids:
                tag = get_object_or_404(ProductTag, id=target_tag_id)
                tag.products.remove(*product_ids)
                create_operation_log(request, 'update', 'product', 0, f'批量为{len(product_ids)}个商品移除标签', '批量移除标签')
        clear_product_all_cache()
        return redirect(f"{request.path}?active_tab={active_tab}")

    # ====================== 标签查询 ======================
    all_tags = ProductTag.all_objects.annotate(
        product_count=Count('products')
    ).order_by('sort_order', '-id')

    # 修改缓存 key，把 no_tag 也加进去
    cache_key = f"{CACHE_PREFIX_PRODUCT_LIST}{keyword}:{page}:{status}:{','.join(tag_ids)}:{no_tag}"

    cached_data = cache.get(cache_key)

    if cached_data:
        product_list_data = cached_data['product_list']
        paginator_data = cached_data['paginator']
        count_stats = cached_data['count_stats']
    else:
        # 基础查询集（优化预加载）
        products_query = Product.all_objects.order_by('name').only(
            'id', 'name', 'price', 'unit', 'specification',
            'stock_system', 'stock_actual', 'is_active', 'pinyin_abbr'
        ).prefetch_related(
            Prefetch('tags', queryset=ProductTag.objects.only('id', 'name', 'color')),
            Prefetch('aliases', queryset=ProductAlias.all_objects.only('id', 'alias_name', 'product_id'))
        )

        # 状态筛选
        if status == 'active':
            products_query = products_query.filter(is_active=True)
        elif status == 'inactive':
            products_query = products_query.filter(is_active=False)

        # 关键词搜索
        if keyword:
            alias_product_ids = ProductAlias.all_objects.filter(
                Q(alias_name__icontains=keyword)
            ).values_list('product_id', flat=True)
            products_query = products_query.filter(
                Q(name__icontains=keyword) | Q(id__in=alias_product_ids)
            )

        # 标签筛选
        if tag_ids:
            products_query = products_query.filter(tags__id__in=tag_ids).distinct()

        # 新增：无标签筛选（使用子查询判断不存在关联标签）
        if no_tag:
            products_query = products_query.annotate(
                tag_count=Count('tags')
            ).filter(tag_count=0)

        # 优化：合并商品状态统计为一次查询（不考虑 no_tag 的统计，仍统计全体）
        count_stats = Product.all_objects.aggregate(
            count_all=Count('id'),
            count_active=Count(Case(When(is_active=True, then=1))),
            count_inactive=Count(Case(When(is_active=False, then=1)))
        )

        # 分页总数缓存
        count_cache_key = f"{CACHE_PREFIX_PRODUCT_COUNT}{keyword}:{status}:{','.join(tag_ids)}:{no_tag}"
        total_count = cache.get(count_cache_key)
        if total_count is None:
            total_count = products_query.count()
            cache.set(count_cache_key, total_count, CACHE_PAGINATION_COUNT)

        # 分页
        page_size = 15
        paginator = Paginator(products_query, page_size)
        try:
            page_products = paginator.page(page)
        except PageNotAnInteger:
            page_products = paginator.page(1)
        except EmptyPage:
            page_products = paginator.page(paginator.num_pages)

        # 序列化商品数据
        product_list_data = []
        for product in page_products:
            product_list_data.append({
                'id': product.id,
                'name': product.name,
                'price': float(product.price),
                'unit': product.unit,
                'specification': product.specification,
                'stock_system': product.stock_system,
                'stock_actual': product.stock_actual,
                'aliases': [...],
                'tags': [...],
                'status': 1 if product.is_active else 0,
                'pinyin_abbr': product.pinyin_abbr,  # 新增
            })

        # 序列化分页器数据
        paginator_data = {
            'num_pages': paginator.num_pages,
            'page_range': list(paginator.page_range),
            'current_page': page_products.number,
            'has_previous': page_products.has_previous(),
            'has_next': page_products.has_next(),
            'previous_page_number': page_products.previous_page_number() if page_products.has_previous() else None,
            'next_page_number': page_products.next_page_number() if page_products.has_next() else None,
        }

        # 缓存数据
        cache.set(cache_key, {
            'product_list': product_list_data,
            'paginator': paginator_data,
            'count_stats': count_stats
        }, CACHE_COMMON)

    # ========== 新增：单位管理相关 ==========
    unit_keyword = request.GET.get('unit_keyword', '').strip()
    unit_status = request.GET.get('unit_status', 'all')

    # 全量统计
    count_unit_all = Unit.all_objects.count()
    count_unit_active = Unit.all_objects.filter(is_active=True).count()
    count_unit_inactive = Unit.all_objects.filter(is_active=False).count()

    # 带过滤的列表
    unit_list_qs = Unit.all_objects.all().order_by('sort_order', 'id')

    if unit_keyword:
        unit_list_qs = unit_list_qs.filter(name__icontains=unit_keyword)
    if unit_status == 'active':
        unit_list_qs = unit_list_qs.filter(is_active=True)
    elif unit_status == 'inactive':
        unit_list_qs = unit_list_qs.filter(is_active=False)
    # 区域缓存
    areas = cache.get(KEY_AREA)
    if not areas:
        areas = list(Area.objects.only('id', 'name'))
        cache.set(KEY_AREA, areas, CACHE_AREA)

    context = {
        'products': product_list_data,
        'paginator_data': paginator_data,  # 前端需适配分页器数据结构
        'keyword': keyword,
        'status': status,
        'tag_ids': list(map(int, tag_ids)),  # 转换为整数列表供前端使用
        'all_tags': all_tags,
        'count_all': count_stats['count_all'],
        'count_active': count_stats['count_active'],
        'count_inactive': count_stats['count_inactive'],
        'areas': areas,
        'active_tab': active_tab,
        'can_add_product': request.user.has_permission(PERM_PRODUCT_ADD),
        'can_edit_product': request.user.has_permission(PERM_PRODUCT_EDIT),
        'can_delete_product': request.user.has_permission(PERM_PRODUCT_DELETE),
        'can_import_product': request.user.has_permission(PERM_PRODUCT_IMPORT),
        'can_stock_operation': request.user.has_permission(PERM_PRODUCT_STOCK_OP),

        'all_units': unit_list_qs,
        'unit_keyword': unit_keyword,
        'unit_status': unit_status,
        'count_unit_all': count_unit_all,
        'count_unit_active': count_unit_active,
        'count_unit_inactive': count_unit_inactive,
        'no_tag': no_tag,
    }

    return render(request, 'product/product_manage.html', context)


# ====================== 商品CRUD ======================
@login_required
@permission_required(PERM_PRODUCT_ADD)
def product_add(request):
    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            price = request.POST.get('price', '0').strip()
            unit = request.POST.get('unit', '件').strip()
            stock = request.POST.get('stock', '77').strip()
            specification = request.POST.get('specification', '').strip()  # 新增

            if not name:
                return JsonResponse({'code': 0, 'msg': '商品名称不能为空'})
            if not price or float(price) < 0:
                return JsonResponse({'code': 0, 'msg': '请输入有效的单价'})

            product = Product.objects.create(
                name=name, price=float(price), unit=unit,
                stock_system=int(stock) if stock.isdigit() else 77,
                stock_actual=int(stock) if stock.isdigit() else 77,
                specification = specification  # 新增
            )

            create_operation_log(
                request=request, op_type='create', obj_type='product',
                obj_id=product.id, obj_name=product.name,
                detail=f"新增商品：名称={product.name}，单价={product.price}，单位={product.unit}，系统库存={product.stock_system}，实际库存={product.stock_actual}"
            )

            clear_product_all_cache()
            return JsonResponse({'code': 1, 'msg': '商品新增成功'})
        except IntegrityError:
            # 查询已存在的同名同单位启用商品
            exist_product = Product.objects.filter(name=name, unit=unit).first()
            if exist_product:
                return JsonResponse({
                    'code': 2,
                    'msg': f'已存在同名同单位商品：{name}（{unit}）',
                    'data': {
                        'product_id': exist_product.id,
                        'product_name': exist_product.name
                    }
                })
            return JsonResponse({'code': 0, 'msg': '商品创建失败，存在唯一约束冲突'})
        except Exception as e:
            return JsonResponse({'code': 0, 'msg': f'新增失败：{str(e)}'})
    return JsonResponse({'code': 0, 'msg': '请求方式错误'})

@login_required
@permission_required(PERM_PRODUCT_EDIT)
def product_edit(request, pk):
    product = get_object_or_404(Product.objects.prefetch_related('aliases'), pk=pk)
    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            price = request.POST.get('price', '0').strip()
            unit = request.POST.get('unit', '件').strip()
            stock = request.POST.get('stock', '77').strip()
            specification = request.POST.get('specification', '').strip()  # 新增
            # 可选：接收备注
            remark = request.POST.get('remark', '后台编辑').strip()

            if not name:
                return JsonResponse({'code': 0, 'msg': '商品名称不能为空'})
            if not price or float(price) < 0:
                return JsonResponse({'code': 0, 'msg': '请输入有效的单价'})
            if Product.objects.filter(name=name, unit=unit).exclude(id=pk).exists():
                return JsonResponse({'code': 0, 'msg': '已存在同名同单位的商品'})

            old_info = f"名称={product.name}，单价={product.price}，单位={product.unit}，系统库存={product.stock_system}"

            # 🔥 核心：检测价格变动
            old_price_val = product.price
            new_price_val = decimal.Decimal(price)

            product.name = name
            product.price = new_price_val
            product.unit = unit
            product.specification = specification  # 新增
            product.stock_system = int(stock) if stock.isdigit() else 77
            product.save()

            # 🔥 如果价格变了，写入历史表
            if old_price_val != new_price_val:
                ProductPriceHistory.objects.create(
                    product=product,
                    old_price=old_price_val,
                    new_price=new_price_val,
                    operator=request.user,
                    remark=remark
                )

            create_operation_log(
                request=request, op_type='update', obj_type='product',
                obj_id=product.id, obj_name=product.name,
                detail=f"编辑商品：原信息[{old_info}] → 新信息[名称={product.name}，单价={product.price}]"
            )

            clear_product_all_cache()
            return JsonResponse({'code': 1, 'msg': '商品编辑成功'})
        except Exception as e:
            return JsonResponse({'code': 0, 'msg': f'编辑失败：{str(e)}'})
    return JsonResponse({'code': 0, 'msg': '请求方式错误'})

@login_required
@permission_required(PERM_PRODUCT_DELETE)
def product_delete(request, pk):
    try:
        product = get_object_or_404(Product.all_objects, pk=pk)
        product.delete()
        create_operation_log(request=request, op_type='delete', obj_type='product', obj_id=pk, obj_name=product.name,
                             detail=f"禁用商品")
        clear_product_all_cache()
        return JsonResponse({'code': 1, 'msg': '商品禁用成功'})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': f'禁用失败：{str(e)}'})

@login_required
@permission_required(PERM_PRODUCT_EDIT)
def product_restore(request, pk):
    try:
        product = get_object_or_404(Product.all_objects, pk=pk)
        product.is_active = True
        product.save(update_fields=['is_active'])
        create_operation_log(request=request, op_type='update', obj_type='product', obj_id=pk, obj_name=product.name,
                             detail=f"启用商品")
        clear_product_all_cache()
        return JsonResponse({'code': 1, 'msg': '商品启用成功'})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': f'启用失败：{str(e)}'})


# ====================== 🔥 行内编辑（仅修改系统库存） ======================
@require_POST
@permission_required(PERM_PRODUCT_EDIT)
def product_inline_update(request):
    try:
        pk = request.POST.get('id')
        field = request.POST.get('field')
        value = request.POST.get('value')
        product = get_object_or_404(Product, pk=pk)

        if field == 'price':
            product.price = float(value)
        elif field == 'stock_system':
            product.stock_system = int(value)
        else:
            return JsonResponse({'code': 0, 'msg': '无效字段'})

        product.save(update_fields=[field])
        clear_product_all_cache()
        return JsonResponse({'code': 1, 'msg': '更新成功'})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': str(e)})


# ====================== 🔥 状态开关 ======================
@require_POST
@permission_required(PERM_PRODUCT_EDIT)
def product_toggle_status(request):
    try:
        pk = request.POST.get('id')
        product = get_object_or_404(Product.all_objects, pk=pk)
        product.is_active = not product.is_active
        product.save(update_fields=['is_active'])
        clear_product_all_cache()
        return JsonResponse({'code': 1, 'status': 1 if product.is_active else 0, 'msg': '状态已更新'})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': str(e)})


# ====================== 🔥 批量操作 ======================
@require_POST
@permission_required(PERM_PRODUCT_DELETE)
def product_batch_operation(request):
    try:
        ids = json.loads(request.POST.get('ids', '[]'))
        action = request.POST.get('action')
        if not ids:
            return JsonResponse({'code': 0, 'msg': '请选择商品'})

        products = Product.all_objects.filter(id__in=ids)
        if action == 'enable':
            products.update(is_active=True)
            msg = '批量启用成功'
        elif action == 'disable':
            products.update(is_active=False)
            msg = '批量停用成功'
        else:
            return JsonResponse({'code': 0, 'msg': '无效操作'})

        clear_product_all_cache()
        return JsonResponse({'code': 1, 'msg': msg})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': str(e)})


# ====================== 🔥 新增：实际库存校准接口 ======================
@require_POST
@permission_required(PERM_PRODUCT_STOCK_OP)
def product_stock_calibrate(request):
    """实际库存校准（核心功能）"""
    try:
        pk = request.POST.get('id')
        actual_stock = request.POST.get('actual_stock')
        product = get_object_or_404(Product, pk=pk)

        if not actual_stock or not actual_stock.isdigit() or int(actual_stock) < 0:
            return JsonResponse({'code': 0, 'msg': '请输入有效的实际库存'})

        old_actual = product.stock_actual
        product.stock_actual = int(actual_stock)
        product.save(update_fields=['stock_actual'])

        # 校准日志
        create_operation_log(
            request=request, op_type='calibrate_stock', obj_type='product',
            obj_id=product.id, obj_name=product.name,
            detail=f"库存校准：原实际库存={old_actual} → 新实际库存={product.stock_actual}，系统库存={product.stock_system}"
        )

        clear_product_all_cache()
        return JsonResponse({'code': 1, 'msg': '库存校准成功'})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': str(e)})


# ====================== 别名CRUD（无修改） ======================
@login_required
@permission_required(PERM_PRODUCT_ALIAS_ADD)
def alias_add(request):
    if request.method == 'POST':
        try:
            product_id = request.POST.get('product_id')
            alias_name = request.POST.get('alias_name').strip()
            product = get_object_or_404(Product, pk=product_id)
            alias = ProductAlias.objects.create(product=product, alias_name=alias_name)
            create_operation_log(request=request, op_type='create', obj_type='product_alias', obj_id=alias.id,
                                 obj_name=f"{product.name}-{alias_name}")
            clear_product_all_cache()
            return JsonResponse(
                {'code': 1, 'msg': '别名添加成功', 'data': {'id': alias.id, 'alias_name': alias.alias_name}})
        except IntegrityError:
            return JsonResponse({'code': 0, 'msg': '别名已存在'})
        except Exception as e:
            return JsonResponse({'code': 0, 'msg': str(e)})
    return JsonResponse({'code': 0, 'msg': '请求方式错误'})

@login_required
@permission_required(PERM_PRODUCT_ALIAS_DELETE)
def alias_delete(request, pk):
    try:
        alias = get_object_or_404(ProductAlias.all_objects, pk=pk)
        alias.delete()
        create_operation_log(request=request, op_type='delete', obj_type='product_alias', obj_id=pk,
                             obj_name=f"{alias.product.name}-{alias.alias_name}")
        clear_product_all_cache()
        return JsonResponse({'code': 1, 'msg': '别名禁用成功'})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': str(e)})

@login_required
@permission_required(PERM_PRODUCT_EDIT)
def product_edit_data(request, pk):
    product = get_object_or_404(Product.objects.prefetch_related('aliases'), pk=pk)
    return JsonResponse({
        'id': product.id,
        'name': product.name,
        'price': float(product.price),
        'unit': product.unit,
        'specification': product.specification,  # 新增
        'stock': product.stock_system,
        'aliases': [{'id': a.id, 'alias_name': a.alias_name} for a in product.aliases.all()]
    })


# ====================== 导入/导出/快速出入库（仅修改系统库存） ======================
# product/views.py

@login_required
@require_POST
@permission_required(PERM_PRODUCT_IMPORT)
def product_import(request):
    """
    商品导入（单批次）- 复用数据迁移的导入逻辑
    """
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'code': 0, 'msg': '请选择Excel文件'})
        file_obj = request.FILES['file']
        # 调用数据迁移导入函数
        result = import_products_from_io(file_obj, strategy='append')
        # 根据结果构建响应
        if result.get('errors'):
            msg = f"成功 {result['success']} 条，跳过 {result['skipped']} 条"
            # 只展示前5条错误
            if result['errors']:
                error_preview = result['errors'][:5]
                msg += f"；错误：{'；'.join(error_preview)}"
            return JsonResponse({'code': 0, 'msg': msg})
        else:
            return JsonResponse({'code': 1, 'msg': f"成功导入 {result['success']} 条商品"})
    except Exception as e:
        logger.error(f"商品导入失败: {str(e)}", exc_info=True)
        return JsonResponse({'code': 0, 'msg': f'导入失败：{str(e)}'})


@login_required
@permission_required('product_export')  # 确保权限定义正确
def product_export(request):
    """
    商品导出（单批次）- 全字段，支持按关键字和状态筛选，包含禁用商品
    """
    try:
        keyword = request.GET.get('keyword', '').strip()
        status = request.GET.get('status', 'all')
        # 使用 all_objects 包含所有商品
        products_query = Product.all_objects.all()
        if status == 'active':
            products_query = products_query.filter(is_active=True)
        elif status == 'inactive':
            products_query = products_query.filter(is_active=False)
        # 关键字过滤（包括别名）
        if keyword:
            from django.db.models import Q
            alias_product_ids = ProductAlias.objects.filter(
                Q(alias_name__icontains=keyword)
            ).values_list('product_id', flat=True)
            products_query = products_query.filter(
                Q(name__icontains=keyword) | Q(id__in=alias_product_ids)
            )
        products = products_query.prefetch_related('aliases', 'tags').order_by('name')
        # 调用导出函数
        buffer = export_products_to_io(products)
        # 构造响应
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'商品列表_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        logger.error(f"导出商品失败: {str(e)}", exc_info=True)
        return JsonResponse({'code': 0, 'msg': f'导出失败：{str(e)}'})

@login_required
@require_POST
@permission_required(PERM_PRODUCT_STOCK_OP)
def quick_stock_operation(request):
    """快速出入库：仅操作系统库存"""
    try:
        data = json.loads(request.body)
        items = data.get('items', [])
        with transaction.atomic():
            pids = [int(i['product_id']) for i in items if i.get('product_id')]
            products = {p.id: p for p in Product.objects.filter(id__in=pids).select_for_update()}
            update_list = []
            for i in items:
                pid = int(i['product_id'])
                if pid not in products: continue
                p = products[pid]
                in_q = int(i.get('in_quantity', 0))
                out_q = int(i.get('out_quantity', 0))
                if out_q > p.stock_system:
                    raise Exception(f'{p.name} 系统库存不足')
                p.stock_system += in_q - out_q
                update_list.append(p)
            Product.objects.bulk_update(update_list, ['stock_system'])
            clear_product_all_cache()
            return JsonResponse({'code': 1, 'msg': '出入库成功'})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': str(e)})


# 导出/详情/排行/库存列表（无修改，仅适配字段）
@login_required
@permission_required(PERM_PRODUCT_IMPORT)
def export_to_excel(data, title, headers, selected_fields, custom_fields, file_name):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    final = selected_fields.copy()
    for cf in custom_fields:
        final.insert(final.index(cf['target']) + 1, f'custom_{cf["name"]}')
    for i, h in enumerate([headers[f] for f in selected_fields], 1):
        ws.cell(1, i, h)
    for r, d in enumerate(data, 2):
        for c, f in enumerate(selected_fields, 1):
            ws.cell(r, c, d.get(f, ''))
    buffer = BytesIO()
    wb.save(buffer)
    return HttpResponse(buffer.getvalue(), content_type='application/vnd.ms-excel')

# ===================== 修改：商品详情主视图 =====================
@login_required
@permission_required(PERM_PRODUCT_DETAIL)
def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)

    # 1. 最近售卖订单列表 (分页20条，展示当时售价)
    # 排除作废订单
    valid_status = ['pending', 'printed', 'reopened']
    recent_sales_qs = OrderItem.objects.filter(
        product_id=pk,
        order__status__in=valid_status
    ).select_related('order', 'order__customer').order_by('-order__create_time')

    paginator = Paginator(recent_sales_qs, 20)
    page = request.GET.get('page', 1)
    try:
        recent_sales = paginator.page(page)
    except PageNotAnInteger:
        recent_sales = paginator.page(1)
    except EmptyPage:
        recent_sales = paginator.page(paginator.num_pages)

    # 2. 价格变更历史
    price_history = ProductPriceHistory.objects.filter(product=product).select_related('operator')[:50]  # 限制展示条数

    return render(request, 'product/product_detail.html', {
        'product': product,
        'recent_sales': recent_sales,
        'price_history': price_history,
    })


@login_required
def sales_rank(request):
    return render(request, 'product/sales_rank.html')


@login_required
@permission_required('product_view')
def sales_rank_data(request):
    data = OrderItem.objects.values('product__name').annotate(total=Sum('quantity')).order_by('-total')[:30]
    return JsonResponse({'data': [{'name': i['product__name'], 'num': i['total']} for i in data]})


@login_required
@permission_required('product_view')
def stock_list(request):
    # 获取搜索关键词
    keyword = request.GET.get('keyword', '')
    # 筛选商品数据
    product_list = Product.objects.filter(name__icontains=keyword).order_by('id')

    # 分页配置：每页10条数据
    paginator = Paginator(product_list, 10)
    # 获取当前页码
    page_number = request.GET.get('page', 1)
    # 获取当前页的商品数据
    page_products = paginator.get_page(page_number)

    # 传递所有前端需要的变量：分页数据、分页器、关键词
    return render(request, 'product/stock.html', {
        'page_products': page_products,  # 匹配前端分页变量名
        'paginator': paginator,  # 传递分页器，用于渲染页码
        'keyword': keyword,  # 传递搜索关键词，保持分页搜索
        'products': page_products  # 保留原变量名，用于表格渲染
    })


# ========== 1. 入库首页（替换开单首页，无客户搜索） ==========
@login_required
@permission_required(PERM_STOCK_IN_CREATE)
def stock_in_index(request):
    """快速入库首页"""
    is_super_admin = request.user.role and request.user.role.code == ROLE_SUPER_ADMIN
    return render(request, 'product/stock_in_index.html', {
        'is_super_admin': is_super_admin
    })


# ========== 2. 保存入库单（核心：增加库存） ==========
@login_required
@permission_required(PERM_STOCK_IN_CREATE)
def save_stock_in(request):
    """保存入库单 + 增加商品库存"""
    if request.method != 'POST':
        return JsonResponse({'code': 0, 'msg': '请求方式错误'})

    try:
        with transaction.atomic():
            data = json.loads(request.body)
            items = data.get('items', [])

            if not items:
                return JsonResponse({'code': 0, 'msg': '请填写入库商品'})

            # 校验商品数据
            product_ids = []
            item_map = {}
            for item in items:
                pid = item.get('id')
                try:
                    pid = int(pid)
                except:
                    return JsonResponse({'code': 0, 'msg': '商品ID格式错误'})

                qty = item.get('qty', 0)
                price = item.get('price', 0)
                if not pid or qty <= 0:
                    return JsonResponse({'code': 0, 'msg': '商品数量必须大于0'})

                product_ids.append(pid)
                item_map[pid] = {'qty': qty, 'price': decimal.Decimal(str(price))}

            # 批量查询商品
            products = Product.objects.filter(id__in=product_ids).in_bulk()
            for pid in product_ids:
                if pid not in products:
                    return JsonResponse({'code': 0, 'msg': f'商品ID {pid} 不存在'})

            # 创建入库单
            stock_in = StockIn()
            stock_in.creator = request.user
            total_amount = 0
            stock_in_items = []

            for pid in product_ids:
                product = products[pid]
                qty = item_map[pid]['qty']
                price = item_map[pid]['price']
                amount = price * qty
                total_amount += amount

                stock_in_items.append(StockInItem(
                    stock_in=stock_in,
                    product=product,
                    quantity=qty,
                    amount=amount,
                    actual_unit_price=price
                ))

            stock_in.total_amount = total_amount
            stock_in.save()
            StockInItem.objects.bulk_create(stock_in_items)

            # ✅ 核心：入库 = 增加库存
            for pid in product_ids:
                products[pid].stock_system += item_map[pid]['qty']
            Product.objects.bulk_update(products.values(), ['stock_system'])

            # 日志+清理缓存
            create_operation_log(request, 'create_stock_in', 'stock_in', str(stock_in.id),
                                 f"入库单-{stock_in.stock_in_no}", "创建入库单")
            clear_stock_cache()
            clear_product_search_cache()

            return JsonResponse({'code': 1, 'msg': '入库成功', 'stock_in_no': stock_in.stock_in_no})

    except Exception as e:
        return JsonResponse({'code': 0, 'msg': f'入库失败：{str(e)}'})


# ========== 3. 入库单列表（替换订单列表） ==========
@login_required
@permission_required(PERM_STOCK_IN_VIEW)
def stock_in_list(request):
    """入库单列表页"""
    stock_in_no = request.GET.get('stock_in_no', '').strip()
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    status = request.GET.get('status', 'all')
    page = request.GET.get('page', 1)

    # 查询集
    stock_ins = StockIn.objects.select_related('creator').order_by('-create_time')

    # 权限控制
    is_super_admin = request.user.role and request.user.role.code == ROLE_SUPER_ADMIN
    if not is_super_admin:
        stock_ins = stock_ins.filter(creator=request.user)

    # 状态筛选
    if status == 'normal':
        stock_ins = stock_ins.filter(status__in=['pending', 'completed'])
    elif status == 'cancelled':
        stock_ins = stock_ins.filter(status='cancelled')

    # 搜索筛选
    if stock_in_no:
        stock_ins = stock_ins.filter(stock_in_no__startswith=stock_in_no)
    if date_from:
        try:
            start = timezone.make_aware(datetime.datetime.strptime(date_from, '%Y-%m-%d'))
            stock_ins = stock_ins.filter(create_time__gte=start)
        except:
            pass
    if date_to:
        try:
            end_date = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()
            end = timezone.make_aware(
                datetime.datetime.combine(end_date + datetime.timedelta(days=1), datetime.datetime.min.time()))
            stock_ins = stock_ins.filter(create_time__lt=end)
        except:
            pass

    # 分页
    paginator = Paginator(stock_ins, 10)
    try:
        page_data = paginator.page(page)
    except:
        page_data = paginator.page(1)

    # 统计
    stats = stock_ins.aggregate(
        total=Count('id'),
        total_amount=Sum('total_amount', default=decimal.Decimal('0.00'))
    )

    # 作废权限
    current_time = timezone.now()
    data_list = list(page_data)
    for item in data_list:
        time_diff = (current_time - item.create_time).total_seconds() / 60
        item.can_cancel = (
                item.status != 'cancelled'
                and is_super_admin
                or (item.creator == request.user and time_diff <= 5)
        )

    context = {
        'stock_ins': data_list,
        'page_data': page_data,
        'paginator': paginator,
        'stock_in_no': stock_in_no,
        'date_from': date_from,
        'date_to': date_to,
        'status': status,
        'total': stats['total'],
        'total_amount': stats['total_amount'],
        'is_super_admin': is_super_admin,
    }
    return render(request, 'product/stock_in_list.html', context)


# ========== 4. 入库单详情（替换订单详情） ==========
@login_required
@permission_required(PERM_STOCK_IN_VIEW)
def stock_in_detail(request, stock_in_no):
    """入库单详情页"""
    stock_in = get_object_or_404(
        StockIn.objects.select_related('creator'),
        stock_in_no=stock_in_no
    )

    # 权限控制
    is_super_admin = request.user.role and request.user.role.code == ROLE_SUPER_ADMIN
    if not is_super_admin and stock_in.creator != request.user:
        return redirect('product:stock_in_list')

    # 作废按钮判断
    current_time = timezone.now()
    time_diff = (current_time - stock_in.create_time).total_seconds() / 60
    show_cancel_btn = (
            stock_in.status != 'cancelled'
            and is_super_admin
            or (stock_in.creator == request.user and time_diff <= 5)
    )

    # 明细
    items = StockInItem.objects.select_related('product').filter(stock_in=stock_in)

    context = {
        'stock_in': stock_in,
        'items': items,
        'is_super_admin': is_super_admin,
        'show_cancel_btn': show_cancel_btn,
    }
    return render(request, 'product/stock_in_detail.html', context)


# ========== 5. 作废入库单（核心：回滚库存） ==========
@login_required
@permission_required(PERM_STOCK_IN_CANCEL)
def cancel_stock_in(request, stock_in_no):
    """作废入库单 + 减少库存"""
    if request.method != 'POST':
        return JsonResponse({'code': 0, 'msg': '仅支持POST'})

    try:
        with transaction.atomic():
            stock_in = get_object_or_404(StockIn, stock_in_no=stock_in_no)

            # 校验
            if stock_in.status == 'cancelled':
                return JsonResponse({'code': 0, 'msg': '已作废，无需重复操作'})

            # 权限
            is_super_admin = request.user.role and request.user.role.code == ROLE_SUPER_ADMIN
            if not is_super_admin and stock_in.creator != request.user:
                return JsonResponse({'code': 0, 'msg': '仅能作废自己的入库单'})

            # 参数
            data = json.loads(request.body)
            reason = data.get('reason', '').strip()
            if not reason:
                return JsonResponse({'code': 0, 'msg': '请填写作废原因'})

            # 更新状态
            stock_in.status = 'cancelled'
            stock_in.cancelled_by = request.user
            stock_in.cancelled_time = timezone.now()
            stock_in.cancelled_reason = reason
            stock_in.save()

            # ✅ 核心：作废入库 = 减少库存
            items = stock_in.items.select_related('product')
            product_list = []
            for item in items:
                if item.product:
                    item.product.stock_system -= item.quantity
                    product_list.append(item.product)
            if product_list:
                Product.objects.bulk_update(product_list, ['stock_system'])

            # 日志+缓存
            create_operation_log(request, 'cancel_stock_in', 'stock_in', str(stock_in.id),
                                 f"入库单-{stock_in.stock_in_no}", f"作废：{reason}")
            clear_stock_cache()

            return JsonResponse({'code': 1, 'msg': '作废成功'})

    except Exception as e:
        return JsonResponse({'code': 0, 'msg': f'作废失败：{str(e)}'})


# ===================== 新增：商品详情统计 API (单个商品) =====================
@login_required
@permission_required(PERM_PRODUCT_DETAIL)
def product_one_statistics_api(request, pk):
    """
    异步统计接口：点击按钮后才计算 (单个商品)
    优化：使用 select_related 关联 order，利用索引
    """
    product = get_object_or_404(Product, pk=pk)

    # 排除作废订单，只统计有效单据
    valid_status = ['pending', 'printed', 'reopened']

    # 利用索引 (product, order) 进行过滤
    items_qs = OrderItem.objects.filter(
        product_id=pk,
        order__status__in=valid_status
    ).select_related('order')  # 减少回表

    stats = items_qs.aggregate(
        total_qty=Coalesce(Sum('quantity'), 0),
        total_amount=Coalesce(Sum('amount'), 0, output_field=DecimalField()),
        count_orders=Count('order', distinct=True),
    )

    return JsonResponse({
        'code': 1,
        'data': {
            'total_qty': stats['total_qty'],
            'total_amount': float(stats['total_amount']),
            'count_orders': stats['count_orders'],
        }
    })


# ===================== 修改：商品统计详情页面视图 (无需PK) =====================
@login_required
@permission_required(PERM_PRODUCT_DETAIL)
def product_statistics_detail(request):
    """全部商品统计页面"""
    all_tags = ProductTag.objects.filter(is_active=True)
    return render(request, 'product/product_statistics.html', {
        'all_tags': all_tags,
    })


@login_required
@permission_required(PERM_PRODUCT_DETAIL)
def product_statistics_api(request):
    tag_ids = request.GET.getlist('tag_ids', [])
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    order_by = request.GET.get('order_by', '-amount')  # 默认按金额降序

    valid_status = ['pending', 'printed', 'reopened']
    items_qs = OrderItem.objects.filter(
        order__status__in=valid_status
    ).select_related('order', 'product')  # 减少回表

    # 时间范围筛选
    if date_from:
        try:
            start = timezone.make_aware(datetime.datetime.strptime(date_from, '%Y-%m-%d'))
            items_qs = items_qs.filter(order__create_time__gte=start)
        except:
            pass
    if date_to:
        try:
            end_date = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()
            end = timezone.make_aware(
                datetime.datetime.combine(end_date + datetime.timedelta(days=1), datetime.datetime.min.time()))
            items_qs = items_qs.filter(order__create_time__lt=end)
        except:
            pass

    # 标签筛选（如果有）
    if tag_ids:
        items_qs = items_qs.filter(product__tags__id__in=tag_ids).distinct()

    # 核心KPI统计
    stats = items_qs.aggregate(
        total_qty=Coalesce(Sum('quantity'), 0),
        total_amount=Coalesce(Sum('amount'), 0, output_field=DecimalField()),
        count_orders=Count('order', distinct=True),
    )

    # 标签占比分析
    all_tags = ProductTag.objects.filter(is_active=True)
    total_orders = stats['count_orders'] or 1

    tag_order_counts = {
        tag['product__tags__id']: tag['order_count']
        for tag in items_qs.filter(
            product__tags__in=all_tags
        ).values('product__tags__id').annotate(
            order_count=Count('order', distinct=True)
        ).values('product__tags__id', 'order_count')
    }

    tag_analysis = []
    for tag in all_tags:
        tag_order_count = tag_order_counts.get(tag.id, 0)
        tag_analysis.append({
            'id': tag.id,
            'name': tag.name,
            'color': tag.color,
            'count': tag_order_count,
            'percentage': round(tag_order_count / total_orders * 100, 2) if total_orders else 0
        })

    # 🔥 修复：销售明细排行（返回订单号、开单时间、数量、金额）
    # 注意：这里直接取订单项明细，不做聚合
    rank_items = items_qs.values(
        'order__order_no',       # 订单号
        'order__create_time',    # 开单时间
        'quantity',              # 销售数量
        'amount'                 # 销售金额
    ).order_by(order_by)[:100]

    return JsonResponse({
        'code': 1,
        'data': {
            'total_qty': stats['total_qty'],
            'total_amount': float(stats['total_amount']),
            'count_orders': stats['count_orders'],
            'tag_analysis': tag_analysis,
            'rank_items': list(rank_items),
        }
    })


@login_required
@permission_required('unit_add')  # 替换为你实际的权限标识
def unit_add(request):
    """新增单位"""
    if request.method != 'POST':
        return JsonResponse({'code': 0, 'msg': '请求方式错误'})

    name = request.POST.get('name', '').strip()
    sort_order = int(request.POST.get('sort_order', 0))

    if not name:
        return JsonResponse({'code': 0, 'msg': '单位名称不能为空'})

    try:
        Unit.objects.create(name=name, sort_order=sort_order)
        return JsonResponse({'code': 1, 'msg': '新增成功'})
    except IntegrityError:
        return JsonResponse({'code': 0, 'msg': '该单位已存在'})


@login_required
@permission_required('unit_edit')
def unit_edit(request):
    """编辑单位"""
    if request.method != 'POST':
        return JsonResponse({'code': 0, 'msg': '请求方式错误'})

    unit_id = int(request.POST.get('id', 0))
    name = request.POST.get('name', '').strip()
    sort_order = int(request.POST.get('sort_order', 0))

    if not name or not unit_id:
        return JsonResponse({'code': 0, 'msg': '参数错误'})

    try:
        unit = Unit.all_objects.get(id=unit_id)
    except Unit.DoesNotExist:
        return JsonResponse({'code': 0, 'msg': '单位不存在'})

    if Unit.all_objects.filter(name=name).exclude(id=unit_id).exists():
        return JsonResponse({'code': 0, 'msg': '单位名称已存在'})

    unit.name = name
    unit.sort_order = sort_order
    unit.save(update_fields=['name', 'sort_order'])
    return JsonResponse({'code': 1, 'msg': '修改成功'})


@login_required
@permission_required('unit_delete')
def unit_toggle_status(request):
    """切换单位启用/禁用（软删除=禁用）"""
    if request.method != 'POST':
        return JsonResponse({'code': 0, 'msg': '请求方式错误'})

    unit_id = int(request.POST.get('id', 0))
    try:
        unit = Unit.all_objects.get(id=unit_id)
    except Unit.DoesNotExist:
        return JsonResponse({'code': 0, 'msg': '单位不存在'})

    unit.is_active = not unit.is_active
    unit.save(update_fields=['is_active'])
    return JsonResponse({'code': 1, 'msg': '状态已更新'})


def export_to_excel_buffer(data, title, headers, selected_fields, custom_fields=None, file_name='导出', total_row=None):
    """
    返回 BytesIO 对象的 Excel 文件
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title

    # 确定最终字段列表（简化，不处理自定义字段位置）
    field_order = selected_fields
    # 若存在自定义字段，简单追加（可根据需求扩展）
    if custom_fields:
        for cf in custom_fields:
            if cf.get('name') not in field_order:
                field_order.append('custom_' + cf['name'])

    # 写入表头
    header_font = Font(bold=True)
    for col_num, field in enumerate(field_order, 1):
        header_text = headers.get(field, field)
        cell = ws.cell(row=1, column=col_num, value=header_text)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # 写入数据
    for row_num, record in enumerate(data, 2):
        for col_num, field in enumerate(field_order, 1):
            value = record.get(field, '')
            ws.cell(row=row_num, column=col_num, value=value)

    # 若有合计行（略）
    if total_row:
        pass

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def export_products_to_io(products=None):
    """
    导出商品数据为 BytesIO 对象（全量字段，包括禁用商品）
    用于一键备份
    """
    if products is None:
        # 使用 all_objects 获取所有商品（包括禁用的）
        products = Product.all_objects.prefetch_related('aliases', 'tags').order_by('name')

    data = []
    seq = 1
    for product in products:
        # 获取别名列表（仅启用别名，若需要禁用的可调整）
        aliases = ','.join([a.alias_name for a in product.aliases.filter(is_active=True)])
        # 获取标签列表（仅启用标签）
        tags = ','.join([tag.name for tag in product.tags.filter(is_active=True)])
        data.append({
            'serial': seq,
            'id': product.id,
            'name': product.name,
            'price': float(product.price) if product.price else 0.0,
            'unit': product.unit,
            'specification': product.specification or '',
            'stock_system': product.stock_system,
            'stock_actual': product.stock_actual,
            'aliases': aliases,
            'status': '启用' if product.is_active else '停用',
            'tags': tags,
        })
        seq += 1

    # 表头映射不变
    headers = {
        'serial': '序号',
        'id': 'ID',
        'name': '商品名称',
        'price': '单价（元）',
        'unit': '单位',
        'specification': '商品规格',
        'stock_system': '系统库存',
        'stock_actual': '实际库存',
        'aliases': '别名',
        'status': '状态',
        'tags': '商品标签'
    }
    selected_fields = ['serial', 'id', 'name', 'price', 'unit', 'specification',
                       'stock_system', 'stock_actual', 'aliases', 'status', 'tags']

    buffer = export_to_excel_buffer(
        data=data,
        title='商品列表',
        headers=headers,
        selected_fields=selected_fields,
        file_name='商品导出'
    )
    return buffer


def import_products_from_io(file_obj, strategy='append'):
    """
    从 BytesIO 对象导入商品数据
    支持自动创建标签（如果不存在）
    策略：'append' 跳过重复（相同名称+单位），'overwrite' 暂不支持
    返回：{'success': int, 'skipped': int, 'errors': list}
    """
    try:
        # 兼容 xlsx 和 xls
        if hasattr(file_obj, 'name') and file_obj.name.endswith('.xls'):
            import xlrd
            wb = xlrd.open_workbook(file_contents=file_obj.read())
            sheet = wb.sheet_by_index(0)
            rows = [sheet.row_values(i) for i in range(sheet.nrows)]
        else:
            wb = openpyxl.load_workbook(file_obj)
            sheet = wb.active
            rows = list(sheet.iter_rows(values_only=True))
    except Exception as e:
        return {'success': 0, 'skipped': 0, 'errors': [f'文件解析失败: {str(e)}']}

    if not rows:
        return {'success': 0, 'skipped': 0, 'errors': ['文件为空']}

    # 映射表头
    header_to_field = {
        '序号': 'serial',
        'ID': 'id',
        '商品名称': 'name',
        '单价（元）': 'price',
        '单位': 'unit',
        '商品规格': 'specification',
        '系统库存': 'stock_system',
        '实际库存': 'stock_actual',
        '别名': 'aliases',
        '状态': 'status',
        '商品标签': 'tags',
    }
    headers = rows[0]
    col_map = {}
    for idx, h in enumerate(headers):
        h = str(h).strip()
        field = header_to_field.get(h)
        if field:
            col_map[field] = idx

    if 'name' not in col_map:
        return {'success': 0, 'skipped': 0, 'errors': ['缺少“商品名称”列']}

    success_count = 0
    fail_count = 0
    errors = []
    new_products = []
    updated_products = []
    tag_cache = {}
    processed_key_map = {}

    for row_idx, row in enumerate(rows[1:], start=2):
        # ---- 新增：跳过完全空行 ----
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue

        try:
            def get_val(field):
                idx = col_map.get(field)
                if idx is not None and idx < len(row):
                    val = row[idx]
                    return str(val).strip() if val is not None else ''
                return ''

            name = get_val('name')
            if not name:
                errors.append(f'第{row_idx}行：商品名称不能为空')
                fail_count += 1
                continue

            unit = get_val('unit') or '件'
            price_str = get_val('price')
            try:
                price = float(price_str) if price_str else 0.0
            except:
                price = 0.0

            specification = get_val('specification')
            stock_system_str = get_val('stock_system')
            try:
                stock_system = int(float(stock_system_str)) if stock_system_str else 0
            except:
                stock_system = 0

            stock_actual_str = get_val('stock_actual')
            try:
                stock_actual = int(float(stock_actual_str)) if stock_actual_str else 0
            except:
                stock_actual = 0

            status_val = get_val('status')
            is_active = status_val != '停用'

            # 处理标签
            tags_str = get_val('tags')
            tag_names = [t.strip() for t in tags_str.split(',') if t.strip()] if tags_str else []
            tag_objs = []
            for tname in tag_names:
                if tname not in tag_cache:
                    tag_obj, created = ProductTag.objects.get_or_create(
                        name=tname,
                        defaults={'color': '#3498db', 'is_active': True}
                    )
                    if not created and not tag_obj.is_active:
                        tag_obj.is_active = True
                        tag_obj.save()
                    tag_cache[tname] = tag_obj
                else:
                    tag_obj = tag_cache[tname]
                tag_objs.append(tag_obj)

            # 查找或创建商品
            key = (name, unit)
            product = None

            # 尝试通过ID更新（若提供）
            id_val = get_val('id')
            if id_val:
                try:
                    pid = int(float(id_val))
                    product = Product.all_objects.filter(id=pid).first()
                except:
                    pass

            if not product and key in processed_key_map:
                product = processed_key_map[key]

            if not product:
                product = Product.objects.filter(name=name, unit=unit).first()
                if not product:
                    product = Product.all_objects.filter(name=name, unit=unit, is_active=False).first()
                if product:
                    processed_key_map[key] = product

            if not product:
                # 新建
                product = Product(
                    name=name,
                    unit=unit,
                    price=price,
                    specification=specification,
                    stock_system=stock_system,
                    stock_actual=stock_actual,
                    is_active=is_active
                )
                new_products.append(product)
                processed_key_map[key] = product
            else:
                # 更新现有
                product.name = name
                product.unit = unit
                product.price = price
                product.specification = specification
                product.stock_system = stock_system
                product.stock_actual = stock_actual
                product.is_active = is_active
                updated_products.append(product)

            # 暂存标签，后续设置
            product._import_tags = tag_objs
            success_count += 1

        except Exception as e:
            fail_count += 1
            errors.append(f'第{row_idx}行：处理错误 - {str(e)}')

    # 持久化
    try:
        with transaction.atomic():
            # 先保存新商品（bulk_create 不触发 save，需单独处理拼音）
            for prod in new_products:
                prod.pinyin_full = ''.join(lazy_pinyin(prod.name, style=0))
                prod.pinyin_abbr = ''.join([p[0] for p in lazy_pinyin(prod.name, style=0)])
            Product.objects.bulk_create(new_products)

            # 更新现有商品
            for prod in updated_products:
                prod.save()

            # 设置标签
            all_products = new_products + updated_products
            for prod in all_products:
                if hasattr(prod, '_import_tags'):
                    prod.tags.set(prod._import_tags)

    except Exception as e:
        logger.error(f"商品导入持久化失败: {str(e)}")
        return {'success': 0, 'skipped': 0, 'errors': [f'数据保存失败: {str(e)}']}

    # 清理缓存
    try:
        from .cache import clear_product_all_cache
        clear_product_all_cache()
    except ImportError:
        pass

    return {
        'success': success_count,
        'skipped': fail_count,
        'errors': errors,
    }

# ---------- 商品审核页面 ----------
@login_required
@permission_required(PERM_PRODUCT_AUDIT)
def product_audit_page(request):
    """渲染商品审核页面"""
    return render(request, 'product/product_audit.html')


@login_required
@permission_required(PERM_PRODUCT_AUDIT)
def product_audit_preview(request):
    """
    返回三类数据：
    1. products: 无标签商品（原有功能）
    2. spec_missing: 规格缺失（单位='件'且规格为空）
    3. duplicate_groups: 同名无规格组（不限单位，每组≥2个商品）
    """
    # ---------- 1. 无标签商品（兼容原有前端） ----------
    products = Product.objects.filter(is_active=True, tags__isnull=True).only(
        'id', 'name', 'unit', 'price', 'specification'
    )
    product_data = []
    for p in products:
        product_data.append({
            'id': p.id,
            'name': p.name,
            'unit': p.unit,
            'price': float(p.price),
            'specification': p.specification or '',
        })

    # ---------- 2. 规格缺失（单位='件'且规格为空） ----------
    empty_spec_condition = Q(specification='') | Q(specification__isnull=True)
    spec_missing = Product.objects.filter(
        is_active=True,
        unit='件'
    ).filter(empty_spec_condition).only('id', 'name', 'unit', 'price', 'specification')

    spec_missing_data = []
    for p in spec_missing:
        spec_missing_data.append({
            'id': p.id,
            'name': p.name,
            'unit': p.unit,
            'price': float(p.price),
            'specification': p.specification or '',
        })

    # ---------- 3. 同名无规格组（不限单位） ----------
    no_spec_products = Product.objects.filter(
        is_active=True
    ).filter(empty_spec_condition)

    name_counts = no_spec_products.values('name').annotate(cnt=Count('id')).filter(cnt__gte=2)
    duplicate_names = [item['name'] for item in name_counts]

    duplicate_products = no_spec_products.filter(name__in=duplicate_names).only(
        'id', 'name', 'unit', 'price', 'specification'
    )

    groups = {}
    for p in duplicate_products:
        groups.setdefault(p.name, []).append({
            'id': p.id,
            'name': p.name,
            'unit': p.unit,
            'price': float(p.price),
            'specification': p.specification or '',
        })

    duplicate_groups = [{'name': name, 'items': items} for name, items in groups.items()]

    # 调试日志（可删除）
    logger.debug(f'无标签商品: {len(product_data)} 个')
    logger.debug(f'规格缺失: {len(spec_missing_data)} 个')
    logger.debug(f'同名组: {len(duplicate_groups)} 组')

    return JsonResponse({
        'code': 1,
        'data': {
            'products': product_data,
            'spec_missing': spec_missing_data,
            'duplicate_groups': duplicate_groups,
        }
    })
# ---------- 规格更新 ----------
@login_required
@permission_required(PERM_PRODUCT_AUDIT)
def product_audit_update_spec(request):
    """
    批量更新商品规格（仅限 unit='件' 且 specification 为空的商品）
    请求体：{"items": [{"product_id": 1, "specification": "新规格"}, ...]}
    """
    if request.method != 'POST':
        return JsonResponse({'code': 0, 'msg': '仅支持POST'})

    try:
        payload = json.loads(request.body)
    except:
        return JsonResponse({'code': 0, 'msg': 'JSON格式错误'})

    items = payload.get('items', [])
    if not items:
        return JsonResponse({'code': 0, 'msg': '未提供有效数据'})

    success_count = 0
    error_msgs = []
    with transaction.atomic():
        for item in items:
            product_id = item.get('product_id')
            spec = item.get('specification', '').strip()
            if not product_id or not spec:
                error_msgs.append(f'商品ID或规格为空')
                continue

            try:
                product = Product.objects.get(
                    id=product_id,
                    is_active=True,
                    unit='件',
                    specification__in=['', None]
                )
            except Product.DoesNotExist:
                error_msgs.append(f'商品ID {product_id} 不存在、已停用、单位不是"件"或已有规格')
                continue

            product.specification = spec
            product.save(update_fields=['specification'])
            success_count += 1

    msg = f'成功更新 {success_count} 个商品的规格。'
    if error_msgs:
        msg += ' 错误：' + '；'.join(error_msgs[:3])
    return JsonResponse({'code': 1, 'msg': msg})

# ---------- 作废同名商品 ----------
@login_required
@permission_required(PERM_PRODUCT_AUDIT)
def product_audit_cancel_duplicate(request):
    """批量作废同名且无规格的商品（仅允许单位='件'且规格为空）"""
    if request.method != 'POST':
        return JsonResponse({'code': 0, 'msg': '仅支持POST'})
    try:
        payload = json.loads(request.body)
    except:
        return JsonResponse({'code': 0, 'msg': 'JSON格式错误'})
    product_ids = payload.get('product_ids', [])
    if not product_ids:
        return JsonResponse({'code': 0, 'msg': '未选择商品'})
    # 仅允许作废符合条件的：单位='件'，规格为空，且属于同名组（即该名称下至少还有另一个符合条件的商品，否则作废后可能只剩一个或零个，但需求允许全部作废？需求说“让用户作废其中的商品”，意味着可以作废全部，但可能保留至少一个。但需求未强制保留，我们允许全部作废）
    # 但需保证这些商品当前是有效的且确实符合条件
    products = Product.objects.filter(
        id__in=product_ids,
        is_active=True,
        unit='件',
        specification__in=['', None]
    )
    if not products.exists():
        return JsonResponse({'code': 0, 'msg': '未找到符合条件的商品'})
    # 进一步校验：这些商品必须属于同名且无规格的组（即其名称下至少有两个符合条件的商品，但为了灵活，可放宽，只要求这些商品本身符合条件即可）
    # 但为了安全，可检查每个商品的名称下是否存在至少两个符合条件的商品（包括自身），但作废后如果少于2个也无所谓，我们允许用户作废任意选择。
    with transaction.atomic():
        count = products.update(is_active=False)
    return JsonResponse({'code': 1, 'msg': f'成功作废 {count} 个商品'})

@login_required
@permission_required(PERM_PRODUCT_AUDIT)
def product_audit_cancel_duplicate(request):
    """
    作废同名无规格组中的选中商品（仅允许作废属于重复组的商品）
    请求体：{"product_ids": [1,2,3]}
    """
    if request.method != 'POST':
        return JsonResponse({'code': 0, 'msg': '仅支持POST'})

    try:
        payload = json.loads(request.body)
    except:
        return JsonResponse({'code': 0, 'msg': 'JSON格式错误'})

    product_ids = payload.get('product_ids', [])
    if not product_ids:
        return JsonResponse({'code': 0, 'msg': '未选择商品'})

    # 所有有效且无规格的商品
    products = Product.objects.filter(id__in=product_ids, is_active=True, specification__in=['', None])
    if not products.exists():
        return JsonResponse({'code': 0, 'msg': '未找到可作废的商品'})

    # 获取这些商品所属的名称列表
    names = set(products.values_list('name', flat=True))

    # 统计每个名称下有效且无规格的商品总数（包括未选中的）
    name_counts = Product.objects.filter(
        name__in=names,
        is_active=True,
        specification__in=['', None]
    ).values('name').annotate(cnt=Count('id'))

    valid_names = {item['name'] for item in name_counts if item['cnt'] >= 2}
    if not valid_names:
        return JsonResponse({'code': 0, 'msg': '所选商品不属于任何重复组，无法作废'})

    # 只作废属于重复组的商品
    to_cancel = products.filter(name__in=valid_names)
    if not to_cancel.exists():
        return JsonResponse({'code': 0, 'msg': '所选商品均不属于重复组，无法作废'})

    count = to_cancel.update(is_active=False)
    return JsonResponse({'code': 1, 'msg': f'成功作废 {count} 个商品'})
# ---------- 添加标签 ----------
@login_required
@permission_required(PERM_PRODUCT_AUDIT)
def product_audit_add_tag(request):
    """批量添加标签，每个商品可指定多个标签（逗号分隔）"""
    if request.method != 'POST':
        return JsonResponse({'code': 0, 'msg': '仅支持POST'})

    try:
        payload = json.loads(request.body)
    except:
        return JsonResponse({'code': 0, 'msg': 'JSON格式错误'})

    items = payload.get('items', [])  # [{"product_id":1, "tag_names":"标签A,标签B"}]
    if not items:
        return JsonResponse({'code': 0, 'msg': '未提供有效数据'})

    success_count = 0
    error_msgs = []
    with transaction.atomic():
        for item in items:
            product_id = item.get('product_id')
            tag_names_str = item.get('tag_names', '').strip()
            if not product_id or not tag_names_str:
                continue
            try:
                product = Product.objects.get(id=product_id, is_active=True)
            except Product.DoesNotExist:
                error_msgs.append(f'商品ID {product_id} 不存在或已作废')
                continue

            # 按逗号分割，去除空格
            tag_names = [name.strip() for name in tag_names_str.split(',') if name.strip()]
            if not tag_names:
                continue

            # 获取或创建标签
            tag_objs = []
            for tname in tag_names:
                tag_obj, created = ProductTag.objects.get_or_create(
                    name=tname,
                    defaults={'color': '#3498db', 'is_active': True}
                )
                tag_objs.append(tag_obj)

            # 添加标签（去重）
            product.tags.add(*tag_objs)
            success_count += 1

    msg = f'成功为 {success_count} 个商品添加标签。'
    if error_msgs:
        msg += ' 错误：' + '；'.join(error_msgs)
    return JsonResponse({'code': 1, 'msg': msg})

# ---------- 作废商品 ----------
@login_required
@permission_required(PERM_PRODUCT_AUDIT)
def product_audit_cancel(request):
    """批量作废商品（设置 is_active=False）"""
    if request.method != 'POST':
        return JsonResponse({'code': 0, 'msg': '仅支持POST'})

    try:
        payload = json.loads(request.body)
    except:
        return JsonResponse({'code': 0, 'msg': 'JSON格式错误'})

    product_ids = payload.get('product_ids', [])
    if not product_ids:
        return JsonResponse({'code': 0, 'msg': '未选择商品'})

    # 只作废有效且无标签的商品（防止误操作）
    products = Product.objects.filter(id__in=product_ids, is_active=True, tags__isnull=True)
    if not products.exists():
        return JsonResponse({'code': 0, 'msg': '未找到可作废的商品'})

    with transaction.atomic():
        count = products.update(is_active=False)
    return JsonResponse({'code': 1, 'msg': f'成功作废 {count} 个商品'})