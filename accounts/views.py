from django.db.models import Q, Sum, Count, Value
from django.db.models.functions import Coalesce
from django.db.models.signals import post_migrate
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout

from django.contrib import messages
from django.db import IntegrityError
import logging
from bill.models import Order, Area


from django.core.cache import cache
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from .models import (
    User, Role, Permission,
    ROLE_SUPER_ADMIN, ROLE_ADMIN, ROLE_OPERATOR
)
from operation_log.models import OperationLog
from django.db.models import DecimalField
from decimal import Decimal

import json
import openpyxl
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.contrib.auth.decorators import login_required

from accounts.models import User, Role

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
import json

logger = logging.getLogger(__name__)

# ========== 缓存常量 ==========
CACHE_ROLE_PERMISSION = 1800
CACHE_SYSTEM_ROLES = 86400  # 角色缓存1天

# ========== 操作常量 ==========
OP_TYPE_LOGIN = 'login'
OP_TYPE_LOGOUT = 'logout'
OP_TYPE_CREATE = 'create'
OP_TYPE_UPDATE = 'update'
OP_TYPE_DELETE = 'delete'
OP_TYPE_RESET_PASSWORD = 'reset_password'
OP_TYPE_CHANGE_PASSWORD = 'change_password'
OP_TYPE_ENABLE_USER = 'enable_user'
OP_TYPE_DISABLE_USER = 'disable_user'
OP_TYPE_UPDATE_ROLE_PERM = 'update_role_permission'

# ========== Excel导出（优化索引匹配） ==========
def export_to_excel(data, title, headers, selected_fields, custom_fields, file_name, total_row=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title

    final_fields = selected_fields.copy()
    final_headers = {field: headers[field] for field in selected_fields}

    if custom_fields:
        for cf in custom_fields:
            cf_name = cf.get('name', '')
            cf_position = cf.get('position', 'after')
            cf_target = cf.get('target', '')
            if not cf_name or not cf_target: continue
            custom_field_key = f'custom_{cf_name.replace(" ", "_")}_{len(final_fields)}'
            final_headers[custom_field_key] = cf_name
            try:
                target_index = final_fields.index(cf_target)
                insert_index = target_index + 1 if cf_position == 'after' else target_index
                final_fields.insert(insert_index, custom_field_key)
            except ValueError:
                final_fields.append(custom_field_key)

    selected_headers = [final_headers[field] for field in final_fields]
    title_font = Font(bold=True, size=12)
    alignment = Alignment(horizontal='center')

    for col, header in enumerate(selected_headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = title_font
        cell.alignment = alignment

    for row, item in enumerate(data, 2):
        for col, field in enumerate(final_fields, 1):
            value = item.get(field, '') if not field.startswith('custom_') else ''
            if isinstance(value, float): value = round(value, 2)
            ws.cell(row=row, column=col, value=value)

    if total_row:
        total_row_num = len(data) + 2
        total_font = Font(bold=True, color="FFFFFF")
        total_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.cell(row=total_row_num, column=1, value="总计").font = total_font
        ws.cell(row=total_row_num, column=1).fill = total_fill
        for col, field in enumerate(final_fields, 1):
            if field in total_row:
                cell = ws.cell(row=total_row_num, column=col, value=round(total_row[field], 2))
                cell.font = total_font
                cell.fill = total_fill
                cell.alignment = Alignment(horizontal='center')

    for col in range(1, len(selected_headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{file_name}.xlsx"'
    return response

# ========== 工具函数 ==========
def get_client_ip(request):
    x_forwarded_for = request.META.get('X-Forwarded-For')
    return x_forwarded_for.split(',')[0] if x_forwarded_for else request.META.get('REMOTE_ADDR', '')

def create_operation_log(request, op_type, obj_type, obj_id=None, obj_name=None, detail=None):
    try:
        # 同步写入保留，轻量系统无压力
        OperationLog.objects.create(
            operator=request.user if request.user.is_authenticated else None,
            operation_time=timezone.now(),
            operation_type=op_type,
            object_type=obj_type,
            object_id=str(obj_id) if obj_id else None,
            object_name=obj_name,
            operation_detail=detail,
            ip_address=get_client_ip(request)
        )
    except Exception as e:
        logger.error(f"日志记录失败：{str(e)}")

# 缓存角色列表（全局复用，消除重复查询）
def get_cached_roles():
    roles = cache.get("system_roles_all")
    if not roles:
        roles = list(Role.objects.all())
        cache.set("system_roles_all", roles, CACHE_SYSTEM_ROLES)
    return roles

# ========== 权限装饰器 ==========
def permission_required(permission_code):
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect(f'/accounts/login/?next={request.path}')
            if request.user.role and request.user.role.code == ROLE_SUPER_ADMIN:
                return view_func(request, *args, **kwargs)
            if not request.user.has_permission(permission_code):
                return redirect('/accounts/no-permission/')
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator

# ========== 认证视图（无性能问题，保留原样） ==========
def login_view(request):
    if request.user.is_authenticated:
        if request.user.force_password_change:
            return redirect('/accounts/force-change-password/')
        return redirect(request.GET.get('next', '/bill/'))

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        user = authenticate(request, username=username, password=password)

        if user and user.is_active:
            login(request, user)
            request.session.update({
                'user_code': user.user_code,
                'user_name': user.name,
                'user_role_code': user.role.code if user.role else '',
                'user_role_name': user.role.name if user.role else '无'
            })
            create_operation_log(
                request=request, op_type=OP_TYPE_LOGIN, obj_type='user',
                obj_id=user.id, obj_name=f"{user.user_code}-{user.username}",
                detail=f"用户登录：编号={user.user_code}，用户名={user.username}"
            )
            if user.force_password_change:
                return redirect('/accounts/force-change-password/')
            return redirect(request.POST.get('next', request.GET.get('next', '/bill/')))
        else:
            messages.error(request, '用户名/密码错误或账户已禁用')
    return render(request, 'accounts/login.html', {'next': request.GET.get('next', '')})

@login_required
def force_change_password(request):
    if not request.user.force_password_change:
        return redirect('/bill/')
    if request.method == 'POST':
        old_pwd = request.POST.get('old_password', '').strip()
        new_pwd = request.POST.get('new_password', '').strip()
        confirm_pwd = request.POST.get('confirm_password', '').strip()
        errors = []
        if not all([old_pwd, new_pwd, confirm_pwd]):
            errors.append('所有密码字段不能为空')
        if new_pwd != confirm_pwd:
            errors.append('两次新密码不一致')
        if len(new_pwd) < 8:
            errors.append('新密码长度至少8位')
        if not request.user.check_password(old_pwd):
            errors.append('原密码输入错误')
        if errors:
            for err in errors:
                messages.error(request, err)
        else:
            request.user.set_password(new_pwd)
            request.user.force_password_change = False
            request.user.save()
            # 清理权限缓存
            cache.delete(f"user_perm_{request.user.id}_*")
            create_operation_log(
                request=request, op_type=OP_TYPE_CHANGE_PASSWORD, obj_type='user',
                obj_id=request.user.id, obj_name=f"{request.user.user_code}-{request.user.username}",
                detail=f"用户强制修改密码：编号={request.user.user_code}"
            )
            login(request, request.user)
            messages.success(request, '密码修改成功！请正常使用系统')
            return redirect('/bill/')
    return render(request, 'accounts/force_change_password.html', {'user': request.user})

def logout_view(request):
    if request.user.is_authenticated:
        create_operation_log(
            request=request, op_type=OP_TYPE_LOGOUT, obj_type='user',
            obj_id=request.user.id, obj_name=f"{request.user.user_code}-{request.user.username}",
            detail=f"用户登出：编号={request.user.user_code}"
        )
    logout(request)
    return redirect('/accounts/login/')

# ========== 用户管理（核心优化） ==========
@login_required
@permission_required('user_view')
def user_list(request):
    """用户列表 - 性能优化版（子查询→LEFT JOIN，命中索引）"""
    keyword = request.GET.get('keyword', '').strip()
    status = request.GET.get('status', 'all')
    page = request.GET.get('page', 1)
    page_size = 10

    # 预加载角色，消除N+1
    queryset = User.objects.select_related('role').all().order_by('-date_joined')

    # 关键词筛选
    if keyword:
        queryset = queryset.filter(
            Q(user_code__icontains=keyword) |
            Q(username__icontains=keyword) |
            Q(phone__icontains=keyword) |
            Q(first_name__icontains=keyword) |
            Q(last_name__icontains=keyword)
        )

    # 状态筛选
    if status == 'active':
        queryset = queryset.filter(is_active=True)
    elif status == 'inactive':
        queryset = queryset.filter(is_active=False)

    is_super_admin = request.user.role and request.user.role.code == ROLE_SUPER_ADMIN
    if is_super_admin:
        # ✅ 修复：移除 is_settled=False，只要不作废都统计
        queryset = queryset.annotate(
            orders=Coalesce(
                Count('created_orders', filter=Q(
                    created_orders__status__in=['pending', 'printed', 'reopened']
                )), Value(0)
            ),
            sales=Coalesce(
                Sum('created_orders__total_amount', filter=Q(
                    created_orders__status__in=['pending', 'printed', 'reopened']
                )), Decimal('0')
            )
        )

    # 分页
    paginator = Paginator(queryset, page_size)
    try:
        users_page = paginator.page(page)
    except PageNotAnInteger:
        users_page = paginator.page(1)
    except EmptyPage:
        users_page = paginator.page(paginator.num_pages)

    # 店铺统计（命中索引）
    shop_stats = {'total_orders': 0, 'total_sales': 0, 'total_cancelled': 0}
    if is_super_admin:
        # ✅ 修复：移除 is_settled=False
        shop_stats = Order.objects.aggregate(
            total_orders=Count('id', filter=Q(status__in=['pending', 'printed', 'reopened'])),
            total_sales=Coalesce(Sum('total_amount', filter=Q(status__in=['pending', 'printed', 'reopened'])), Decimal('0')),
            total_cancelled=Count('id', filter=Q(status='cancelled'))
        )

    # 销售占比
    if is_super_admin:
        total_sales = shop_stats['total_sales']
        for user in users_page:
            user.ratio = round((user.sales / total_sales) * 100) if total_sales > 0 else 0

    return render(request, 'accounts/user_list.html', {
        'users': users_page, 'roles': get_cached_roles(), 'keyword': keyword, 'status': status,
        'current_user': request.user, 'is_super_admin': is_super_admin, 'shop_stats': shop_stats,
        'paginator': paginator, 'page_obj': users_page,
    })

@login_required
@permission_required('user_add')
def user_add(request):
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        user_code = request.POST.get('user_code', '').strip()
        password = request.POST.get('password', '').strip()
        phone = request.POST.get('phone', '').strip()
        role_id = request.POST.get('role_id')
        is_active = request.POST.get('is_active') == 'on'
        is_staff = request.POST.get('is_staff') == 'on'

        if not all([username, user_code, password]):
            messages.error(request, '用户名、用户编号、初始密码不能为空！')
            return render(request, 'accounts/user_form.html', {
                'roles': get_cached_roles(), 'form_data': request.POST, 'is_add': True
            })
        try:
            user = User.objects.create_user(
                username=username, user_code=user_code, password=password,
                phone=phone, is_active=is_active, is_staff=is_staff
            )
            role_name = '无'
            if role_id:
                role = get_object_or_404(Role, id=role_id)
                user.role = role
                user.save()
                role_name = role.name
            # 清理角色缓存
            cache.delete("system_roles_all")
            create_operation_log(
                request=request, op_type=OP_TYPE_CREATE, obj_type='user',
                obj_id=user.id, obj_name=f"{user_code}-{username}",
                detail=f"新增用户：编号={user_code}，角色={role_name}"
            )
            messages.success(request, f'用户 {user_code} - {username} 创建成功！')
            return redirect('/accounts/user-list/')
        except IntegrityError:
            messages.error(request, '用户编号已存在！')
        except Exception as e:
            messages.error(request, f'创建失败：{str(e)}')
    return render(request, 'accounts/user_form.html', {'roles': get_cached_roles(), 'is_add': True})

@login_required
@permission_required('user_edit')
def user_edit(request, user_id):
    user = get_object_or_404(User.objects.select_related('role'), id=user_id)
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        user_code = request.POST.get('user_code', '').strip()
        phone = request.POST.get('phone', '').strip()
        role_id = request.POST.get('role_id')
        is_active = request.POST.get('is_active') == 'on'
        is_staff = request.POST.get('is_staff') == 'on'
        new_password = request.POST.get('new_password', '').strip()

        old_info = {
            'user_code': user.user_code, 'username': user.username, 'phone': user.phone,
            'role': user.role.name if user.role else '无', 'is_active': user.is_active
        }
        try:
            user.username = username
            user.user_code = user_code
            user.phone = phone
            user.is_active = is_active
            user.is_staff = is_staff
            pwd_changed = False
            if new_password:
                user.set_password(new_password)
                pwd_changed = True
                cache.delete(f"user_perm_{user.id}_*")

            new_role_name = old_info['role']
            if role_id:
                role = get_object_or_404(Role, id=role_id)
                user.role = role
                new_role_name = role.name

            user.save()
            cache.delete("system_roles_all")
            create_operation_log(
                request=request, op_type=OP_TYPE_UPDATE, obj_type='user',
                obj_id=user.id, obj_name=f"{user_code}-{username}",
                detail=f"编辑用户：原编号={old_info['user_code']}→新编号={user_code}"
            )
            messages.success(request, f'用户 {user_code} - {username} 修改成功！')
            return redirect('/accounts/user-list/')
        except IntegrityError:
            messages.error(request, '用户编号已存在！')
        except Exception as e:
            messages.error(request, f'修改失败：{str(e)}')
    return render(request, 'accounts/user_form.html', {
        'user': user, 'roles': get_cached_roles(), 'role_id': user.role.id if user.role else '', 'is_edit': True
    })

@login_required
@permission_required('user_view')
def user_detail(request, user_id):
    """用户详情 - 优化N+1查询，命中索引"""
    user = get_object_or_404(User.objects.select_related('role'), id=user_id)
    is_super_admin = request.user.role and request.user.role.code == ROLE_SUPER_ADMIN
    user_stats = {'total_orders': 0, 'total_sales': 0, 'total_cancelled': 0, 'sales_ratio': 0}

    if is_super_admin and user.is_active:
        # ✅ 修复：移除所有 is_settled=False 限制
        stats = Order.objects.aggregate(
            shop_total=Coalesce(Sum('total_amount', filter=Q(status__in=['pending', 'printed', 'reopened'])), 0, output_field=DecimalField()),
            user_orders=Count('id', filter=Q(creator=user, status__in=['pending', 'printed', 'reopened'])),
            user_sales=Coalesce(Sum('total_amount', filter=Q(creator=user, status__in=['pending', 'printed', 'reopened'])), 0, output_field=DecimalField()),
            user_canceled=Count('id', filter=Q(creator=user, status='cancelled'))
        )
        user_stats['total_orders'] = stats['user_orders']
        user_stats['total_sales'] = stats['user_sales']
        user_stats['total_cancelled'] = stats['user_canceled']
        shop_total = stats['shop_total']
        user_stats['sales_ratio'] = round((user_stats['total_sales'] / shop_total) * 100) if shop_total > 0 else 0

    recent_orders = []
    if is_super_admin:
        # 🔥 优化：select_related 消除N+1查询
        recent_orders = Order.objects.filter(
            creator=user,
            status__in=['pending', 'printed', 'reopened', 'cancelled']
        ).select_related('customer', 'area', 'creator').order_by('-create_time')[:15]

    return render(request, 'accounts/user_detail.html', {
        'user': user, 'is_super_admin': is_super_admin, 'user_stats': user_stats,
        'recent_orders': recent_orders, 'roles': get_cached_roles()
    })

@login_required
@permission_required('user_toggle_status')
def user_toggle_status(request, user_id):
    if request.method != 'POST':
        return JsonResponse({'code': 0, 'msg': '仅支持POST请求'})
    try:
        user = get_object_or_404(User.objects.select_related('role'), id=user_id)
        user.is_active = not user.is_active
        user.save()
        op_type = OP_TYPE_ENABLE_USER if user.is_active else OP_TYPE_DISABLE_USER
        status_text = '启用' if user.is_active else '禁用'
        create_operation_log(
            request=request, op_type=op_type, obj_type='user',
            obj_id=user.id, obj_name=f"{user.user_code}-{user.username}",
            detail=f"{status_text}用户：编号={user.user_code}"
        )
        return JsonResponse({'code': 1, 'msg': f'用户 {user.user_code} 已{status_text}！'})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': f'操作失败：{str(e)}'})

@login_required
@permission_required('user_reset_password')
def reset_password(request, user_id):
    if request.method != 'POST':
        return JsonResponse({'code': 0, 'msg': '仅支持POST请求'})
    try:
        user = get_object_or_404(User.objects.select_related('role'), id=user_id)
        temp_pwd = "123456"
        user.set_password(temp_pwd)
        user.force_password_change = True
        user.save()
        cache.delete(f"user_perm_{user.id}_*")
        create_operation_log(
            request=request, op_type=OP_TYPE_RESET_PASSWORD, obj_type='user',
            obj_id=user.id, obj_name=f"{user.user_code}-{user.username}",
            detail=f"重置用户密码：编号={user.user_code}"
        )
        return JsonResponse({'code': 1, 'msg': f'密码已重置为：{temp_pwd}，登录后强制修改！'})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': f'重置失败：{str(e)}'})

@login_required
@permission_required('role_permission_config')
def role_permission_config(request, role_code):
    if not (request.user.role and request.user.role.code == ROLE_SUPER_ADMIN):
        return redirect('/accounts/no-permission/')
    role = get_object_or_404(Role, code=role_code)
    all_roles = Role.objects.filter(code__in=[ROLE_SUPER_ADMIN, ROLE_ADMIN, ROLE_OPERATOR]).order_by('id')
    cache_key = f"role_perms_{role_code}"

    if request.method == 'POST' and 'save' in request.POST:
        try:
            selected_perm_ids = request.POST.getlist('permissions', [])
            if not selected_perm_ids and role.code != ROLE_SUPER_ADMIN:
                messages.error(request, '禁止空权限！')
                return redirect(f'/accounts/role-permission/{role_code}/')
            if role.code == ROLE_SUPER_ADMIN:
                selected_perm_ids = Permission.objects.filter(is_active=True).values_list('id', flat=True)
            role.permissions.clear()
            if selected_perm_ids:
                selected_perms = Permission.objects.filter(id__in=selected_perm_ids)
                role.permissions.add(*selected_perms)
            cache.delete(cache_key)
            cache.delete("system_roles_all")
            create_operation_log(
                request=request, op_type=OP_TYPE_UPDATE_ROLE_PERM, obj_type='role',
                obj_id=role.id, obj_name=role.name, detail=f"配置角色权限：{role.name}"
            )
            messages.success(request, f'角色 {role.name} 权限已保存！')
            return redirect(f'/accounts/role-permission/{role_code}/')
        except Exception as e:
            messages.error(request, f'保存失败：{str(e)}')

    perm_groups = cache.get(cache_key)
    if not perm_groups:
        all_perms = Permission.objects.filter(is_active=True).order_by('category', 'code')
        selected_perm_ids = role.permissions.values_list('id', flat=True)
        if role.code == ROLE_SUPER_ADMIN:
            selected_perm_ids = all_perms.values_list('id', flat=True)
        perm_groups = {}
        for perm in all_perms:
            if perm.category not in perm_groups:
                perm_groups[perm.category] = {'name': perm.get_category_display(), 'permissions': []}
            perm_groups[perm.category]['permissions'].append({
                'id': perm.id, 'code': perm.code, 'name': perm.name, 'description': perm.description,
                'is_active': perm.is_active, 'is_selected': perm.id in selected_perm_ids
            })
        cache.set(cache_key, perm_groups, CACHE_ROLE_PERMISSION)

    return render(request, 'accounts/role_permission_config.html', {
        'current_role': role, 'all_roles': all_roles, 'perm_groups': perm_groups,
        'is_super_admin_role': role.code == ROLE_SUPER_ADMIN,
        'ROLE_SUPER_ADMIN': ROLE_SUPER_ADMIN, 'ROLE_ADMIN': ROLE_ADMIN, 'ROLE_OPERATOR': ROLE_OPERATOR
    })

# ========== 其他视图（无性能问题） ==========
@login_required
def profile(request):
    user = request.user
    if request.method == 'POST':
        user.first_name = request.POST.get('first_name', user.first_name).strip()
        user.last_name = request.POST.get('last_name', user.last_name).strip()
        user.phone = request.POST.get('phone', user.phone).strip()
        user.address = request.POST.get('address', user.address).strip()
        user.email = request.POST.get('email', user.email).strip()
        new_pwd = request.POST.get('new_password', '').strip()
        pwd_changed = False
        if new_pwd:
            old_pwd = request.POST.get('old_password', '').strip()
            if not user.check_password(old_pwd):
                messages.error(request, '原密码错误！')
                return render(request, 'accounts/profile.html', {'user': user})
            user.set_password(new_pwd)
            pwd_changed = True
            cache.delete(f"user_perm_{user.id}_*")
        user.save()
        create_operation_log(
            request=request, op_type=OP_TYPE_UPDATE, obj_type='user',
            obj_id=user.id, obj_name=f"{user.user_code}-{user.username}",
            detail=f"修改个人信息"
        )
        messages.success(request, '个人信息修改成功！')
        if pwd_changed:
            login(request, user)
    return render(request, 'accounts/profile.html', {'user': user})

@login_required
def no_permission(request):
    return render(request, 'accounts/no_permission.html')






# ===================== 用户管理：导入导出新增代码 =====================
@login_required
@permission_required('area_add')  # 可根据实际权限配置修改，如 'user_add'
def user_import(request):
    """
    用户批量导入：
    读取Excel，格式：[序号, 用户编号, 用户名, 姓名, 联系电话, 邮箱, 所属角色, 状态]
    用户编号重复则跳过，默认密码为 123456
    """
    if request.method == 'POST':
        try:
            file = request.FILES.get('file')
            if not file:
                return JsonResponse({'code': 0, 'msg': '请选择文件'})

            wb = openpyxl.load_workbook(file)
            ws = wb.active

            imported_count = 0
            skipped_count = 0

            # 预加载所有角色到内存，加速匹配
            role_map = {r.name: r for r in Role.objects.only('id', 'name')}

            # 从第2行开始遍历（第1行是表头）
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if len(row) < 3:
                    continue

                # 提取数据，忽略第一列序号
                user_code = str(row[1]).strip() if row[1] else ''
                username = str(row[2]).strip() if row[2] else ''
                name = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                phone = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                email = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                role_name = str(row[6]).strip() if len(row) > 6 and row[6] else ''
                status_str = str(row[7]).strip() if len(row) > 7 and row[7] else '正常'

                if not user_code or not username:
                    skipped_count += 1
                    continue

                # 检查用户编号是否已存在
                if User.objects.filter(user_code=user_code).exists():
                    skipped_count += 1
                    continue

                # 查找角色
                role = role_map.get(role_name) if role_name else None

                # 解析状态
                is_active = status_str != '禁用'

                # 拆分姓名为 first_name 和 last_name
                first_name = name[:1] if name else ''
                last_name = name[1:] if len(name) > 1 else ''

                # 创建新用户
                user = User.objects.create(
                    user_code=user_code,
                    username=username,
                    first_name=first_name,
                    last_name=last_name,
                    phone=phone,
                    email=email,
                    role=role,
                    is_active=is_active,
                    force_password_change=True
                )
                user.set_password('123456')
                user.save()

                imported_count += 1

            # 记录日志
            create_operation_log(
                request=request, op_type='import', obj_type='user',
                obj_id=0, obj_name='批量导入',
                detail=f"导入成功：新增{imported_count}条，跳过{skipped_count}条重复"
            )

            return JsonResponse({
                'code': 1,
                'msg': f'导入完成！新增 {imported_count} 条，跳过 {skipped_count} 条重复数据'
            })

        except Exception as e:
            logger.error(f"导入用户失败：{str(e)}", exc_info=True)
            return JsonResponse({'code': 0, 'msg': f'导入失败：文件格式错误或数据异常'})
    return JsonResponse({'code': 0, 'msg': '仅支持POST请求'})


@login_required
@permission_required('area_view')  # 可根据实际权限配置修改，如 'user_view'
def user_export(request):
    """
    用户批量导出（支持字段选择和自定义字段）
    """
    try:
        if request.method == 'POST':
            # 1. 获取选中的字段
            selected_fields = request.POST.getlist('fields[]')
            if not selected_fields:
                return JsonResponse({'code': 0, 'msg': '请至少选择一个导出字段'})

            # 2. 获取自定义字段
            custom_fields_json = request.POST.get('custom_fields', '[]')
            try:
                custom_fields = json.loads(custom_fields_json)
            except json.JSONDecodeError:
                custom_fields = []

            # 3. 定义表头映射
            headers = {
                'serial': '序号',
                'id': 'ID',
                'user_code': '用户编号',
                'username': '用户名',
                'name': '姓名',
                'phone': '联系电话',
                'email': '邮箱',
                'role': '所属角色',
                'status': '状态',
                'create_time': '创建时间'
            }

            # 4. 查询并格式化数据
            users = User.objects.select_related('role').only(
                'id', 'user_code', 'username', 'first_name', 'last_name',
                'phone', 'email', 'role', 'is_active', 'date_joined'
            ).order_by('id')

            data = []
            seq = 1
            for user in users:
                data.append({
                    'serial': seq,
                    'id': user.id,
                    'user_code': user.user_code,
                    'username': user.username,
                    'name': user.name,
                    'phone': user.phone or '',
                    'email': user.email or '',
                    'role': user.role.name if user.role else '未分配',
                    'status': '正常' if user.is_active else '禁用',
                    'create_time': user.date_joined.strftime('%Y-%m-%d %H:%M:%S') if user.date_joined else ''
                })
                seq += 1

            # 5. 生成文件名
            date_str = timezone.now().strftime('%Y年%m月%d日')
            file_name = f'{date_str}用户管理导出'

            # 6. 调用通用导出函数
            response = export_to_excel(
                data=data,
                title='用户列表',
                headers=headers,
                selected_fields=selected_fields,
                custom_fields=custom_fields,
                file_name=file_name,
                total_row=None
            )

            # 7. 记录日志
            create_operation_log(
                request=request, op_type='export', obj_type='user',
                obj_id=0, obj_name='批量导出', detail=f"导出用户数据共{len(data)}条"
            )

            return response
        else:
            return JsonResponse({'code': 0, 'msg': '请求方式错误'})
    except Exception as e:
        logger.error(f"导出用户失败：{str(e)}", exc_info=True)
        return JsonResponse({'code': 0, 'msg': '导出失败'})