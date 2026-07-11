# customer_manage\views.py
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.core.cache import cache

from accounts.models import ROLE_SUPER_ADMIN, PERM_LOG_VIEW_ALL
from bill.models import OrderItem, Order
from product.models import Product, ProductAlias
from customer_manage.models import Customer, CustomerPrice, RepaymentRecord, CustomerPhone
from area_manage.models import Area

import datetime
from django.contrib.auth.decorators import login_required
from accounts.views import permission_required, create_operation_log

from django.db.models import Sum, F, Q, OuterRef, Subquery, DecimalField, Max, Count
from django.db.models.functions import Coalesce
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage

from django.utils import timezone
import json
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import openpyxl

from django.db import transaction, models
from django.db.models import F, Sum
from decimal import Decimal

# ========== 缓存时长常量配置 ==========
CACHE_HIGH_PRIORITY = 300  # 复杂聚合查询 5分钟
CACHE_MID_PRIORITY = 600  # 静态数据/搜索接口 10分钟

# ========== 全局统一缓存 Key 定义 ==========
CACHE_KEY_AREA_LIST = "global:area_list"  # 统一区域列表缓存
CACHE_PREFIX_CUSTOMER_LIST = "customer_list_"
CACHE_PREFIX_CUSTOMER_DETAIL = "customer_detail_"
CACHE_PREFIX_CUSTOMER_PRICE = "customer_price_"
# 新增：辅助接口缓存前缀
CACHE_PREFIX_PRODUCT_LIST_FOR_PRICE = "product_list_for_price_"
CACHE_PREFIX_SEARCH_CUSTOMER_FOR_PRICE = "search_customer_for_price_"
CACHE_PREFIX_SEARCH_PRODUCT_FOR_PRICE = "search_product_for_price_"
CACHE_PREFIX_CUSTOMER_SALES_RANK = "customer_sales_rank_"

import logging

logger = logging.getLogger(__name__)


# ========== 统一时间格式化工具函数 ==========
def format_datetime(dt, fmt='%Y-%m-%d %H:%M:%S'):
    """统一时间格式化 - 先转为上海本地时区再格式化"""
    if not dt:
        return ''
    return timezone.localtime(dt).strftime(fmt)


# ========== 统一缓存清理函数 ==========
def clear_customer_cache(customer_id: int = None):
    """
    清理客户相关缓存（含客户列表、详情、消费排行）
    """
    # 1. 清理所有客户列表缓存
    cache.delete_pattern(f"{CACHE_PREFIX_CUSTOMER_LIST}*")

    # 2. 清理指定客户的详情缓存
    if customer_id:
        cache.delete_pattern(f"{CACHE_PREFIX_CUSTOMER_DETAIL}{customer_id}*")

    # 3. 清理客户消费排行缓存（数据变化会影响排行）
    cache.delete_pattern(f"{CACHE_PREFIX_CUSTOMER_SALES_RANK}*")

    logger.info(f"已清理客户缓存: {customer_id if customer_id else '全列表'}")


def clear_customer_price_cache():
    """
    清理客户专属价格相关缓存（含列表、商品/客户搜索）
    """
    # 1. 清理客户专属价格列表缓存
    cache.delete_pattern(f"{CACHE_PREFIX_CUSTOMER_PRICE}*")

    # 2. 清理价格页辅助接口缓存
    cache.delete_pattern(f"{CACHE_PREFIX_PRODUCT_LIST_FOR_PRICE}*")
    cache.delete_pattern(f"{CACHE_PREFIX_SEARCH_CUSTOMER_FOR_PRICE}*")
    cache.delete_pattern(f"{CACHE_PREFIX_SEARCH_PRODUCT_FOR_PRICE}*")

    logger.info(f"已清理客户专属价格全量缓存")


def full_to_half(s):
    """将全角字符转换为半角"""
    if not s:
        return s
    result = []
    for char in s:
        code_point = ord(char)
        if code_point == 0x3000:
            code_point = 0x20
        elif 0xFF01 <= code_point <= 0xFF5E:
            code_point -= 0xFEE0
        result.append(chr(code_point))
    return ''.join(result)


# ========== 客户列表（手动缓存） ==========
@login_required
@permission_required('customer_view')
def customer_list(request):
    """极简版：仅返回客户基本信息，零算力消耗；支持 with_debt 参数返回欠款"""
    try:
        keyword = request.GET.get('keyword', '').strip()
        status = request.GET.get('status', 'all')
        page = request.GET.get('page', 1)
        page_size = 10
        with_debt = request.GET.get('with_debt', '0') == '1'

        cache_key = f"{CACHE_PREFIX_CUSTOMER_LIST}{request.user.id}_{keyword}_{status}_{page}_{with_debt}"
        cached_data = cache.get(cache_key)
        if cached_data:
            return JsonResponse(cached_data, safe=False)

        # ✅ 修改：预加载电话关联，避免N+1查询
        customers = Customer.all_objects.all().select_related('area').prefetch_related('phones')

        # 状态筛选
        if status == 'active':
            customers = customers.filter(is_active=True)
        elif status == 'disabled':
            customers = customers.filter(is_active=False)

        # 关键词筛选
        if keyword:
            id_q = Q()
            if keyword.isdigit():
                id_q = Q(id=int(keyword))
            # ✅ 修改：电话筛选改为关联子表查询 + 去重
            customers = customers.filter(
                Q(name__icontains=keyword) |
                Q(pinyin_full__icontains=keyword) |  # 新增
                Q(pinyin_abbr__icontains=keyword) |  # 新增
                Q(phones__phone__icontains=keyword) |
                id_q |
                Q(area__name__icontains=keyword)
            ).distinct()

        # 统计数量
        all_count = Customer.all_objects.count()
        active_count = Customer.objects.count()
        disabled_count = all_count - active_count

        # 分页
        paginator = Paginator(customers, page_size)
        try:
            customer_page = paginator.page(page)
        except PageNotAnInteger:
            customer_page = paginator.page(1)
        except EmptyPage:
            customer_page = paginator.page(paginator.num_pages)

        # 批量计算欠款（仅需2次查询，替代N*2次）
        unpaid_dict = {}
        paid_dict = {}
        if with_debt:
            customer_ids = [c.id for c in customer_page]

            # 批量查询未付款金额
            unpaid_orders = Order.objects.filter(
                customer_id__in=customer_ids,
                status__in=['pending', 'printed', 'reopened'],
                is_settled=False
            ).values('customer_id').annotate(
                total=Sum(F('total_amount') - F('received_amount'), output_field=models.DecimalField())
            )
            unpaid_dict = {item['customer_id']: item['total'] or 0 for item in unpaid_orders}

            # 批量查询已还款金额
            paid_records = RepaymentRecord.objects.filter(
                customer_id__in=customer_ids
            ).values('customer_id').annotate(
                total=Sum('repayment_amount')
            )
            paid_dict = {item['customer_id']: item['total'] or 0 for item in paid_records}

        # 组装返回数据
        result = []
        for c in customer_page:
            item = {
                'id': c.id,
                'name': c.name,
                'area_id': c.area.id if c.area else '',
                'area_name': c.area.name if c.area else '',
                # ✅ 修改：使用主号码属性替代原 phone 字段
                'phone': c.primary_phone,
                'remark': c.remark or '',
                'is_active': c.is_active,
                'page': int(page),
                'total': paginator.count,
                'counts': {
                    'all': all_count,
                    'active': active_count,
                    'disabled': disabled_count
                }
            }

            # 从预查询的字典中取欠款数据
            if with_debt:
                unpaid_amount = unpaid_dict.get(c.id, 0)
                paid_amount = paid_dict.get(c.id, 0)
                item['total_debt'] = max(float(unpaid_amount) - float(paid_amount), 0)
            else:
                item['total_debt'] = 0.00

            result.append(item)

        cache.set(cache_key, result, CACHE_HIGH_PRIORITY if not with_debt else 10)
        return JsonResponse(result, safe=False)
    except Exception as e:
        logger.error(f"客户列表查询失败: {str(e)}", exc_info=True)
        return JsonResponse({'code': 0, 'msg': f'查询失败：{str(e)}'})


# ========== 启用客户 ==========
@login_required
@permission_required('customer_delete')  # 复用删除权限
def customer_enable(request, pk):
    """启用客户接口"""
    try:
        customer = get_object_or_404(Customer.all_objects, pk=pk)
        customer_name = customer.name

        # 启用操作
        customer.is_active = True
        customer.disabled_time = None
        customer.save()

        create_operation_log(
            request=request,
            op_type='enable',
            obj_type='customer',
            obj_id=pk,
            obj_name=customer_name,
            detail=f"启用客户：{customer_name}"
        )

        clear_customer_cache(customer_id=pk)
        return JsonResponse({'code': 1, 'msg': '启用客户成功'}, content_type='application/json')
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': f'启用失败：{str(e)}'}, content_type='application/json')


# ========== 客户详情（手动缓存） ==========
@login_required
@permission_required('customer_view')
def customer_detail(request, pk):
    """客户详情接口 - 手动缓存版"""
    try:
        settle_status = request.GET.get('settle_status', 'all')
        page = request.GET.get('page', 1)
        cache_key = f"{CACHE_PREFIX_CUSTOMER_DETAIL}{pk}_{settle_status}_{page}"

        cached_data = cache.get(cache_key)
        if cached_data:
            return JsonResponse(cached_data, safe=False)

        customer = get_object_or_404(Customer, pk=pk)
        page_size = 10

        base_orders = Order.objects.filter(
            customer=customer,
            status__in=['pending', 'printed', 'reopened']
        ).select_related('creator__role')

        if settle_status == 'settled':
            orders_query = base_orders.filter(is_settled=True)
        elif settle_status == 'unsettled':
            orders_query = base_orders.filter(is_settled=False)
        else:
            orders_query = base_orders

        unpaid_orders = base_orders.filter(is_settled=False)
        unpaid_order_count = unpaid_orders.count()
        # 精准计算：总欠款 = Sum(订单总额 - 已收金额)
        unpaid_amount = unpaid_orders.aggregate(
            total=Sum(F('total_amount') - F('received_amount'), output_field=models.DecimalField())
        )['total'] or 0
        paid_amount = RepaymentRecord.objects.filter(customer=customer).aggregate(total=Sum('repayment_amount'))[
                          'total'] or 0
        total_debt = max(float(unpaid_amount) - float(paid_amount), 0)

        orders_query = orders_query.order_by('-create_time')
        paginator = Paginator(orders_query, page_size)
        try:
            order_page = paginator.page(page)
        except PageNotAnInteger:
            order_page = paginator.page(1)
        except EmptyPage:
            order_page = paginator.page(paginator.num_pages)

        order_list = []
        for order in order_page:
            order_list.append({
                'order_no': order.order_no or '',
                'create_time': format_datetime(order.create_time, '%Y-%m-%d %H:%M'),
                'total_amount': float(order.total_amount) if order.total_amount else 0.0,
                'is_settled': order.is_settled,
                'status': order.status,
                'status_text': dict(Order.ORDER_STATUS).get(order.status, '未知'),
                'overdue_days': order.get_overdue_days(),
                'order_date': format_datetime(order.create_time, '%Y-%m-%d'),
                'creator_name': order.creator.username if order.creator else '未知',
                'creator_role': order.creator.role.name if (order.creator and order.creator.role) else '未知'
            })

        repayment_list = []
        repayments = RepaymentRecord.objects.filter(customer=customer).select_related('operator__role').order_by(
            '-repayment_time')[:100]
        for repay in repayments:
            repayment_list.append({
                'id': repay.id,
                'repayment_amount': float(repay.repayment_amount) if repay.repayment_amount else 0.0,
                'repayment_time': format_datetime(repay.repayment_time, '%Y-%m-%d %H:%M'),
                'repayment_remark': repay.repayment_remark or '',
                'operator': repay.operator.username if repay.operator else '未知',
                'operator_role': repay.operator.role.name if (repay.operator and repay.operator.role) else '未知',
                'create_time': format_datetime(repay.create_time, '%Y-%m-%d %H:%M')
            })

        product_stats = OrderItem.objects.filter(
            order__customer=customer, product__isnull=False
        ).values(
            'product__id', 'product__name', 'product__unit'
        ).annotate(
            total_quantity=Coalesce(Sum('quantity'), 0),
            last_purchase_time=Coalesce(Max('order__create_time'), None)
        ).order_by('-total_quantity')

        product_stats_list = [{
            'product_name': stat['product__name'],
            'total_quantity': stat['total_quantity'],
            'unit': stat['product__unit'],
            'last_purchase_time': format_datetime(stat['last_purchase_time'], '%Y-%m-%d') if stat[
                'last_purchase_time'] else '无'
        } for stat in product_stats]

        response_data = {
            'code': 1, 'msg': '查询成功',
            'customer_info': {
                'id': customer.id, 'name': customer.name,
                'area_name': customer.area.name if customer.area else '',
                # ✅ 修改：使用主号码
                'phone': customer.primary_phone,
                'remark': customer.remark or ''
            },
            'debt_info': {
                'total_debt': total_debt, 'unpaid_order_count': unpaid_order_count,
                'unpaid_amount': float(unpaid_amount), 'paid_amount': float(paid_amount)
            },
            'orders': order_list,
            'current_page': order_page.number,
            'total_pages': paginator.num_pages,
            'total_orders': paginator.count,
            'repayments': repayment_list,
            'product_stats': product_stats_list
        }

        cache.set(cache_key, response_data, CACHE_HIGH_PRIORITY)
        return JsonResponse(response_data, safe=False)

    except Exception as e:
        logger.error(f"客户详情查询失败: {str(e)}", exc_info=True)
        return JsonResponse({'code': 0, 'msg': f'查询失败：{str(e)}'}, safe=False)


# ========== 还款登记 ==========
@login_required
@permission_required('customer_repayment')
def repayment_register(request):
    if request.method == 'POST':
        try:
            customer_id = request.POST.get('customer_id', '').strip()
            repayment_amount = request.POST.get('repayment_amount', '').strip()
            repayment_time = request.POST.get('repayment_time', '').strip()
            repayment_remark = request.POST.get('repayment_remark', '').strip()

            if not customer_id or not repayment_amount:
                return JsonResponse({'code': 0, 'msg': '客户和还款金额不能为空'}, content_type='application/json')

            # 统一使用 Decimal 计算，避免浮点精度问题
            try:
                repayment_amount = Decimal(repayment_amount)
                if repayment_amount <= 0:
                    return JsonResponse({'code': 0, 'msg': '还款金额必须大于0'}, content_type='application/json')
            except:
                return JsonResponse({'code': 0, 'msg': '还款金额必须是数字'}, content_type='application/json')

            customer = get_object_or_404(Customer, id=customer_id)

            # 时间解析：前端传入本地时间，标记为上海时区后存入数据库
            if repayment_time:
                try:
                    if 'T' in repayment_time:
                        dt = datetime.datetime.strptime(repayment_time, '%Y-%m-%dT%H:%M')
                    else:
                        dt = datetime.datetime.strptime(repayment_time, '%Y-%m-%d %H:%M')
                    repayment_time = timezone.make_aware(dt)
                except Exception as e:
                    return JsonResponse({'code': 0, 'msg': '还款时间格式错误，请重新选择'},
                                        content_type='application/json')
            else:
                repayment_time = timezone.now()

            with transaction.atomic():
                repayment = RepaymentRecord.objects.create(
                    customer=customer,
                    repayment_amount=repayment_amount,
                    repayment_time=repayment_time,
                    repayment_remark=repayment_remark,
                    operator=request.user if request.user.is_authenticated else None
                )

                remaining_to_allocate = repayment_amount

                unsettled_orders = Order.objects.filter(
                    customer=customer,
                    is_settled=False,
                    status__in=['pending', 'printed', 'reopened']
                ).select_for_update().order_by('create_time')

                for order in unsettled_orders:
                    if remaining_to_allocate <= 0:
                        break

                    order_unpaid = order.total_amount - order.received_amount
                    if order_unpaid <= 0:
                        continue

                    if remaining_to_allocate >= order_unpaid:
                        order.received_amount = order.total_amount
                        order.is_settled = True
                        order.settled_by = request.user
                        order.settled_time = timezone.now()
                        order.settled_remark = f"系统自动核销（还款ID:{repayment.id}）"
                        remaining_to_allocate -= order_unpaid
                    else:
                        order.received_amount += remaining_to_allocate
                        remaining_to_allocate = Decimal('0')
                    order.save()

            create_operation_log(
                request=request,
                op_type='repayment_register',
                obj_type='repayment',
                obj_id=repayment.id,
                obj_name=f'{customer.name} - 还款¥{repayment_amount}',
                detail=f"登记还款¥{repayment_amount}，系统自动核销订单"
            )

            clear_customer_cache(customer_id=int(customer_id))
            return JsonResponse({'code': 1, 'msg': '还款登记成功！系统已自动核销旧账'}, content_type='application/json')
        except Exception as e:
            logger.error(f"还款登记失败: {str(e)}", exc_info=True)
            return JsonResponse({'code': 0, 'msg': f'登记失败：{str(e)}'}, content_type='application/json')
    return JsonResponse({'code': 0, 'msg': '仅支持POST请求'}, content_type='application/json')


# ========== 页面入口 ==========
@login_required
@permission_required('customer_view')
def customer_detail_page(request, pk):
    """客户详情页面"""
    return render(request, 'customer_manage/customer_detail.html', {'customer_id': pk})


@login_required
@permission_required('customer_repayment')
def repayment_page(request):
    """还款登记页面"""
    return render(request, 'customer_manage/repayment.html')


# ========== 新增客户 ==========
@login_required
@permission_required('customer_add')
def customer_add(request):
    """新增客户接口（电话选填）"""
    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            area_id = request.POST.get('area_id', '').strip()
            phone = request.POST.get('phone', '').strip()
            remark = request.POST.get('remark', '').strip()

            if not name:
                return JsonResponse({'code': 0, 'msg': '客户名称不能为空'}, content_type='application/json')
            if not area_id:
                return JsonResponse({'code': 0, 'msg': '所属区域不能为空'}, content_type='application/json')
            # ✅ 移除：电话必填校验，改为选填

            if Customer.objects.filter(name=name).exists():
                return JsonResponse({'code': 0, 'msg': '客户名称已存在'}, content_type='application/json')

            area = get_object_or_404(Area, id=area_id)
            customer = Customer.objects.create(
                name=name,
                area=area,
                remark=remark
            )

            # ✅ 修改：仅当电话非空时，才创建主号码记录
            if phone:
                CustomerPhone.objects.create(
                    customer=customer,
                    phone=phone.strip(),
                    is_primary=True
                )

            create_operation_log(
                request=request,
                op_type='create',
                obj_type='customer',
                obj_id=customer.id,
                obj_name=customer.name,
                detail=f"新增客户：名称={customer.name}，主电话={phone if phone else '无'}"
            )

            clear_customer_cache()
            return JsonResponse({'code': 1, 'msg': '新增客户成功'}, content_type='application/json')
        except Exception as e:
            logger.error(f"新增客户失败: {str(e)}", exc_info=True)
            return JsonResponse({'code': 0, 'msg': f'新增失败：{str(e)}'}, content_type='application/json')
    return JsonResponse({'code': 0, 'msg': '仅支持POST请求'}, content_type='application/json')


# ========== 编辑客户 ==========
@login_required
@permission_required('customer_edit')
def customer_edit(request, pk):
    """编辑客户接口（支持清空电话）"""
    try:
        customer = get_object_or_404(Customer.all_objects, pk=pk)
        if request.method == 'POST':
            name = request.POST.get('name', '').strip()
            area_id = request.POST.get('area_id', '').strip()
            phone = request.POST.get('phone', '').strip()
            remark = request.POST.get('remark', '').strip()

            if not name:
                return JsonResponse({'code': 0, 'msg': '客户名称不能为空'}, content_type='application/json')
            if not area_id:
                return JsonResponse({'code': 0, 'msg': '所属区域不能为空'}, content_type='application/json')
            # ✅ 移除：电话必填校验，改为选填

            if Customer.objects.filter(name=name).exclude(pk=pk).exists():
                return JsonResponse({'code': 0, 'msg': '客户名称已存在'}, content_type='application/json')

            area = get_object_or_404(Area, id=area_id)
            customer.name = name
            customer.area = area
            customer.remark = remark
            customer.save()

            # ✅ 修改：处理主号码，支持清空
            primary_phone = customer.phones.filter(is_primary=True).first()
            if phone:
                # 有新电话：更新或创建主号码
                if primary_phone:
                    primary_phone.phone = phone.strip()
                    primary_phone.is_active = True
                    primary_phone.save()
                else:
                    CustomerPhone.objects.create(
                        customer=customer,
                        phone=phone.strip(),
                        is_primary=True
                    )
            else:
                # 电话为空：删除原主号码（物理删除，避免冗余数据）
                if primary_phone:
                    primary_phone.delete()

            create_operation_log(
                request=request,
                op_type='update',
                obj_type='customer',
                obj_id=customer.id,
                obj_name=customer.name,
                detail=f"编辑客户信息，主电话更新为：{phone if phone else '无'}"
            )

            clear_customer_cache(customer_id=pk)
            return JsonResponse({'code': 1, 'msg': '编辑客户成功'}, content_type='application/json')
        return JsonResponse({'code': 0, 'msg': '仅支持POST请求'}, content_type='application/json')
    except Exception as e:
        logger.error(f"编辑客户失败: {str(e)}", exc_info=True)
        return JsonResponse({'code': 0, 'msg': f'编辑失败：{str(e)}'}, content_type='application/json')


# ========== 禁用客户（软删除） ==========
@login_required
@permission_required('customer_delete')
def customer_delete(request, pk):
    """禁用客户接口（软删除）"""
    try:
        customer = get_object_or_404(Customer, pk=pk)
        customer_name = customer.name

        # 软删除操作
        customer.is_active = False
        customer.disabled_time = timezone.now()
        customer.save()

        create_operation_log(
            request=request,
            op_type='disable',
            obj_type='customer',
            obj_id=pk,
            obj_name=customer_name,
            detail=f"禁用客户：{customer_name}"
        )

        clear_customer_cache(customer_id=pk)
        return JsonResponse({'code': 1, 'msg': '禁用客户成功'}, content_type='application/json')
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': f'禁用失败：{str(e)}'}, content_type='application/json')


# ========== 区域列表（共用缓存） ==========
@login_required
@permission_required('customer_view')
def area_list_for_customer(request):
    """供客户管理页面获取区域下拉列表 - 手动缓存版（共用）"""
    try:
        cached_data = cache.get(CACHE_KEY_AREA_LIST)
        if cached_data:
            return JsonResponse(cached_data, safe=False, content_type='application/json')

        areas = Area.objects.all().order_by('name')
        result = [{'id': a.id, 'name': a.name} for a in areas]
        cache.set(CACHE_KEY_AREA_LIST, result, CACHE_MID_PRIORITY)

        return JsonResponse(result, safe=False, content_type='application/json')
    except Exception as e:
        return JsonResponse(
            {'code': 0, 'msg': f'查询区域失败：{str(e)}'},
            safe=False,
            content_type='application/json'
        )


@login_required
@permission_required('customer_view')
def customer_page(request):
    """客户管理页面"""
    return render(request, 'customer_manage/customer.html')


# ========== 客户专属价格列表 ==========
@login_required
@permission_required('customer_price_view')
def customer_price_list(request):
    """获取客户专属价格列表 - 手动缓存版 + 状态筛选"""
    try:
        keyword = request.GET.get('keyword', '').strip()
        min_price = request.GET.get('min_price', '').strip()
        max_price = request.GET.get('max_price', '').strip()
        area_id = request.GET.get('area_id', '').strip()
        status = request.GET.get('status', 'all')
        page = request.GET.get('page', 1)

        cache_key = f"{CACHE_PREFIX_CUSTOMER_PRICE}{request.user.id}_{keyword}_{min_price}_{max_price}_{area_id}_{status}_{page}"
        cached_data = cache.get(cache_key)
        if cached_data:
            logger.info(f"命中专属价格缓存: {cache_key}")
            return JsonResponse(cached_data, safe=False, content_type='application/json')

        page_size = 15
        prices = CustomerPrice.all_objects.all() \
            .select_related('customer__area', 'product') \
            .prefetch_related('product__aliases')

        if not request.user.has_permission(PERM_LOG_VIEW_ALL):
            pass

        # 状态筛选
        if status == 'active':
            prices = prices.filter(is_active=True)
        elif status == 'disabled':
            prices = prices.filter(is_active=False)

        if keyword:
            keyword = full_to_half(keyword).strip()
            keywords = [k for k in keyword.split() if k]
            base_q = Q()
            for kw in keywords:
                customer_q = Q(customer__name__icontains=kw) | \
                             Q(customer__pinyin_full__icontains=kw) | \
                             Q(customer__pinyin_abbr__icontains=kw)
                if kw.isdigit():
                    customer_q |= Q(customer__id=int(kw))
                product_q = Q(product__name__icontains=kw)
                if kw.isdigit():
                    product_q |= Q(product__id=int(kw))
                alias_ids = ProductAlias.objects.filter(
                    Q(alias_name__icontains=kw) |
                    Q(alias_pinyin_full__icontains=kw) |
                    Q(alias_pinyin_abbr__icontains=kw)
                ).values_list('product_id', flat=True)
                if alias_ids:
                    product_q |= Q(product__id__in=alias_ids)
                kw_q = customer_q | product_q
                base_q &= kw_q
            prices = prices.filter(base_q)

        if min_price:
            try:
                min_price = float(min_price)
                prices = prices.filter(custom_price__gte=min_price)
            except:
                pass
        if max_price:
            try:
                max_price = float(max_price)
                prices = prices.filter(custom_price__lte=max_price)
            except:
                pass

        if area_id and area_id.isdigit():
            prices = prices.filter(customer__area_id=int(area_id))

        # 计算各状态数量
        all_count = CustomerPrice.all_objects.count()
        active_count = CustomerPrice.objects.count()
        disabled_count = all_count - active_count

        paginator = Paginator(prices, page_size)
        try:
            price_page = paginator.page(page)
        except PageNotAnInteger:
            price_page = paginator.page(1)
        except EmptyPage:
            price_page = paginator.page(paginator.num_pages)

        result = []
        for cp in price_page:
            result.append({
                'id': cp.id,
                'customer_id': cp.customer.id,
                'customer_name': cp.customer.name,
                'customer_area_id': cp.customer.area.id if cp.customer.area else '',
                'customer_area_name': cp.customer.area.name if cp.customer.area else '',
                'product_id': cp.product.id,
                'product_name': cp.product.name,
                'product_aliases': [alias.alias_name for alias in cp.product.aliases.all()],
                'custom_price': float(cp.custom_price),
                'standard_price': float(cp.product.price),
                'remark': cp.remark or '',
                'is_active': cp.is_active,
            })

        response_data = {
            'code': 1, 'msg': '查询成功', 'data': result, 'keyword': keyword,
            'page': int(page), 'total': paginator.count, 'total_pages': paginator.num_pages,
            'has_next': price_page.has_next(), 'has_previous': price_page.has_previous(),
            'counts': {
                'all': all_count,
                'active': active_count,
                'disabled': disabled_count
            }
        }

        cache.set(cache_key, response_data, CACHE_HIGH_PRIORITY)
        logger.info(f"设置专属价格缓存: {cache_key}")
        return JsonResponse(response_data, safe=False, content_type='application/json')

    except Exception as e:
        logger.error(f"查询专属价列表失败: {str(e)}", exc_info=True)
        return JsonResponse(
            {'code': 0, 'msg': f'查询失败：{str(e)}', 'data': []},
            safe=False, content_type='application/json'
        )


# ========== 启用客户专属价 ==========
@login_required
@permission_required('customer_price_delete')
def customer_price_enable(request, pk):
    """启用客户专属价格接口"""
    try:
        cp = get_object_or_404(CustomerPrice.all_objects, pk=pk)
        customer_name = cp.customer.name
        product_name = cp.product.name
        custom_price = float(cp.custom_price)
        product_standard_price = float(cp.product.price)
        remark = cp.remark if cp.remark else '无'

        cp.is_active = True
        cp.disabled_time = None
        cp.save()

        create_operation_log(
            request=request,
            op_type='enable',
            obj_type='customer_price',
            obj_id=pk,
            obj_name=f"{customer_name}-{product_name}",
            detail=f"启用客户专属价：ID={pk}，客户={customer_name}，商品={product_name}，标准价={product_standard_price}元，专属价={custom_price}元，备注={remark}"
        )
        clear_customer_price_cache()
        return JsonResponse({'code': 1, 'msg': '启用专属价成功'}, content_type='application/json')
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': f'启用失败：{str(e)}'}, content_type='application/json')


# ========== 新增客户价格 ==========
@login_required
@permission_required('customer_price_add')
def customer_price_add(request):
    """新增客户专属价格"""
    if request.method == 'POST':
        try:
            customer_id = request.POST.get('customer_id', '').strip()
            product_id = request.POST.get('product_id', '').strip()
            custom_price = request.POST.get('custom_price', '').strip()
            remark = request.POST.get('remark', '').strip()

            if not customer_id or not product_id or not custom_price:
                return JsonResponse({'code': 0, 'msg': '客户、商品、专属价不能为空'}, content_type='application/json')

            try:
                custom_price = float(custom_price)
                if custom_price < 0:
                    return JsonResponse({'code': 0, 'msg': '专属价不能为负数'}, content_type='application/json')
            except:
                return JsonResponse({'code': 0, 'msg': '专属价必须是数字'}, content_type='application/json')

            customer = get_object_or_404(Customer, id=customer_id)
            product = get_object_or_404(Product, id=product_id)
            product_standard_price = float(product.price)

            if CustomerPrice.objects.filter(customer=customer, product=product).exists():
                return JsonResponse({'code': 0, 'msg': '该客户已设置过此商品的专属价'}, content_type='application/json')

            cp = CustomerPrice.objects.create(
                customer=customer,
                product=product,
                custom_price=custom_price,
                remark=remark
            )

            create_operation_log(
                request=request,
                op_type='create',
                obj_type='customer_price',
                obj_id=cp.id,
                obj_name=f"{customer.name}-{product.name}",
                detail=f"新增客户专属价：客户={customer.name}，商品={product.name}，标准价={product_standard_price}元，专属价={custom_price}元，备注={remark if remark else '无'}"
            )
            clear_customer_price_cache()
            return JsonResponse({'code': 1, 'msg': '新增专属价成功'}, content_type='application/json')
        except Exception as e:
            return JsonResponse({'code': 0, 'msg': f'新增失败：{str(e)}'}, content_type='application/json')
    return JsonResponse({'code': 0, 'msg': '仅支持POST请求'}, content_type='application/json')


# ========== 编辑客户价格 ==========
@login_required
@permission_required('customer_price_edit')
def customer_price_edit(request, pk):
    """编辑客户专属价格"""
    try:
        cp = get_object_or_404(CustomerPrice, pk=pk)
        if request.method == 'POST':
            custom_price = request.POST.get('custom_price', '').strip()
            remark = request.POST.get('remark', '').strip()

            if not custom_price:
                return JsonResponse({'code': 0, 'msg': '专属价不能为空'}, content_type='application/json')
            try:
                custom_price = float(custom_price)
                if custom_price < 0:
                    return JsonResponse({'code': 0, 'msg': '专属价不能为负数'}, content_type='application/json')
            except:
                return JsonResponse({'code': 0, 'msg': '专属价必须是数字'}, content_type='application/json')

            old_price = float(cp.custom_price)
            old_remark = cp.remark if cp.remark else '无'
            customer_name = cp.customer.name
            product_name = cp.product.name
            product_standard_price = float(cp.product.price)

            cp.custom_price = custom_price
            cp.remark = remark
            cp.save()

            create_operation_log(
                request=request,
                op_type='update',
                obj_type='customer_price',
                obj_id=cp.id,
                obj_name=f"{customer_name}-{product_name}",
                detail=f"编辑客户专属价：客户={customer_name}，商品={product_name}，标准价={product_standard_price}元，原专属价={old_price}元→新专属价={custom_price}元，原备注={old_remark}→新备注={remark if remark else '无'}"
            )
            clear_customer_price_cache()
            return JsonResponse({'code': 1, 'msg': '编辑专属价成功'}, content_type='application/json')
        return JsonResponse({'code': 0, 'msg': '仅支持POST请求'}, content_type='application/json')
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': f'编辑失败：{str(e)}'}, content_type='application/json')


# ========== 禁用专属价（软删除） ==========
@login_required
@permission_required('customer_price_delete')
def customer_price_delete(request, pk):
    """禁用客户专属价格（软删除）"""
    try:
        cp = get_object_or_404(CustomerPrice, pk=pk)
        customer_name = cp.customer.name
        product_name = cp.product.name
        custom_price = float(cp.custom_price)
        product_standard_price = float(cp.product.price)
        remark = cp.remark if cp.remark else '无'

        cp.is_active = False
        cp.disabled_time = timezone.now()
        cp.save()

        create_operation_log(
            request=request,
            op_type='disable',
            obj_type='customer_price',
            obj_id=pk,
            obj_name=f"{customer_name}-{product_name}",
            detail=f"禁用客户专属价：ID={pk}，客户={customer_name}，商品={product_name}，标准价={product_standard_price}元，专属价={custom_price}元，备注={remark}"
        )
        clear_customer_price_cache()
        return JsonResponse({'code': 1, 'msg': '禁用专属价成功'}, content_type='application/json')
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': f'禁用失败：{str(e)}'}, content_type='application/json')


@login_required
@permission_required('customer_price_view')
def customer_price_page(request):
    """客户专属价格管理页面"""
    return render(request, 'customer_manage/customer_price.html', {
        'can_add_price': request.user.has_permission('customer_price_add'),
        'can_edit_price': request.user.has_permission('customer_price_edit'),
        'can_delete_price': request.user.has_permission('customer_price_delete'),
    })


# ========== 价格页辅助接口 ==========
@login_required
@permission_required('customer_price_view')
def product_list_for_price(request):
    """供客户价格管理页面获取商品列表 - 手动缓存版"""
    try:
        cache_key = CACHE_PREFIX_PRODUCT_LIST_FOR_PRICE
        cached_data = cache.get(cache_key)
        if cached_data:
            return JsonResponse(cached_data, safe=False, content_type='application/json')

        products = Product.objects.all().order_by('name')
        result = [{'id': p.id, 'name': p.name, 'price': float(p.price)} for p in products]
        cache.set(cache_key, result, CACHE_MID_PRIORITY)

        return JsonResponse(result, safe=False, content_type='application/json')
    except Exception as e:
        return JsonResponse(
            {'code': 0, 'msg': f'查询商品失败：{str(e)}'},
            safe=False,
            content_type='application/json'
        )


@login_required
@permission_required('customer_price_view')
def search_customer_for_price(request):
    """客户搜索：匹配名称/区域/电话 - 手动缓存版"""
    keyword = request.GET.get('keyword', '').strip()
    if not keyword:
        return JsonResponse({'code': 0, 'data': []})

    cache_key = f"{CACHE_PREFIX_SEARCH_CUSTOMER_FOR_PRICE}{keyword}"
    cached_data = cache.get(cache_key)
    if cached_data:
        return JsonResponse(cached_data, safe=False, content_type='application/json')

    # ✅ 修改：支持按电话搜索，加去重
    customer_matches = Customer.objects.select_related('area').prefetch_related('phones').filter(
        Q(name__icontains=keyword) |
        Q(pinyin_full__icontains=keyword) |  # 新增
        Q(pinyin_abbr__icontains=keyword) |  # 新增
        Q(area__name__icontains=keyword) |
        Q(phones__phone__icontains=keyword)
    ).distinct()[:8]

    data = []
    for customer in customer_matches:
        area_name = customer.area.name if customer.area else '无区域'
        full_name = f"{area_name} | {customer.name}"
        data.append({
            'id': customer.id,
            'name': customer.name,
            'area_name': area_name,
            'full_name': full_name
        })

    cache.set(cache_key, {'code': 1, 'data': data}, CACHE_MID_PRIORITY)
    return JsonResponse({'code': 1, 'data': data}, content_type='application/json')

@login_required
@permission_required('customer_price_view')
def search_product_for_price(request):
    """商品搜索：匹配名称/拼音/别名 - 手动缓存版"""
    keyword = request.GET.get('keyword', '').strip()
    if not keyword:
        return JsonResponse({'code': 0, 'data': []})

    cache_key = f"{CACHE_PREFIX_SEARCH_PRODUCT_FOR_PRICE}{keyword}"
    cached_data = cache.get(cache_key)
    if cached_data:
        return JsonResponse(cached_data, safe=False, content_type='application/json')

    all_products = Product.objects.filter(
        Q(name__icontains=keyword) |
        Q(pinyin_full__icontains=keyword) |
        Q(pinyin_abbr__icontains=keyword) |
        Q(id__in=ProductAlias.objects.filter(
            Q(alias_name__icontains=keyword) |
            Q(alias_pinyin_full__icontains=keyword) |
            Q(alias_pinyin_abbr__icontains=keyword)
        ).values('product_id'))
    ).distinct()[:8]

    data = []
    for product in all_products:
        data.append({
            'id': product.id,
            'name': product.name,
            'price': float(product.price),
            'unit': product.unit
        })

    cache.set(cache_key, {'code': 1, 'data': data}, CACHE_MID_PRIORITY)
    return JsonResponse({'code': 1, 'data': data}, content_type='application/json')


@login_required
@permission_required('customer_price_view')
def area_list_for_price(request):
    """供专属价高级筛选获取区域列表 - 手动缓存版（共用）"""
    try:
        cached_data = cache.get(CACHE_KEY_AREA_LIST)
        if cached_data:
            return JsonResponse({'code': 1, 'data': cached_data}, content_type='application/json')

        areas = Area.objects.all().order_by('name')
        result = [{'id': a.id, 'name': a.name} for a in areas]
        cache.set(CACHE_KEY_AREA_LIST, result, CACHE_MID_PRIORITY)

        return JsonResponse({'code': 1, 'data': result}, content_type='application/json')
    except Exception as e:
        return JsonResponse(
            {'code': 0, 'msg': f'查询区域失败：{str(e)}', 'data': []},
            content_type='application/json'
        )


# ========== 客户消费TOP30 ==========
@login_required
@permission_required('customer_sales_rank')
def customer_sales_rank_page(request):
    """客户消费TOP30排行页面"""
    areas = Area.objects.all().order_by('name')
    return render(request, 'customer_manage/customer_sales_rank.html', {
        'areas': areas,
        'is_super_admin': request.user.role and request.user.role.code == ROLE_SUPER_ADMIN
    })


@login_required
@permission_required('customer_sales_rank')
def customer_sales_rank_data(request):
    """获取客户消费TOP30数据 - 手动缓存版"""
    try:
        area_id = request.GET.get('area_id', '').strip()
        time_range = request.GET.get('time_range', 'year').strip()

        cache_key = f"{CACHE_PREFIX_CUSTOMER_SALES_RANK}{area_id}_{time_range}"
        cached_data = cache.get(cache_key)
        if cached_data:
            return JsonResponse(cached_data, safe=False, content_type='application/json')

        base_orders = Order.objects.filter(
            status__in=['pending', 'printed', 'reopened'],
            is_settled__in=[True, False],
            customer__isnull=False
        )

        # ✅ 修复：使用上海本地日期，避免UTC日期偏差1天
        today = timezone.localdate()
        today_start = timezone.make_aware(datetime.datetime.combine(today, datetime.time.min))
        today_end = timezone.make_aware(datetime.datetime.combine(today, datetime.time.max))

        if time_range == 'today':
            base_orders = base_orders.filter(create_time__gte=today_start, create_time__lte=today_end)
        elif time_range == 'week':
            week_start = today - datetime.timedelta(days=today.weekday())
            week_end = week_start + datetime.timedelta(days=6)
            week_start_dt = timezone.make_aware(datetime.datetime.combine(week_start, datetime.time.min))
            week_end_dt = timezone.make_aware(datetime.datetime.combine(week_end, datetime.time.max))
            base_orders = base_orders.filter(create_time__gte=week_start_dt, create_time__lte=week_end_dt)
        elif time_range == 'month':
            month_start = datetime.date(today.year, today.month, 1)
            if today.month == 12:
                month_end = datetime.date(today.year, 12, 31)
            else:
                month_end = datetime.date(today.year, today.month + 1, 1) - datetime.timedelta(days=1)
            month_start_dt = timezone.make_aware(datetime.datetime.combine(month_start, datetime.time.min))
            month_end_dt = timezone.make_aware(datetime.datetime.combine(month_end, datetime.time.max))
            base_orders = base_orders.filter(create_time__gte=month_start_dt, create_time__lte=month_end_dt)

        if area_id and area_id.isdigit():
            base_orders = base_orders.filter(customer__area_id=int(area_id))

        customer_sales = base_orders.values(
            'customer__id', 'customer__name', 'customer__area__name'
        ).annotate(
            total_amount=Sum('total_amount')
        ).order_by('-total_amount')[:30]

        if customer_sales:
            customer_ids = [item['customer__id'] for item in customer_sales]
            unpaid_data = Order.objects.filter(
                customer_id__in=customer_ids,
                status__in=['pending', 'printed', 'reopened'],
                is_settled=False
            ).values('customer_id').annotate(total=Sum('total_amount'))
            unpaid_dict = {item['customer_id']: float(item['total'] or 0) for item in unpaid_data}

            paid_data = RepaymentRecord.objects.filter(
                customer_id__in=customer_ids
            ).values('customer_id').annotate(total=Sum('repayment_amount'))
            paid_dict = {item['customer_id']: float(item['total'] or 0) for item in paid_data}
        else:
            unpaid_dict = {}
            paid_dict = {}

        result = []
        for idx, item in enumerate(customer_sales, 1):
            customer_id = item['customer__id']
            unpaid_amount = unpaid_dict.get(customer_id, 0)
            paid_amount = paid_dict.get(customer_id, 0)
            total_debt = max(unpaid_amount - paid_amount, 0)

            result.append({
                'rank': idx,
                'customer_id': customer_id,
                'customer_name': item['customer__name'],
                'area_name': item['customer__area__name'] or '无区域',
                'total_amount': float(item['total_amount'] or 0.0),
                'total_debt': total_debt
            })

        response_data = {'code': 1, 'msg': '查询成功', 'data': result}
        cache.set(cache_key, response_data, CACHE_HIGH_PRIORITY)
        return JsonResponse(response_data, content_type='application/json')

    except Exception as e:
        return JsonResponse({
            'code': 0, 'msg': f'查询失败：{str(e)}', 'data': []
        }, content_type='application/json')


# ========== Excel导出通用函数 ==========
def export_to_excel(data, title, headers, selected_fields, custom_fields, file_name, total_row=None):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = title

    final_fields = selected_fields.copy()
    final_headers = {field: headers[field] for field in selected_fields}

    if custom_fields:
        for cf in custom_fields:
            cf_name = cf.get('name', '')
            cf_position = cf.get('position', 'after')
            cf_target = cf.get('target', '')
            if not cf_name or not cf_target: continue
            custom_field_key = f'custom_{cf_name.replace(" ", "_")}_{len(final_fields)}'
            final_headers[custom_field_key] = cf_name
            try:
                target_index = final_fields.index(cf_target)
                insert_index = target_index + 1 if cf_position == 'after' else target_index
                final_fields.insert(insert_index, custom_field_key)
            except ValueError:
                final_fields.append(custom_field_key)

    selected_headers = [final_headers[field] for field in final_fields]
    title_font = Font(bold=True, size=12)
    alignment = Alignment(horizontal='center')

    for col, header in enumerate(selected_headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = title_font
        cell.alignment = alignment

    for row, item in enumerate(data, 2):
        for col, field in enumerate(final_fields, 1):
            value = item.get(field, '') if not field.startswith('custom_') else ''
            if isinstance(value, float): value = round(value, 2)
            ws.cell(row=row, column=col, value=value)

    if total_row:
        total_row_num = len(data) + 2
        total_font = Font(bold=True, color="FFFFFF")
        total_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.cell(row=total_row_num, column=1, value="总计").font = total_font
        ws.cell(row=total_row_num, column=1).fill = total_fill
        for col, field in enumerate(final_fields, 1):
            if field in total_row:
                cell = ws.cell(row=total_row_num, column=col, value=round(total_row[field], 2))
                cell.font = total_font
                cell.fill = total_fill
                cell.alignment = Alignment(horizontal='center')

    for col in range(1, len(selected_headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{file_name}.xlsx"'
    return response

# ========== 客户导出 ==========
@login_required
def customer_export(request):
    if request.method == 'POST':
        try:
            data = request.POST
            selected_fields = data.getlist('fields[]')
            custom_fields = json.loads(data.get('custom_fields', '[]'))

            if not selected_fields:
                return JsonResponse({'code': 0, 'msg': '请至少选择一个导出字段'})

            headers = {
                'serial': '序号',
                'id': 'ID',
                'name': '客户名称',
                'area_name': '所属区域',
                'phone': '联系电话',
                'remark': '备注'
            }

            # ✅ 修改：预加载电话
            customers = Customer.objects.select_related('area').prefetch_related('phones').order_by('-create_time')
            export_data = []
            for idx, customer in enumerate(customers, 1):
                export_data.append({
                    'serial': idx,
                    'id': customer.id,
                    'name': customer.name,
                    'area_name': customer.area.name if customer.area else '无',
                    # ✅ 修改：使用主号码
                    'phone': customer.primary_phone,
                    'remark': customer.remark or ''
                })

            file_date_str = timezone.localdate().strftime("%Y%m%d")
            return export_to_excel(
                data=export_data,
                title='客户列表',
                headers=headers,
                selected_fields=selected_fields,
                custom_fields=custom_fields,
                file_name=f'{file_date_str}客户管理导出',
                total_row=None
            )
        except Exception as e:
            logger.error(f"导出客户失败：{str(e)}", exc_info=True)
            return JsonResponse({'code': 0, 'msg': f'导出失败：{str(e)}'}, status=500)
    return JsonResponse({'code': 0, 'msg': '请求方式错误'})

@login_required
def customer_import(request):
    if request.method == 'POST':
        try:
            file_obj = request.FILES.get('file')
            if not file_obj:
                return JsonResponse({'code': 0, 'msg': '请上传文件'})

            wb = load_workbook(file_obj)
            ws = wb.active
            new_count = 0
            skip_count = 0
            error_list = []
            area_map = {area.name: area for area in Area.objects.all()}

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not any(row):
                    continue

                name = ''
                area_name = ''
                phone = ''
                remark = ''
                cells = [str(cell).strip() if cell else '' for cell in row]

                if len(cells) >= 4:
                    name = cells[1]
                    area_name = cells[2]
                    phone = cells[3]
                    if len(cells) > 4:
                        remark = cells[4]

                # ✅ 修改：仅校验名称必填，电话可为空
                if not name:
                    error_list.append(f"第{row_idx}行：客户名称为空，跳过")
                    continue

                if Customer.objects.filter(name=name).exists():
                    skip_count += 1
                    continue

                area_obj = None
                if area_name and area_name in area_map:
                    area_obj = area_map[area_name]

                try:
                    customer = Customer.objects.create(
                        name=name,
                        area=area_obj,
                        remark=remark
                    )
                    # ✅ 修改：仅电话非空时创建主号码
                    if phone:
                        CustomerPhone.objects.create(
                            customer=customer,
                            phone=phone.strip(),
                            is_primary=True
                        )
                    new_count += 1
                except Exception as e:
                    error_list.append(f"第{row_idx}行：保存失败（{str(e)}）")

            msg = f"导入完成！新增：{new_count} 条，跳过重复：{skip_count} 条。"
            if error_list:
                msg += f" 异常：{len(error_list)} 条。"

            if new_count > 0:
                clear_customer_cache()

            return JsonResponse({'code': 1, 'msg': msg})

        except Exception as e:
            logger.error(f"导入客户失败：{str(e)}", exc_info=True)
            return JsonResponse({'code': 0, 'msg': f'导入失败：{str(e)}'})
    return JsonResponse({'code': 0, 'msg': '请求方式错误'})

# ========== 客户专属价格导出/导入 ==========
@login_required
def customer_price_export(request):
    if request.method == 'POST':
        try:
            data = request.POST
            selected_fields = data.getlist('fields[]')
            custom_fields = json.loads(data.get('custom_fields', '[]'))

            if not selected_fields:
                return JsonResponse({'code': 0, 'msg': '请至少选择一个导出字段'})

            headers = {
                'serial': '序号',
                'id': 'ID',
                'customer_name': '客户名称',
                'customer_area_name': '所属区域',
                'product_name': '商品名称',
                'standard_price': '商品标准价',
                'custom_price': '客户专属价',
                'remark': '备注'
            }

            prices = CustomerPrice.objects.select_related(
                'customer', 'customer__area', 'product'
            ).order_by('-create_time')

            export_data = []
            for idx, cp in enumerate(prices, 1):
                export_data.append({
                    'serial': idx,
                    'id': cp.id,
                    'customer_name': cp.customer.name if cp.customer else '未知',
                    'customer_area_name': cp.customer.area.name if (cp.customer and cp.customer.area) else '无',
                    'product_name': cp.product.name if cp.product else '未知',
                    'standard_price': float(cp.product.price) if cp.product else 0.0,
                    'custom_price': float(cp.custom_price),
                    'remark': cp.remark or ''
                })

            file_date_str = timezone.localdate().strftime("%Y%m%d")
            return export_to_excel(
                data=export_data,
                title='客户专属价格',
                headers=headers,
                selected_fields=selected_fields,
                custom_fields=custom_fields,
                file_name=f'{file_date_str}客户专属价格导出',
                total_row=None
            )
        except Exception as e:
            logger.error(f"导出客户专属价格失败：{str(e)}", exc_info=True)
            return JsonResponse({'code': 0, 'msg': f'导出失败：{str(e)}'}, status=500)
    return JsonResponse({'code': 0, 'msg': '请求方式错误'})

@login_required
def customer_price_import(request):
    if request.method == 'POST':
        try:
            file_obj = request.FILES.get('file')
            if not file_obj:
                return JsonResponse({'code': 0, 'msg': '请上传文件'})

            wb = load_workbook(file_obj, data_only=True)
            ws = wb.active
            new_count = 0
            skip_count = 0
            error_list = []

            customer_map = {c.name: c for c in Customer.objects.all()}
            product_map = {p.name: p for p in Product.objects.all()}
            existing_price_keys = set(
                CustomerPrice.objects.values_list('customer_id', 'product_id')
            )

            logger.info(f"开始导入专属价格，预加载客户数: {len(customer_map)}, 商品数: {len(product_map)}")

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not any(row):
                    continue

                cells = [str(cell).strip() if cell is not None else '' for cell in row]
                while len(cells) < 7:
                    cells.append('')

                customer_name = cells[1]
                product_name = cells[3]
                custom_price_str = cells[5]
                remark = cells[6]

                if not customer_name:
                    error_list.append(f"第{row_idx}行：客户名称为空，跳过")
                    continue
                if not product_name:
                    error_list.append(f"第{row_idx}行：商品名称为空，跳过")
                    continue
                if not custom_price_str:
                    error_list.append(f"第{row_idx}行：专属价格为空，跳过")
                    continue

                try:
                    price_clean = custom_price_str.replace('¥', '').replace(',', '').strip()
                    custom_price = float(price_clean)
                    if custom_price < 0:
                        raise ValueError("价格为负")
                except Exception as e:
                    error_list.append(f"第{row_idx}行：专属价格格式错误 ({custom_price_str})")
                    continue

                if customer_name not in customer_map:
                    error_list.append(f"第{row_idx}行：客户【{customer_name}】在系统中不存在，跳过")
                    continue
                if product_name not in product_map:
                    error_list.append(f"第{row_idx}行：商品【{product_name}】在系统中不存在，跳过")
                    continue

                customer = customer_map[customer_name]
                product = product_map[product_name]

                if (customer.id, product.id) in existing_price_keys:
                    skip_count += 1
                    continue

                try:
                    CustomerPrice.objects.create(
                        customer=customer,
                        product=product,
                        custom_price=custom_price,
                        remark=remark
                    )
                    new_count += 1
                    existing_price_keys.add((customer.id, product.id))
                except Exception as e:
                    error_list.append(f"第{row_idx}行：数据库保存失败（{str(e)}）")

            msg = f"导入完成！新增：{new_count} 条，跳过重复/错误：{skip_count + len(error_list)} 条。"
            if error_list:
                msg += f" (前10个错误: {'; '.join(error_list[:10])})"

            logger.info(msg)
            if new_count > 0:
                clear_customer_price_cache()
                logger.info("已触发专属价格缓存清理")

            return JsonResponse({'code': 1, 'msg': msg})

        except Exception as e:
            logger.error(f"导入客户专属价格系统异常：{str(e)}", exc_info=True)
            return JsonResponse({'code': 0, 'msg': f'导入系统异常：{str(e)}'})
    return JsonResponse({'code': 0, 'msg': '请求方式错误'})

# ==================== 客户统计功能 ====================
@login_required
@permission_required('customer_view')
def customer_stats_page(request):
    """客户统计详情页面"""
    areas = Area.objects.all().order_by('name')
    return render(request, 'customer_manage/customer_stats.html', {
        'areas': areas
    })


@login_required
@permission_required('customer_view')
def calculate_customer_stats(request):
    """
    客户统计接口
    默认：最近30天
    支持：时间筛选/地区筛选/状态筛选
    """
    try:
        # 筛选参数
        time_range = request.GET.get('time_range', '30days')
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')
        area_id = request.GET.get('area_id', '')
        status = request.GET.get('status', 'all')
        page = request.GET.get('page', 1)
        page_size = 20

        # 基础查询：有效订单 + 有效客户
        base_orders = Order.objects.filter(
            status__in=['pending', 'printed', 'reopened'],
            customer__isnull=False
        )
        # ✅ 修改：预加载电话
        base_customers = Customer.all_objects.all().select_related('area').prefetch_related('phones')

        # 1. 时间筛选
        today = timezone.localdate()
        if time_range == '30days':
            thirty_days_ago = today - datetime.timedelta(days=30)
            base_orders = base_orders.filter(create_time__date__gte=thirty_days_ago)
        elif time_range == 'today':
            base_orders = base_orders.filter(create_time__date=today)
        elif time_range == 'month':
            base_orders = base_orders.filter(create_time__year=today.year, create_time__month=today.month)
        elif time_range == 'year':
            base_orders = base_orders.filter(create_time__year=today.year)
        elif time_range == 'custom' and start_date and end_date:
            base_orders = base_orders.filter(create_time__date__gte=start_date, create_time__date__lte=end_date)

        # 2. 地区筛选
        if area_id and area_id.isdigit():
            base_customers = base_customers.filter(area_id=area_id)
            base_orders = base_orders.filter(area_id=area_id)

        # 3. 客户状态筛选
        if status == 'active':
            base_customers = base_customers.filter(is_active=True)
        elif status == 'disabled':
            base_customers = base_customers.filter(is_active=False)

        # 数据库层面分页，仅处理当前页客户
        paginator = Paginator(base_customers, page_size)
        try:
            customer_page = paginator.page(page)
        except PageNotAnInteger:
            customer_page = paginator.page(1)
        except EmptyPage:
            customer_page = paginator.page(paginator.num_pages)

        # 提取当前页客户ID
        current_page_customer_ids = [c.id for c in customer_page]

        # 仅查询当前页客户的订单统计
        order_stats = base_orders.filter(customer_id__in=current_page_customer_ids).values('customer_id').annotate(
            total_consume=Sum('total_amount'),
            total_order=Count('id'),
            finished_order=Count('id', filter=Q(is_settled=True)),
            unsettled_order=Count('id', filter=Q(is_settled=False)),
            last_consume_time=Max('create_time')
        )
        order_dict = {item['customer_id']: item for item in order_stats}

        # 仅查询当前页客户的还款统计
        repay_stats = RepaymentRecord.objects.filter(customer_id__in=current_page_customer_ids).values(
            'customer_id').annotate(
            total_repay=Sum('repayment_amount')
        )
        repay_dict = {item['customer_id']: item['total_repay'] or Decimal('0') for item in repay_stats}

        # 组装当前页客户数据
        customer_list = []
        for customer in customer_page:
            stats = order_dict.get(customer.id, {
                'total_consume': Decimal('0'),
                'total_order': 0,
                'finished_order': 0,
                'unsettled_order': 0,
                'last_consume_time': None
            })
            total_repay = repay_dict.get(customer.id, Decimal('0'))

            total_debt = max(stats['total_consume'] - total_repay, Decimal('0'))

            customer_list.append({
                'id': customer.id,
                'name': customer.name,
                'area_name': customer.area.name if customer.area else '无区域',
                # ✅ 修改：使用主号码
                'phone': customer.primary_phone,
                'is_active': customer.is_active,
                'total_consume': float(stats['total_consume']),
                'total_debt': float(total_debt),
                'total_order': stats.get('total_order', 0),
                'finished_order': stats.get('finished_order', 0),
                'unsettled_order': stats.get('unsettled_order', 0),
                'last_consume_time': format_datetime(stats.get('last_consume_time'), '%Y-%m-%d %H:%M') if stats.get(
                    'last_consume_time') else '无消费'
            })

        # 全局统计聚合
        global_total_consume = base_orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
        global_total_repay = RepaymentRecord.objects.filter(
            customer_id__in=base_customers.values_list('id', flat=True)
        ).aggregate(total=Sum('repayment_amount'))['total'] or Decimal('0')
        global_total_debt = max(global_total_consume - global_total_repay, Decimal('0'))
        global_total_order = base_orders.count()

        return JsonResponse({
            'code': 1,
            'global_stats': {
                'total_consume': round(float(global_total_consume), 2),
                'total_debt': round(float(global_total_debt), 2),
                'total_order': global_total_order
            },
            'customers': customer_list,
            'page': customer_page.number,
            'total_pages': paginator.num_pages,
            'total_count': paginator.count
        })

    except Exception as e:
        logger.error(f"客户统计失败：{str(e)}", exc_info=True)
        return JsonResponse({'code': 0, 'msg': f'统计失败：{str(e)}'})
